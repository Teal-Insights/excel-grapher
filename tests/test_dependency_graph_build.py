from __future__ import annotations

from pathlib import Path

import openpyxl

from excel_grapher import create_dependency_graph


def _fixture_path(name: str) -> Path:
    return Path(__file__).parent / "data" / name


def _create_fixture_workbook(path: Path) -> None:
    """
    Create a small workbook with a simple dependency chain:

    - Sheet1!A1 = 2           (leaf)
    - Sheet1!A2 = 3           (leaf)
    - Sheet1!A3 = =A1 + A2    (formula depends on A1, A2)
    - Sheet1!A4 = =A3 * 2     (formula depends on A3)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"].value = 2
    ws["A2"].value = 3
    ws["A3"].value = "=A1+A2"
    ws["A4"].value = "=A3*2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def test_create_dependency_graph_traces_dependencies(tmp_path: Path) -> None:
    excel_path = tmp_path / "simple_chain.xlsx"
    _create_fixture_workbook(excel_path)

    graph = create_dependency_graph(excel_path, ["Sheet1!A4"], load_values=False)

    assert "Sheet1!A4" in graph
    assert "Sheet1!A3" in graph
    assert "Sheet1!A2" in graph
    assert "Sheet1!A1" in graph

    assert graph.dependencies("Sheet1!A4") == {"Sheet1!A3"}
    assert graph.dependencies("Sheet1!A3") == {"Sheet1!A1", "Sheet1!A2"}
    assert graph.dependencies("Sheet1!A2") == set()
    assert graph.dependencies("Sheet1!A1") == set()


def test_evaluation_order_is_dependency_first(tmp_path: Path) -> None:
    excel_path = tmp_path / "simple_chain.xlsx"
    _create_fixture_workbook(excel_path)

    graph = create_dependency_graph(excel_path, ["Sheet1!A4"], load_values=False)
    order = graph.evaluation_order()

    assert order.index("Sheet1!A1") < order.index("Sheet1!A3")
    assert order.index("Sheet1!A2") < order.index("Sheet1!A3")
    assert order.index("Sheet1!A3") < order.index("Sheet1!A4")


def test_range_dependencies_are_expanded(tmp_path: Path) -> None:
    """
    Excel range references should expand to individual cell dependencies so we
    don't miss intermediate inputs inside SUM/MIN/MAX/etc.
    """
    excel_path = tmp_path / "range_chain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"].value = 1
    ws["A2"].value = 2
    ws["A3"].value = 3
    ws["A4"].value = "=SUM(A1:A3)"

    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A4"], load_values=False)
    assert graph.dependencies("Sheet1!A4") == {"Sheet1!A1", "Sheet1!A2", "Sheet1!A3"}


def test_cross_sheet_range_dependencies_are_expanded(tmp_path: Path) -> None:
    excel_path = tmp_path / "cross_sheet_range.xlsx"
    wb = openpyxl.Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s2 = wb.create_sheet("Sheet 2")

    s2["A1"].value = 10
    s2["A2"].value = 20
    s2["B1"].value = 30
    s2["B2"].value = 40

    s1["A1"].value = "x"
    s1["A2"].value = "=SUM('Sheet 2'!A1:B2)"

    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A2"], load_values=False)
    # Sheet names with spaces are quoted in keys to match Excel formula syntax
    assert graph.dependencies("Sheet1!A2") == {
        "'Sheet 2'!A1",
        "'Sheet 2'!A2",
        "'Sheet 2'!B1",
        "'Sheet 2'!B2",
    }


def test_named_range_is_resolved(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_range.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"].value = 41
    ws["A2"].value = "=MyInput+1"

    # Define name: MyInput -> Sheet1!$A$1
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names.add(DefinedName("MyInput", attr_text="Sheet1!$A$1"))

    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A2"], load_values=False)
    assert graph.dependencies("Sheet1!A2") == {"Sheet1!A1"}


def test_load_values_reads_cached_formula_results(tmp_path: Path) -> None:
    """
    When load_values=True, formula nodes should include cached computed values.

    We generate the workbook with XlsxWriter so cached results are embedded.
    """
    import xlsxwriter

    excel_path = tmp_path / "cached_values.xlsx"
    wb = xlsxwriter.Workbook(excel_path)
    ws = wb.add_worksheet("Sheet1")

    # A1=2, A2=3
    ws.write_number(0, 0, 2)
    ws.write_number(1, 0, 3)

    # A3 = A1+A2 (cached result 5)
    ws.write_formula(2, 0, "=A1+A2", None, 5)
    # A4 = A3*2 (cached result 10)
    ws.write_formula(3, 0, "=A3*2", None, 10)

    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A4"], load_values=True)

    n3 = graph.get_node("Sheet1!A3")
    n4 = graph.get_node("Sheet1!A4")
    assert n3 is not None and n4 is not None
    assert n3.value == 5
    assert n4.value == 10


def test_parse_target_handles_quoted_sheet_name(tmp_path: Path) -> None:
    """
    Target strings with quoted sheet names (e.g., 'Sheet Name'!A1) should be
    parsed correctly. Keys in the graph use quoted format to match Excel syntax.
    """
    excel_path = tmp_path / "quoted_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Sheet"

    ws["A1"].value = 42

    wb.save(excel_path)
    wb.close()

    # Target uses quoted format as Excel would show it
    graph = create_dependency_graph(excel_path, ["'My Sheet'!A1"], load_values=False)

    # Keys are quoted when sheet names contain spaces
    assert "'My Sheet'!A1" in graph
    node = graph.get_node("'My Sheet'!A1")
    assert node is not None
    assert node.value == 42
    # Node.sheet stores the unquoted name
    assert node.sheet == "My Sheet"

