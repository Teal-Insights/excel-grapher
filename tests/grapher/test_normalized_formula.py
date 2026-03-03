"""
Tests for Node.normalized_formula — the transpiler-friendly form of the formula.
"""

from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

from excel_grapher import create_dependency_graph


def test_normalized_formula_prefixes_same_sheet_refs(tmp_path: Path) -> None:
    """Same-sheet refs like A1 should become Sheet1!A1 in normalized form."""
    excel_path = tmp_path / "local_refs.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 1
    ws["A2"].value = 2
    ws["A3"].value = "=A1+A2"
    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A3"], load_values=False)
    node = graph.get_node("Sheet1!A3")
    assert node is not None
    assert node.formula == "=A1+A2"
    assert node.normalized_formula == "=Sheet1!A1+Sheet1!A2"


def test_normalized_formula_resolves_named_ranges(tmp_path: Path) -> None:
    """Named ranges should be replaced with their fully-qualified target."""
    excel_path = tmp_path / "named_range.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 10
    ws["A2"].value = "=MyInput*2"
    wb.defined_names.add(DefinedName("MyInput", attr_text="Sheet1!$A$1"))
    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A2"], load_values=False)
    node = graph.get_node("Sheet1!A2")
    assert node is not None
    assert node.formula == "=MyInput*2"
    assert node.normalized_formula == "=Sheet1!A1*2"


def test_normalized_formula_strips_absolute_markers(tmp_path: Path) -> None:
    """Absolute markers ($) should be stripped in normalized form."""
    excel_path = tmp_path / "absolute.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 5
    ws["A2"].value = "=$A$1+$A1+A$1"
    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A2"], load_values=False)
    node = graph.get_node("Sheet1!A2")
    assert node is not None
    # All three refs should normalize to Sheet1!A1
    assert node.normalized_formula == "=Sheet1!A1+Sheet1!A1+Sheet1!A1"


def test_normalized_formula_qualifies_range_endpoints(tmp_path: Path) -> None:
    """Range endpoints should be sheet-qualified in normalized form."""
    excel_path = tmp_path / "range.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 1
    ws["A2"].value = 2
    ws["A3"].value = 3
    ws["A4"].value = "=SUM(A1:A3)"
    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A4"], load_values=False)
    node = graph.get_node("Sheet1!A4")
    assert node is not None
    assert node.formula == "=SUM(A1:A3)"
    assert node.normalized_formula == "=SUM(Sheet1!A1:Sheet1!A3)"


def test_normalized_formula_preserves_cross_sheet_refs(tmp_path: Path) -> None:
    """Already-qualified refs should be preserved (just stripped of $)."""
    excel_path = tmp_path / "cross_sheet.xlsx"
    wb = openpyxl.Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s2 = wb.create_sheet("Other Sheet")
    s2["B5"].value = 100
    s1["A1"].value = "='Other Sheet'!$B$5+1"
    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A1"], load_values=False)
    node = graph.get_node("Sheet1!A1")
    assert node is not None
    # Cross-sheet ref should stay quoted but lose $ markers
    assert node.normalized_formula == "='Other Sheet'!B5+1"


def test_normalized_formula_is_none_for_leaves(tmp_path: Path) -> None:
    """Leaf nodes (constants) should have normalized_formula=None."""
    excel_path = tmp_path / "leaf.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].value = 42
    ws["A2"].value = "=A1"
    wb.save(excel_path)
    wb.close()

    graph = create_dependency_graph(excel_path, ["Sheet1!A2"], load_values=False)
    leaf = graph.get_node("Sheet1!A1")
    assert leaf is not None
    assert leaf.is_leaf
    assert leaf.formula is None
    assert leaf.normalized_formula is None
