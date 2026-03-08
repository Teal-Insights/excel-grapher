"""Tests for named range resolution and strict validation."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from excel_grapher import create_dependency_graph
from excel_grapher.grapher.resolver import build_named_range_map


def _new_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb


def test_named_range_map_allows_simple_range(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_ranges_ok.xlsx"
    wb = _new_workbook()
    wb.defined_names.add(DefinedName("Foo", attr_text="Sheet1!$A$1"))
    wb.defined_names.add(DefinedName("Bar", attr_text="Sheet1!$A$1:$B$2"))
    wb.save(excel_path)

    maps = build_named_range_map(openpyxl.load_workbook(excel_path, data_only=False))
    assert maps.cell_map["Foo"] == ("Sheet1", "A1")
    assert maps.range_map["Bar"] == ("Sheet1", "A1", "B2")


def test_named_range_range_is_expanded(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_ranges_range_dep.xlsx"
    wb = _new_workbook()
    ws = wb["Sheet1"]
    ws["A1"].value = 1
    ws["B1"].value = 2
    ws["C1"].value = "=SUM(Range1)"
    wb.defined_names.add(DefinedName("Range1", attr_text="Sheet1!$A$1:$B$1"))
    wb.save(excel_path)

    graph = create_dependency_graph(excel_path, ["Sheet1!C1"], load_values=False)
    deps = graph.dependencies("Sheet1!C1")
    assert deps == {"Sheet1!A1", "Sheet1!B1"}


def test_named_range_map_raises_on_multi_area(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_ranges_multi.xlsx"
    wb = _new_workbook()
    wb.defined_names.add(DefinedName("Multi", attr_text="Sheet1!$A$1,Sheet1!$B$1"))
    ws = wb["Sheet1"]
    ws["C1"].value = "=SUM(Multi)"
    wb.save(excel_path)

    with pytest.raises(ValueError) as exc:
        create_dependency_graph(excel_path, ["Sheet1!C1"], load_values=False)
    assert "Multi" in str(exc.value)


def test_named_range_map_resolves_offset_counta_formula(tmp_path: Path) -> None:
    """Formula-based name OFFSET(Sheet!A1,0,0,COUNTA(Sheet!A:A),COUNTA(Sheet!1:1)) resolves to range."""
    excel_path = tmp_path / "offset_counta_name.xlsx"
    wb = _new_workbook()
    wb.create_sheet("Country_Information")
    ci = wb["Country_Information"]
    ci["A1"].value = 1
    ci["A2"].value = 2
    ci["B1"].value = 3
    ci["C1"].value = 4
    attr = "OFFSET(Country_Information!$A$1,0,0,COUNTA(Country_Information!$A:$A),COUNTA(Country_Information!$1:$1))"
    wb.defined_names.add(DefinedName("DSF__Country_Info", attr_text=attr))
    wb.save(excel_path)

    wb_loaded = openpyxl.load_workbook(excel_path, data_only=False)
    maps = build_named_range_map(wb_loaded)
    assert "DSF__Country_Info" in maps.range_map
    sheet, start, end = maps.range_map["DSF__Country_Info"]
    assert sheet == "Country_Information"
    assert start == "A1"
    assert end == "C2"


def test_named_range_map_resolves_offset_counta_fixed_width(tmp_path: Path) -> None:
    """OFFSET with COUNTA height and literal width (e.g. DSF__COMMODITY_TABLE) resolves when result is wider than sheet used range."""
    excel_path = tmp_path / "offset_counta_fixed_width.xlsx"
    wb = _new_workbook()
    wb.create_sheet("COM")
    com = wb["COM"]
    com["A1"].value = 1
    com["A2"].value = 2
    com["A3"].value = "header"
    attr = "OFFSET(COM!$A$3,0,0,COUNTA(COM!$A:$A),7)"
    wb.defined_names.add(DefinedName("DSF__COMMODITY_TABLE", attr_text=attr))
    wb.save(excel_path)

    wb_loaded = openpyxl.load_workbook(excel_path, data_only=False)
    maps = build_named_range_map(wb_loaded)
    assert "DSF__COMMODITY_TABLE" in maps.range_map
    sheet, start, end = maps.range_map["DSF__COMMODITY_TABLE"]
    assert sheet == "COM"
    assert start == "A3"
    assert end == "G5"


def test_dependency_graph_expands_formula_based_named_range(tmp_path: Path) -> None:
    """A formula that references an OFFSET/COUNTA defined name resolves without ValueError."""
    excel_path = tmp_path / "graph_offset_name.xlsx"
    wb = _new_workbook()
    wb.create_sheet("Country_Information")
    ci = wb["Country_Information"]
    ci["A1"].value = "a"
    ci["A2"].value = "b"
    ci["B1"].value = "c"
    ci["C1"].value = "d"
    wb.defined_names.add(
        DefinedName(
            "DSF__Country_Info",
            attr_text="OFFSET(Country_Information!$A$1,0,0,COUNTA(Country_Information!$A:$A),COUNTA(Country_Information!$1:$1))",
        )
    )
    ws = wb["Sheet1"]
    ws["D1"].value = "=COUNTA(DSF__Country_Info)"
    wb.save(excel_path)

    graph = create_dependency_graph(excel_path, ["Sheet1!D1"], load_values=False)
    deps = graph.dependencies("Sheet1!D1")
    assert "Country_Information!A1" in deps
    assert "Country_Information!B1" in deps
    assert "Country_Information!C1" in deps
    assert "Country_Information!A2" in deps

