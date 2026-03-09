"""Cell reading utilities for Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tests.utils._helpers import parse_cell_ref


def read_cell_value(
    workbook_path: Path,
    cell_ref: str,
    *,
    data_only: bool = True,
) -> Any:
    """Read a single cell value from an Excel workbook.

    Args:
        workbook_path: Path to the Excel workbook.
        cell_ref: Cell reference in format "'Sheet Name'!$A$1" or "Sheet!A1".
        data_only: If True, return cached values instead of formulas.

    Returns:
        The cell value (number, string, bool, datetime, or None).
    """
    sheet_name, cell_address = parse_cell_ref(cell_ref)

    wb = load_workbook(str(workbook_path), data_only=data_only, read_only=True)
    try:
        ws = wb[sheet_name]
        value = ws[cell_address].value
        return value
    finally:
        wb.close()


def read_range_values(
    workbook_path: Path,
    range_ref: str,
    *,
    data_only: bool = True,
) -> list[list[Any]]:
    """Read a range of cell values from an Excel workbook.

    Args:
        workbook_path: Path to the Excel workbook.
        range_ref: Range reference in format "Sheet!A1:B10".
        data_only: If True, return cached values instead of formulas.

    Returns:
        2D list of cell values.
    """
    from excel_grapher.evaluator.name_utils import parse_address

    if ":" in range_ref:
        sheet_name, range_part = parse_address(range_ref)
        start_cell, end_cell = range_part.replace("$", "").split(":")
        cell_range = f"{start_cell}:{end_cell}"
    else:
        sheet_name, cell_address = parse_cell_ref(range_ref)
        cell_range = cell_address

    wb = load_workbook(str(workbook_path), data_only=data_only, read_only=True)
    try:
        ws = wb[sheet_name]
        values = []
        for row in ws[cell_range]:
            if not isinstance(row, tuple):
                return [[row.value]]
            values.append([cell.value for cell in row])
        return values
    finally:
        wb.close()
