"""Tests for named range resolution and strict validation."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName

from excel_grapher import create_dependency_graph
from excel_grapher.resolver import build_named_range_map


def _new_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    return wb


def test_named_range_map_allows_simple_range(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_ranges_ok.xlsx"
    wb = _new_workbook()
    wb.defined_names.add(DefinedName("Foo", attr_text="Sheet1!$A$1"))
    wb.defined_names.add(DefinedName("Bar", attr_text="Sheet1!$A$1:$B$2"))
    wb.save(excel_path)

    maps = build_named_range_map(openpyxl.load_workbook(excel_path, data_only=False))
    assert maps.cell_map["Foo"] == ("Sheet1", "A1")
    assert maps.range_map["Bar"] == ("Sheet1", "A1", "B2")


def test_named_range_range_is_expanded(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_ranges_range_dep.xlsx"
    wb = _new_workbook()
    ws = wb["Sheet1"]
    ws["A1"].value = 1
    ws["B1"].value = 2
    ws["C1"].value = "=SUM(Range1)"
    wb.defined_names.add(DefinedName("Range1", attr_text="Sheet1!$A$1:$B$1"))
    wb.save(excel_path)

    graph = create_dependency_graph(excel_path, ["Sheet1!C1"], load_values=False)
    deps = graph.dependencies("Sheet1!C1")
    assert deps == {"Sheet1!A1", "Sheet1!B1"}


def test_named_range_map_raises_on_multi_area(tmp_path: Path) -> None:
    excel_path = tmp_path / "named_ranges_multi.xlsx"
    wb = _new_workbook()
    wb.defined_names.add(DefinedName("Multi", attr_text="Sheet1!$A$1,Sheet1!$B$1"))
    ws = wb["Sheet1"]
    ws["C1"].value = "=SUM(Multi)"
    wb.save(excel_path)

    with pytest.raises(ValueError) as exc:
        create_dependency_graph(excel_path, ["Sheet1!C1"], load_values=False)
    assert "Multi" in str(exc.value)

