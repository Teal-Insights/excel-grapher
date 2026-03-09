from __future__ import annotations

from pathlib import Path
from typing import Any, Literal, TypedDict, get_type_hints

import openpyxl

from example.map_lic_dsf_indicators import constrain_constant_range


class _TestConstraints(TypedDict, total=False):
    pass


def _build_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "lookup"
    ws["C4"] = "Afghanistan"
    ws["C5"] = "Benin"
    ws["C6"] = "Chad"
    wb.save(path)
    wb.close()


def test_constrain_constant_range_populates_annotations_and_data(tmp_path: Path) -> None:
    excel_path = tmp_path / "countries.xlsx"
    _build_workbook(excel_path)

    data: dict[str, Any] = {}

    constrain_constant_range(
        _TestConstraints,
        data,
        excel_path,
        sheet_name="lookup",
        range_a1="C4:C6",
    )

    hints = get_type_hints(_TestConstraints, include_extras=True)

    assert hints["lookup!C4"] == Literal["Afghanistan"]
    assert hints["lookup!C5"] == Literal["Benin"]
    assert hints["lookup!C6"] == Literal["Chad"]

    assert data == {
        "lookup!C4": "Afghanistan",
        "lookup!C5": "Benin",
        "lookup!C6": "Chad",
    }

