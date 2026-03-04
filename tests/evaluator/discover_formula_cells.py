from pathlib import Path

import openpyxl
import openpyxl.utils.cell


def discover_formula_cells_in_rows(
    wb_path: Path,
    sheet_name: str,
    rows: list[int],
) -> list[str]:
    """
    Scan specified rows and return sheet-qualified addresses for formula cells.

    Only includes cells that contain formulas (start with '=') and whose cached
    calculated value is numeric.
    """
    wb_formulas = openpyxl.load_workbook(wb_path, data_only=False, keep_vba=True)
    wb_values = openpyxl.load_workbook(wb_path, data_only=True, keep_vba=True)
    try:
        if (
            sheet_name not in wb_formulas.sheetnames
            or sheet_name not in wb_values.sheetnames
        ):
            print(f"  Warning: Sheet '{sheet_name}' not found")
            return []

        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]
        targets: list[str] = []

        for row in rows:
            # Scan all columns up to max_column
            max_col = ws_formulas.max_column or 1
            for col_idx in range(1, max_col + 1):
                cell_formula = ws_formulas.cell(row=row, column=col_idx)
                if isinstance(
                    cell_formula.value, str
                ) and cell_formula.value.startswith("="):
                    cached_value = ws_values.cell(row=row, column=col_idx).value
                    if not isinstance(cached_value, (int, float)) or isinstance(
                        cached_value, bool
                    ):
                        continue
                    col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
                    targets.append(f"{sheet_name}!{col_letter}{row}")

        return targets
    finally:
        wb_formulas.close()
        wb_values.close()