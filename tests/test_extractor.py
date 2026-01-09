"""Tests for the extractor module.

Tests cell key format normalization and CellDict functionality.
Keys should use quoted sheet names when the sheet name contains spaces,
hyphens, or other special characters that require quoting in Excel formulas.

Examples:
    - Simple sheet: Sheet1!A1 (unquoted)
    - Sheet with space: 'Sales Data'!A1 (quoted)
    - Sheet with hyphen: 'Q4-Results'!A1 (quoted)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import xlsxwriter

from excel_grapher import (
    CellDict,
    CellInfo,
    build_cell_dict,
    discover_formula_cells_in_rows,
    format_cell_key,
    graph_to_cell_dict,
    create_dependency_graph,
    needs_quoting,
)


class TestNeedsQuoting:
    """Tests for the needs_quoting helper function."""

    def test_simple_alphanumeric_no_quoting(self) -> None:
        """Simple alphanumeric sheet names don't need quoting."""
        assert needs_quoting("Sheet1") is False
        assert needs_quoting("Data") is False
        assert needs_quoting("MySheet123") is False

    def test_sheet_with_space_needs_quoting(self) -> None:
        """Sheet names with spaces need quoting."""
        assert needs_quoting("Sales Data") is True
        assert needs_quoting("Input Values") is True
        assert needs_quoting("Q4 Results") is True

    def test_sheet_with_hyphen_needs_quoting(self) -> None:
        """Sheet names with hyphens need quoting."""
        assert needs_quoting("Q4-Results") is True
        assert needs_quoting("2024-Budget") is True
        assert needs_quoting("Baseline - external") is True

    def test_sheet_with_underscore_no_quoting(self) -> None:
        """Sheet names with underscores don't need quoting."""
        assert needs_quoting("Sales_Data") is False
        assert needs_quoting("Q4_Results") is False


class TestFormatCellKey:
    """Tests for the format_cell_key helper function."""

    def test_simple_sheet_unquoted(self) -> None:
        """Simple sheet names should produce unquoted keys."""
        assert format_cell_key("Sheet1", "A", 1) == "Sheet1!A1"
        assert format_cell_key("Data", "B", 10) == "Data!B10"
        assert format_cell_key("MySheet", "AA", 100) == "MySheet!AA100"

    def test_sheet_with_space_quoted(self) -> None:
        """Sheet names with spaces should produce quoted keys."""
        assert format_cell_key("Sales Data", "A", 1) == "'Sales Data'!A1"
        assert format_cell_key("Input Values", "C", 5) == "'Input Values'!C5"

    def test_sheet_with_hyphen_quoted(self) -> None:
        """Sheet names with hyphens should produce quoted keys."""
        assert format_cell_key("Q4-Results", "A", 1) == "'Q4-Results'!A1"
        assert format_cell_key("2024-Budget", "B", 2) == "'2024-Budget'!B2"

    def test_complex_sheet_name_quoted(self) -> None:
        """Sheet names with multiple special characters should be quoted."""
        assert format_cell_key("Baseline - external", "M", 35) == "'Baseline - external'!M35"
        assert format_cell_key("GDP Forecast", "A", 1) == "'GDP Forecast'!A1"


class TestCellInfo:
    """Tests for the CellInfo dataclass."""

    def test_is_formula_true_when_formula_present(self) -> None:
        """is_formula should be True when formula is not None."""
        info = CellInfo(formula="=A1+B1", normalized_formula="Sheet1!A1+Sheet1!B1", value=10)
        assert info.is_formula is True

    def test_is_formula_false_when_formula_none(self) -> None:
        """is_formula should be False when formula is None."""
        info = CellInfo(formula=None, normalized_formula=None, value=42)
        assert info.is_formula is False


class TestCellDict:
    """Tests for the CellDict class."""

    def test_formula_cells_filters_correctly(self) -> None:
        """formula_cells() should return only cells with formulas."""
        cells = CellDict()
        cells["Sheet1!A1"] = CellInfo(formula="=B1", normalized_formula="Sheet1!B1", value=5)
        cells["Sheet1!B1"] = CellInfo(formula=None, normalized_formula=None, value=10)
        cells["Sheet1!C1"] = CellInfo(formula="=A1+B1", normalized_formula="Sheet1!A1+Sheet1!B1", value=15)

        formula_cells = cells.formula_cells()
        assert len(formula_cells) == 2
        assert "Sheet1!A1" in formula_cells
        assert "Sheet1!C1" in formula_cells
        assert "Sheet1!B1" not in formula_cells

    def test_value_cells_filters_correctly(self) -> None:
        """value_cells() should return only cells without formulas."""
        cells = CellDict()
        cells["Sheet1!A1"] = CellInfo(formula="=B1", normalized_formula="Sheet1!B1", value=5)
        cells["Sheet1!B1"] = CellInfo(formula=None, normalized_formula=None, value=10)
        cells["Sheet1!C1"] = CellInfo(formula=None, normalized_formula=None, value=20)

        value_cells = cells.value_cells()
        assert len(value_cells) == 2
        assert "Sheet1!B1" in value_cells
        assert "Sheet1!C1" in value_cells
        assert "Sheet1!A1" not in value_cells

    def test_formula_keys_returns_sorted_list(self) -> None:
        """formula_keys() should return sorted list of formula cell keys."""
        cells = CellDict()
        cells["Sheet1!C1"] = CellInfo(formula="=A1", normalized_formula="Sheet1!A1", value=1)
        cells["Sheet1!A1"] = CellInfo(formula="=B1", normalized_formula="Sheet1!B1", value=2)
        cells["Sheet1!B1"] = CellInfo(formula=None, normalized_formula=None, value=3)

        assert cells.formula_keys() == ["Sheet1!A1", "Sheet1!C1"]

    def test_value_keys_returns_sorted_list(self) -> None:
        """value_keys() should return sorted list of value cell keys."""
        cells = CellDict()
        cells["Sheet1!C1"] = CellInfo(formula=None, normalized_formula=None, value=1)
        cells["Sheet1!A1"] = CellInfo(formula="=B1", normalized_formula="Sheet1!B1", value=2)
        cells["Sheet1!B1"] = CellInfo(formula=None, normalized_formula=None, value=3)

        assert cells.value_keys() == ["Sheet1!B1", "Sheet1!C1"]


class TestDiscoverFormulaCellsInRows:
    """Tests for discover_formula_cells_in_rows function."""

    def test_discovers_formula_cells_with_numeric_values(self, tmp_path: Path) -> None:
        """Should discover formula cells that have numeric cached values."""
        excel_path = tmp_path / "discover_test.xlsx"

        # Use xlsxwriter to create workbook with cached values
        wb = xlsxwriter.Workbook(excel_path)
        ws = wb.add_worksheet("Sheet1")

        # Row 1: value, formula with numeric result, formula with text result
        ws.write_number(0, 0, 10)  # A1 = 10 (value)
        ws.write_formula(0, 1, "=A1*2", None, 20)  # B1 = =A1*2 (cached: 20)
        ws.write_formula(0, 2, '="text"', None, "text")  # C1 = ="text" (cached: "text")

        # Row 2: more formulas with numeric results
        ws.write_number(1, 0, 5)  # A2 = 5 (value)
        ws.write_formula(1, 1, "=A2+10", None, 15)  # B2 = =A2+10 (cached: 15)

        wb.close()

        targets = discover_formula_cells_in_rows(excel_path, "Sheet1", [1, 2])

        # Should find B1 and B2 (formulas with numeric values)
        # Should NOT find A1, A2 (values) or C1 (formula with text result)
        assert "Sheet1!B1" in targets
        assert "Sheet1!B2" in targets
        assert "Sheet1!A1" not in targets
        assert "Sheet1!A2" not in targets
        assert "Sheet1!C1" not in targets

    def test_returns_empty_for_nonexistent_sheet(self, tmp_path: Path) -> None:
        """Should return empty list for non-existent sheet."""
        excel_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(excel_path)
        wb.close()

        targets = discover_formula_cells_in_rows(excel_path, "NonExistent", [1])
        assert targets == []

    def test_quoted_sheet_names(self, tmp_path: Path) -> None:
        """Should return properly quoted keys for sheets with special chars."""
        excel_path = tmp_path / "quoted_test.xlsx"

        wb = xlsxwriter.Workbook(excel_path)
        ws = wb.add_worksheet("Sales Data")
        ws.write_number(0, 0, 100)  # A1 = 100
        ws.write_formula(0, 1, "=A1*2", None, 200)  # B1 = =A1*2
        wb.close()

        targets = discover_formula_cells_in_rows(excel_path, "Sales Data", [1])

        assert len(targets) == 1
        assert targets[0] == "'Sales Data'!B1"


class TestGraphToCellDict:
    """Tests for graph_to_cell_dict function."""

    def test_converts_graph_to_cell_dict(self, tmp_path: Path) -> None:
        """Should convert DependencyGraph to CellDict correctly."""
        excel_path = tmp_path / "graph_test.xlsx"

        wb = xlsxwriter.Workbook(excel_path)
        ws = wb.add_worksheet("Sheet1")
        ws.write_number(0, 0, 2)  # A1 = 2
        ws.write_number(1, 0, 3)  # A2 = 3
        ws.write_formula(2, 0, "=A1+A2", None, 5)  # A3 = =A1+A2
        wb.close()

        graph = create_dependency_graph(excel_path, ["Sheet1!A3"], load_values=True)
        cells = graph_to_cell_dict(graph)

        assert isinstance(cells, CellDict)
        assert "Sheet1!A1" in cells
        assert "Sheet1!A2" in cells
        assert "Sheet1!A3" in cells

        # Check formula cell
        assert cells["Sheet1!A3"].is_formula is True
        assert cells["Sheet1!A3"].formula == "=A1+A2"
        assert cells["Sheet1!A3"].value == 5

        # Check value cells
        assert cells["Sheet1!A1"].is_formula is False
        assert cells["Sheet1!A1"].value == 2


class TestBuildCellDict:
    """Tests for build_cell_dict function."""

    def test_builds_cell_dict_from_workbook(self, tmp_path: Path) -> None:
        """Should build CellDict from workbook with specified rows."""
        excel_path = tmp_path / "build_test.xlsx"

        wb = xlsxwriter.Workbook(excel_path)
        ws = wb.add_worksheet("Sheet1")

        # Row 1: inputs
        ws.write_number(0, 0, 10)  # A1 = 10
        ws.write_number(0, 1, 20)  # B1 = 20

        # Row 2: formulas (this is our target row)
        ws.write_formula(1, 0, "=A1*2", None, 20)  # A2 = =A1*2
        ws.write_formula(1, 1, "=B1+A2", None, 40)  # B2 = =B1+A2

        wb.close()

        cells = build_cell_dict(excel_path, {"Sheet1": [2]}, load_values=True)

        # Should have formula cells from row 2 and their dependencies from row 1
        assert "Sheet1!A2" in cells
        assert "Sheet1!B2" in cells
        assert "Sheet1!A1" in cells
        assert "Sheet1!B1" in cells

        # Verify formula cells
        assert cells["Sheet1!A2"].is_formula is True
        assert cells["Sheet1!B2"].is_formula is True

        # Verify value cells
        assert cells["Sheet1!A1"].is_formula is False
        assert cells["Sheet1!B1"].is_formula is False

    def test_returns_empty_for_no_formulas(self, tmp_path: Path) -> None:
        """Should return empty CellDict when no formula cells found."""
        excel_path = tmp_path / "no_formulas.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = 10
        ws["B1"].value = 20
        wb.save(excel_path)
        wb.close()

        cells = build_cell_dict(excel_path, {"Sheet1": [1]})
        assert len(cells) == 0

    def test_handles_quoted_sheet_names(self, tmp_path: Path) -> None:
        """Should handle sheet names that require quoting."""
        excel_path = tmp_path / "quoted_sheet.xlsx"

        wb = xlsxwriter.Workbook(excel_path)
        ws = wb.add_worksheet("My Data")
        ws.write_number(0, 0, 5)  # A1 = 5
        ws.write_formula(1, 0, "=A1+10", None, 15)  # A2 = =A1+10
        wb.close()

        cells = build_cell_dict(excel_path, {"My Data": [2]}, load_values=True)

        # Keys should be properly quoted
        assert "'My Data'!A2" in cells
        assert "'My Data'!A1" in cells

        # Verify the formula cell
        assert cells["'My Data'!A2"].is_formula is True
        assert cells["'My Data'!A2"].value == 15

    def test_cross_sheet_references_correct_quoting(self, tmp_path: Path) -> None:
        """Should handle cross-sheet references with correct quoting."""
        excel_path = tmp_path / "cross_sheet.xlsx"

        wb = xlsxwriter.Workbook(excel_path)

        # Simple sheet (no quoting needed)
        ws1 = wb.add_worksheet("Data")
        ws1.write_number(0, 0, 100)  # Data!A1 = 100

        # Sheet with space (needs quoting)
        ws2 = wb.add_worksheet("Output Sheet")
        ws2.write_formula(0, 0, "=Data!A1*2", None, 200)  # 'Output Sheet'!A1 = =Data!A1*2

        wb.close()

        cells = build_cell_dict(excel_path, {"Output Sheet": [1]}, load_values=True)

        # Data sheet key should NOT be quoted
        assert "Data!A1" in cells
        # Output Sheet key SHOULD be quoted
        assert "'Output Sheet'!A1" in cells

        # Verify proper quoting in all keys
        for key in cells:
            if "Output Sheet" in key:
                assert key.startswith("'Output Sheet'!"), f"Key {key} should be quoted"
            elif "Data" in key:
                assert not key.startswith("'Data'!"), f"Key {key} should NOT be quoted"
