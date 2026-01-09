"""
Flatten Excel dependency graphs for transpilation.

Builds dependency graphs from Excel workbooks and represents them as flat
dictionaries keyed by normalized cell address, ready for formula expansion
and Python transpilation.
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import openpyxl.utils.cell

from .builder import create_dependency_graph
from .graph import DependencyGraph
from .parser import format_cell_key

__all__ = [
    "CellInfo",
    "CellDict",
    "graph_to_cell_dict",
    "discover_formula_cells_in_rows",
    "build_cell_dict",
]


@dataclass
class CellInfo:
    """Cell information extracted from the dependency graph."""

    formula: str | None  # Original formula (None for non-formula cells)
    normalized_formula: str | None  # Sheet-qualified formula for transpilation
    value: Any  # Cached or hardcoded value

    @property
    def is_formula(self) -> bool:
        """True if this cell contains a formula."""
        return self.formula is not None


class CellDict(dict[str, CellInfo]):
    """
    Flat dictionary of cells keyed by normalized address (e.g., "'Sheet Name'!A1").

    Provides helper methods for filtering formula vs non-formula cells.
    """

    def formula_cells(self) -> dict[str, CellInfo]:
        """Return only cells that contain formulas."""
        return {k: v for k, v in self.items() if v.is_formula}

    def value_cells(self) -> dict[str, CellInfo]:
        """Return only cells that are hardcoded values (non-formula)."""
        return {k: v for k, v in self.items() if not v.is_formula}

    def formula_keys(self) -> list[str]:
        """Return sorted list of keys for formula cells."""
        return sorted(k for k, v in self.items() if v.is_formula)

    def value_keys(self) -> list[str]:
        """Return sorted list of keys for value cells."""
        return sorted(k for k, v in self.items() if not v.is_formula)


def graph_to_cell_dict(graph: DependencyGraph) -> CellDict:
    """
    Convert a DependencyGraph to a flat CellDict.

    Args:
        graph: DependencyGraph from excel-grapher

    Returns:
        CellDict keyed by normalized cell address (e.g., "'Sheet Name'!A1")
    """
    result = CellDict()
    for key in graph:
        node = graph.get_node(key)
        if node is None:
            continue
        result[key] = CellInfo(
            formula=node.formula,
            normalized_formula=node.normalized_formula,
            value=node.value,
        )
    return result


def discover_formula_cells_in_rows(
    wb_path: Path,
    sheet_name: str,
    rows: list[int],
) -> list[str]:
    """
    Scan specified rows and return sheet-qualified addresses for formula cells.

    Only includes cells that contain formulas (start with '=') and whose cached
    calculated value is numeric.

    Args:
        wb_path: Path to the Excel workbook
        sheet_name: Name of the sheet to scan
        rows: List of row numbers to scan

    Returns:
        List of sheet-qualified cell addresses (e.g., "'Sheet Name'!A1")
    """
    wb_formulas = openpyxl.load_workbook(wb_path, data_only=False, keep_vba=True)
    wb_values = openpyxl.load_workbook(wb_path, data_only=True, keep_vba=True)
    try:
        if sheet_name not in wb_formulas.sheetnames:
            return []

        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]
        targets: list[str] = []

        for row in rows:
            max_col = ws_formulas.max_column or 1
            for col_idx in range(1, max_col + 1):
                cell_formula = ws_formulas.cell(row=row, column=col_idx)
                if isinstance(cell_formula.value, str) and cell_formula.value.startswith("="):
                    cached_value = ws_values.cell(row=row, column=col_idx).value
                    if not isinstance(cached_value, (int, float)) or isinstance(cached_value, bool):
                        continue
                    col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
                    targets.append(format_cell_key(sheet_name, col_letter, row))

        return targets
    finally:
        wb_formulas.close()
        wb_values.close()


def build_cell_dict(
    workbook_path: Path,
    sheet_rows: dict[str, list[int]],
    load_values: bool = True,
    max_depth: int = 50,
) -> CellDict:
    """
    Build a CellDict from specified output rows in an Excel workbook.

    Args:
        workbook_path: Path to the Excel file
        sheet_rows: Dict mapping sheet names to lists of output row numbers
        load_values: Whether to load cached values (requires second workbook load)
        max_depth: Maximum dependency traversal depth

    Returns:
        CellDict with all cells participating in the dependency graph
    """
    # Discover all formula cells in the specified rows
    all_targets: list[str] = []
    for sheet_name, rows in sheet_rows.items():
        targets = discover_formula_cells_in_rows(workbook_path, sheet_name, rows)
        all_targets.extend(targets)

    if not all_targets:
        return CellDict()

    # Build the dependency graph
    graph = create_dependency_graph(
        workbook_path,
        all_targets,
        load_values=load_values,
        max_depth=max_depth,
    )

    return graph_to_cell_dict(graph)
