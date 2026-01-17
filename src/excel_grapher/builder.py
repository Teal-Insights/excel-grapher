from __future__ import annotations

import re
from collections import deque
from collections.abc import Iterable
from pathlib import Path

import openpyxl
import openpyxl.utils.cell
from openpyxl.worksheet.formula import ArrayFormula

from .graph import DependencyGraph, NodeHook
from .guard import And, Compare, GuardExpr, Literal, Not
from .node import Node
from .parser import (
    expand_range,
    format_key,
    mask_spans,
    normalize_formula,
    parse_cell_refs,
    parse_dynamic_range_refs_with_spans,
    parse_guard_expr,
    parse_range_refs_with_spans,
    split_top_level_choose,
    split_top_level_if,
    split_top_level_ifs,
    split_top_level_switch,
)
from .resolver import build_named_range_map


def create_dependency_graph(
    workbook: Path | str | openpyxl.Workbook,
    targets: Iterable[str],
    *,
    max_depth: int = 50,
    expand_ranges: bool = True,
    max_range_cells: int = 5000,
    hooks: list[NodeHook] | None = None,
    load_values: bool = True,
) -> DependencyGraph:
    """
    Build a dependency graph starting from target cells.

    Current implementation supports basic A1 references and sheet-qualified references.
    Range expansion and named range resolution will be added incrementally via TDD.
    """

    def load_wb(data_only: bool) -> openpyxl.Workbook:
        if isinstance(workbook, openpyxl.Workbook):
            if data_only:
                raise ValueError("load_values=True is not supported when passing a Workbook instance")
            return workbook
        path = Path(workbook)
        keep_vba = path.suffix.lower() == ".xlsm"
        return openpyxl.load_workbook(path, data_only=data_only, keep_vba=keep_vba)

    wb_formulas = load_wb(data_only=False)
    wb_values = load_wb(data_only=True) if load_values and not isinstance(workbook, openpyxl.Workbook) else None

    graph = DependencyGraph()
    for h in hooks or []:
        graph.register_hook(h)

    named_range_maps = build_named_range_map(wb_formulas)
    named_ranges = named_range_maps.cell_map
    named_range_ranges = named_range_maps.range_map
    defined_names = {str(name) for name in wb_formulas.defined_names}
    _NAME_TOKEN_RE = re.compile(r"\b([A-Za-z_][A-Za-z0-9_]*)\b(?!\s*!)")

    def parse_target(t: str) -> tuple[str, str]:
        if "!" not in t:
            raise ValueError(f"Target must be sheet-qualified: {t}")
        # Handle quoted sheet names: 'Sheet Name'!A1 or 'Sheet!Name'!A1
        if t.startswith("'"):
            # Find the closing quote
            end_quote = t.index("'", 1)
            sheet = t[1:end_quote]
            a1 = t[end_quote + 2:]  # Skip '!
        else:
            sheet, a1 = t.split("!", 1)
        if sheet not in wb_formulas.sheetnames:
            raise ValueError(f"Sheet not found: {sheet}")
        return sheet, a1

    def extract_deps_with_guards(
        formula: str, current_sheet: str, current_a1: str
    ) -> list[tuple[str, str, GuardExpr | None]]:
        if not formula.startswith("="):
            return []

        def resolve_cached_value(sheet: str, a1: str) -> object | None:
            nonlocal wb_values
            if wb_values is None and not isinstance(workbook, openpyxl.Workbook):
                wb_values = load_wb(data_only=True)
            if wb_values is None:
                return None
            return wb_values[sheet][a1].value

        def extract_expr_deps(expr: str) -> list[tuple[str, str]]:
            """
            Extract dependencies from an expression fragment (no leading '=').
            """
            f = "=" + expr if not expr.startswith("=") else expr
            deps: list[tuple[str, str]] = []

            masked = f

            # 0) Expand dynamic refs (OFFSET/INDIRECT) and mask them so underlying
            # refs don't get parsed as direct dependencies.
            dyn_spans: list[tuple[int, int]] = []
            for start, end, span in parse_dynamic_range_refs_with_spans(
                f,
                current_sheet=current_sheet,
                current_cell_a1=current_a1,
                named_ranges=named_ranges,
                named_range_ranges=named_range_ranges,
                value_resolver=resolve_cached_value,
            ):
                dyn_spans.append(span)
                sheet = start.sheet if start.sheet is not None else current_sheet
                deps.extend(
                    expand_range(
                        sheet=sheet,
                        start_col=start.column,
                        start_row=start.row,
                        end_col=end.column,
                        end_row=end.row,
                        max_cells=max_range_cells,
                    )
                )
            masked = mask_spans(masked, dyn_spans)

            # 1) Expand ranges, then mask them so later cell-ref parsing doesn't
            # misinterpret the range endpoint as a same-sheet ref.
            if expand_ranges:
                spans: list[tuple[int, int]] = []
                for start, end, span in parse_range_refs_with_spans(masked):
                    spans.append(span)
                    sheet = start.sheet if start.sheet is not None else current_sheet
                    deps.extend(
                        expand_range(
                            sheet=sheet,
                            start_col=start.column,
                            start_row=start.row,
                            end_col=end.column,
                            end_row=end.row,
                            max_cells=max_range_cells,
                        )
                    )
                masked = mask_spans(masked, spans)

            for ref in parse_cell_refs(masked):
                sh = ref.sheet if ref.sheet is not None else current_sheet
                deps.append((sh, f"{ref.column}{ref.row}"))

            # 3) Named ranges
            for m in _NAME_TOKEN_RE.finditer(masked):
                token = m.group(1)
                resolved = named_ranges.get(token)
                if resolved is not None:
                    deps.append(resolved)
                    continue
                resolved_range = named_range_ranges.get(token)
                if resolved_range is not None:
                    if expand_ranges:
                        sheet, start_a1, end_a1 = resolved_range
                        start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_a1)
                        end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_a1)
                        deps.extend(
                            expand_range(
                                sheet=sheet,
                                start_col=start_col,
                                start_row=int(start_row),
                                end_col=end_col,
                                end_row=int(end_row),
                                max_cells=max_range_cells,
                            )
                        )
                    continue
                if token in defined_names:
                    raise ValueError(f"Unsupported defined name referenced in formula: {token}")

            # Deduplicate while preserving order
            seen: set[tuple[str, str]] = set()
            out: list[tuple[str, str]] = []
            for d in deps:
                if d in seen:
                    continue
                seen.add(d)
                out.append(d)
            return out

        # 1) IF(cond, then, else)
        if_parts = split_top_level_if(formula)
        if if_parts is not None:
            cond_s, then_s, else_s = if_parts
            cond_guard = parse_guard_expr(cond_s, current_sheet=current_sheet, named_ranges=named_ranges)

            unconditional = set(extract_expr_deps(cond_s))
            out: dict[tuple[str, str], GuardExpr | None] = {(sh, a1): None for (sh, a1) in unconditional}

            # If the condition can't be parsed, branch deps are still conditional, but opaque.
            then_guard: GuardExpr | None = cond_guard
            else_guard: GuardExpr | None = None if cond_guard is None else Not(cond_guard)

            for sh, a1 in extract_expr_deps(then_s):
                key = (sh, a1)
                if key in out:
                    continue
                out[key] = then_guard

            if else_s:
                for sh, a1 in extract_expr_deps(else_s):
                    key = (sh, a1)
                    if key in out:
                        continue
                    out[key] = else_guard

            return [(sh, a1, g) for (sh, a1), g in out.items()]

        # 2) IFS(cond1, value1, cond2, value2, ..., [default])
        ifs_args = split_top_level_ifs(formula)
        if ifs_args is not None:
            # All condition expressions may be evaluated (sequentially), so include deps from all
            # conditions as unconditional to avoid missing required inputs.
            conditions: list[str] = []
            values: list[str] = []
            default_expr: str | None = None
            if len(ifs_args) >= 2:
                pairs = ifs_args
                if len(pairs) % 2 == 1:
                    default_expr = pairs[-1]
                    pairs = pairs[:-1]
                for i in range(0, len(pairs), 2):
                    conditions.append(pairs[i])
                    values.append(pairs[i + 1])

            unconditional: set[tuple[str, str]] = set()
            for c in conditions:
                unconditional |= set(extract_expr_deps(c))

            out: dict[tuple[str, str], GuardExpr | None] = {(sh, a1): None for (sh, a1) in unconditional}

            prev_negations: list[GuardExpr] = []
            for _idx, (cond_s, val_s) in enumerate(zip(conditions, values, strict=False), start=1):
                cond_guard = parse_guard_expr(cond_s, current_sheet=current_sheet, named_ranges=named_ranges)
                # Build sequential guard: cond_i AND NOT(cond_1) AND ... NOT(cond_{i-1})
                g: GuardExpr | None
                if cond_guard is None:
                    g = None
                else:
                    ops: list[GuardExpr] = [cond_guard, *prev_negations]
                    g = ops[0] if len(ops) == 1 else And(tuple(ops))
                    prev_negations.append(Not(cond_guard))

                for sh, a1 in extract_expr_deps(val_s):
                    key = (sh, a1)
                    if key in out:
                        continue
                    out[key] = g

            if default_expr is not None:
                if prev_negations:
                    default_guard: GuardExpr = prev_negations[0] if len(prev_negations) == 1 else And(tuple(prev_negations))
                else:
                    default_guard = Literal(True)
                for sh, a1 in extract_expr_deps(default_expr):
                    key = (sh, a1)
                    if key in out:
                        continue
                    out[key] = default_guard

            return [(sh, a1, g) for (sh, a1), g in out.items()]

        # 3) CHOOSE(index, value1, value2, ...)
        choose_args = split_top_level_choose(formula)
        if choose_args is not None and len(choose_args) >= 2:
            index_s = choose_args[0]
            choices = choose_args[1:]

            index_expr = parse_guard_expr(index_s, current_sheet=current_sheet, named_ranges=named_ranges)
            unconditional = set(extract_expr_deps(index_s))
            out: dict[tuple[str, str], GuardExpr | None] = {(sh, a1): None for (sh, a1) in unconditional}

            for i, choice_s in enumerate(choices, start=1):
                guard: GuardExpr | None = None
                if index_expr is not None:
                    guard = Compare(left=index_expr, op="=", right=Literal(i))
                for sh, a1 in extract_expr_deps(choice_s):
                    key = (sh, a1)
                    if key in out:
                        continue
                    out[key] = guard

            return [(sh, a1, g) for (sh, a1), g in out.items()]

        # 4) SWITCH(expr, value1, result1, ..., [default])
        switch_args = split_top_level_switch(formula)
        if switch_args is not None and len(switch_args) >= 3:
            expr_s = switch_args[0]
            expr_ge = parse_guard_expr(expr_s, current_sheet=current_sheet, named_ranges=named_ranges)
            unconditional = set(extract_expr_deps(expr_s))
            out: dict[tuple[str, str], GuardExpr | None] = {(sh, a1): None for (sh, a1) in unconditional}

            pairs = switch_args[1:]
            default_expr: str | None = None
            if len(pairs) % 2 == 1:
                default_expr = pairs[-1]
                pairs = pairs[:-1]

            prev_negations: list[GuardExpr] = []
            for i in range(0, len(pairs), 2):
                val_s = pairs[i]
                res_s = pairs[i + 1]
                val_ge = parse_guard_expr(val_s, current_sheet=current_sheet, named_ranges=named_ranges)
                match: GuardExpr | None = None
                if expr_ge is not None and val_ge is not None:
                    match = Compare(left=expr_ge, op="=", right=val_ge)

                guard: GuardExpr | None = None
                if match is not None:
                    ops2: list[GuardExpr] = [match, *prev_negations]
                    guard = ops2[0] if len(ops2) == 1 else And(tuple(ops2))
                    prev_negations.append(Not(match))

                for sh, a1 in extract_expr_deps(res_s):
                    key = (sh, a1)
                    if key in out:
                        continue
                    out[key] = guard

            if default_expr is not None:
                if prev_negations:
                    default_guard2: GuardExpr = prev_negations[0] if len(prev_negations) == 1 else And(tuple(prev_negations))
                else:
                    default_guard2 = Literal(True)
                for sh, a1 in extract_expr_deps(default_expr):
                    key = (sh, a1)
                    if key in out:
                        continue
                    out[key] = default_guard2

            return [(sh, a1, g) for (sh, a1), g in out.items()]

        return [(sh, a1, None) for (sh, a1) in extract_expr_deps(formula)]

    visited: set[str] = set()
    q: deque[tuple[str, str, int]] = deque()
    for t in targets:
        sh, a1 = parse_target(str(t))
        q.append((sh, a1, 0))

    try:
        while q:
            sheet, a1, depth = q.popleft()
            key = format_key(sheet, a1)
            if key in visited:
                continue
            visited.add(key)
            if depth > max_depth:
                continue

            ws_f = wb_formulas[sheet]
            raw = ws_f[a1].value
            if isinstance(raw, ArrayFormula):
                raw = raw.text or ""
                if raw and not raw.startswith("="):
                    raw = f"={raw}"
            is_formula = isinstance(raw, str) and raw.startswith("=")

            if is_formula:
                formula_str = str(raw)
                formula = formula_str
                normalized = normalize_formula(formula_str, sheet, named_ranges)
                value = None
                if wb_values is not None:
                    value = wb_values[sheet][a1].value
                is_leaf = False
            else:
                formula_str = ""
                formula = None
                normalized = None
                value = raw
                is_leaf = True

            col, row = openpyxl.utils.cell.coordinate_from_string(a1)
            node = Node(
                sheet=sheet,
                column=col,
                row=int(row),
                formula=formula,
                normalized_formula=normalized,
                value=value,
                is_leaf=is_leaf,
            )
            graph.add_node(node)

            if not is_formula:
                continue

            for dep_sheet, dep_a1, guard in extract_deps_with_guards(formula_str, sheet, a1):
                dep_key = format_key(dep_sheet, dep_a1)
                graph.add_edge(key, dep_key, guard=guard)
                if dep_key not in visited:
                    if dep_sheet not in wb_formulas.sheetnames:
                        continue
                    q.append((dep_sheet, dep_a1, depth + 1))
    finally:
        if wb_values is not None:
            wb_values.close()
        if not isinstance(workbook, openpyxl.Workbook):
            wb_formulas.close()

    return graph

