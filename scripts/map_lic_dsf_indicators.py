#!/usr/bin/env python3
"""
Map dependencies for LIC-DSF indicator rows using excel-grapher.

This script traces the dependency closure for key indicator rows across
B1, B3, and B4 sheets and validates against calcChain.xml.
"""

from pathlib import Path

import openpyxl
import openpyxl.utils.cell

from excel_grapher import (
    create_dependency_graph,
    validate_graph,
    to_graphviz,
)

# Configuration: sheets and indicator rows to trace
INDICATOR_CONFIG = [
    {"sheet": "B1_GDP_ext", "indicator_rows": [35, 36, 39, 40]},
    {"sheet": "B3_Exports_ext", "indicator_rows": [35, 36, 39, 40]},
    {"sheet": "B4_other flows_ext", "indicator_rows": [35, 36, 39, 40]},
]

WORKBOOK_PATH = Path("Gold-Standard-LIC-DSF.xlsm")


def discover_formula_cells_in_rows(
    wb_path: Path,
    sheet_name: str,
    rows: list[int],
) -> list[str]:
    """
    Scan specified rows and return sheet-qualified addresses for formula cells.
    
    Only includes cells that contain formulas (start with '=').
    """
    wb = openpyxl.load_workbook(wb_path, data_only=False, keep_vba=True)
    try:
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: Sheet '{sheet_name}' not found")
            return []
        
        ws = wb[sheet_name]
        targets: list[str] = []
        
        for row in rows:
            # Scan all columns up to max_column
            max_col = ws.max_column or 1
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
                    targets.append(f"{sheet_name}!{col_letter}{row}")
        
        return targets
    finally:
        wb.close()


def main() -> None:
    print("=" * 70)
    print("LIC-DSF Indicator Dependency Mapping")
    print("=" * 70)
    
    if not WORKBOOK_PATH.exists():
        print(f"Error: Workbook not found at {WORKBOOK_PATH}")
        print("Make sure Gold-Standard-LIC-DSF.xlsm is in the project root.")
        return
    
    # Discover all formula cells in indicator rows
    print("\n1. Discovering formula cells in indicator rows...")
    all_targets: list[str] = []
    
    for config in INDICATOR_CONFIG:
        sheet = config["sheet"]
        rows = config["indicator_rows"]
        print(f"   {sheet}: rows {rows}")
        
        targets = discover_formula_cells_in_rows(WORKBOOK_PATH, sheet, rows)
        print(f"      Found {len(targets)} formula cells")
        all_targets.extend(targets)
    
    print(f"\n   Total targets: {len(all_targets)}")
    
    if not all_targets:
        print("No formula cells found. Exiting.")
        return
    
    # Build dependency graph
    print("\n2. Building dependency graph...")
    graph = create_dependency_graph(
        WORKBOOK_PATH,
        all_targets,
        load_values=False,  # Skip cached values for speed
        max_depth=50,
    )
    
    print(f"   Nodes in graph: {len(graph)}")
    print(f"   Leaf nodes: {sum(1 for _ in graph.leaves())}")
    print(f"   Formula nodes: {len(graph) - sum(1 for _ in graph.leaves())}")
    
    # Group nodes by sheet
    sheets: dict[str, int] = {}
    for key in graph:
        node = graph.get_node(key)
        if node:
            sheets[node.sheet] = sheets.get(node.sheet, 0) + 1
    
    print(f"\n   Nodes by sheet:")
    for sheet_name in sorted(sheets.keys()):
        print(f"      {sheet_name}: {sheets[sheet_name]}")
    
    # Validate against calcChain.xml
    print("\n3. Validating against calcChain.xml...")
    scope = {config["sheet"] for config in INDICATOR_CONFIG}
    result = validate_graph(graph, WORKBOOK_PATH, scope=scope)
    
    print(f"   Valid: {result.is_valid}")
    for msg in result.messages:
        print(f"   {msg}")
    
    if result.in_graph_not_in_chain:
        print(f"\n   Cells in graph but not in calcChain ({len(result.in_graph_not_in_chain)}):")
        for cell in sorted(result.in_graph_not_in_chain)[:10]:
            print(f"      {cell}")
        if len(result.in_graph_not_in_chain) > 10:
            print(f"      ... and {len(result.in_graph_not_in_chain) - 10} more")
    
    # Evaluation order stats
    print("\n4. Computing evaluation order...")
    try:
        order = graph.evaluation_order()
        print(f"   Evaluation order computed: {len(order)} nodes")
        print(f"   First 5 (leaves): {order[:5]}")
        print(f"   Last 5 (targets): {order[-5:]}")
    except ValueError as e:
        print(f"   Error: {e}")
    
    # Optional: save a small subgraph visualization
    print("\n5. Sample visualization (first target's immediate deps)...")
    if all_targets:
        sample_target = all_targets[0]
        sample_deps = graph.dependencies(sample_target)
        print(f"   {sample_target} depends on {len(sample_deps)} cells:")
        for dep in sorted(sample_deps)[:5]:
            print(f"      {dep}")
        if len(sample_deps) > 5:
            print(f"      ... and {len(sample_deps) - 5} more")
    
    print("\n" + "=" * 70)
    print("Done.")


if __name__ == "__main__":
    main()
