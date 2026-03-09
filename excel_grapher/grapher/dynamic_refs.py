from __future__ import annotations

import re
from collections.abc import Callable, Iterable, Mapping
from dataclasses import dataclass
from itertools import product
from pathlib import Path
from typing import Any, get_args, get_origin, get_type_hints

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple

from excel_grapher.core.addressing import (
    WorkbookBoundsProtocol,
    indirect_text_to_range,
    offset_range,
)
from excel_grapher.core.cell_types import (
    CellKind,
    CellType,
    CellTypeEnv,
    EnumDomain,
    IntIntervalDomain,
    constraints_to_cell_type_env,
)
from excel_grapher.core.excel_function_meta import is_ref_only_arg
from excel_grapher.core.expr_eval import Unsupported, evaluate_expr
from excel_grapher.core.formula_ast import (
    AstNode,
    CellRefNode,
    FormulaParseError,
    FunctionCallNode,
    RangeNode,
)
from excel_grapher.core.formula_ast import (
    parse as parse_ast,
)
from excel_grapher.core.types import ExcelRange, XlError

from .parser import _find_function_calls_with_spans, format_key


class DynamicRefError(ValueError):
    """Raised when dynamic reference analysis cannot proceed.

    When building a dependency graph, pass a :class:`DynamicRefConfig` (e.g. via
    :meth:`DynamicRefConfig.from_constraints`) or set ``use_cached_dynamic_refs=True``
    to resolve OFFSET/INDIRECT instead of raising.
    """


@dataclass(frozen=True)
class DynamicRefLimits:
    max_branches: int = 1024
    max_cells: int = 10_000
    max_depth: int = 10


@dataclass(frozen=True)
class DynamicRefConfig:
    """Configuration for resolving OFFSET/INDIRECT via constraint-based inference.

    Prefer building via :meth:`from_constraints`; the constructor is for internal use.
    """

    cell_type_env: CellTypeEnv
    limits: DynamicRefLimits

    @classmethod
    def from_constraints(
        cls,
        constraints_type: type[Any],
        constraints_data: Mapping[str, Any],
        *,
        limits: DynamicRefLimits | None = None,
    ) -> DynamicRefConfig:
        """Build a config from a TypedDict (or type with type hints) and a validated instance.

        Keys of constraints_type (from get_type_hints(..., include_extras=True)) should be
        address-style cell addresses (e.g. \"Sheet1!B1\"). constraints_data must have the
        same keys. No validation of constraints_data is performed; use TypeAdapter etc. if needed.
        """
        env = constraints_to_cell_type_env(constraints_type, constraints_data)
        return cls(cell_type_env=env, limits=limits or DynamicRefLimits())

    @classmethod
    def from_constraints_and_workbook(
        cls,
        constraints_type: type[Any],
        workbook_path: str | Path,
        *,
        limits: DynamicRefLimits | None = None,
        data_only: bool = True,
    ) -> DynamicRefConfig:
        """Build config from constraints type plus workbook values for constant cells.

        Constraints whose annotations carry a FromWorkbook marker are treated as
        singleton domains derived from the current cached value in the workbook.
        Other constraints are interpreted via constraints_to_cell_type_env as usual.
        """
        hints = get_type_hints(constraints_type, include_extras=True)
        # Start from the type-based environment (Literal/Between/etc.).
        dummy_data: dict[str, Any] = {k: None for k in hints}
        env = constraints_to_cell_type_env(constraints_type, dummy_data)

        wb = openpyxl.load_workbook(Path(workbook_path), data_only=data_only, keep_vba=True)
        try:
            for addr, annotated_type in hints.items():
                if not _has_from_workbook_marker(annotated_type):
                    continue
                sheet_name, coord = _split_addr_sheet_coord(addr)
                if sheet_name not in wb.sheetnames:
                    raise DynamicRefError(
                        f"Sheet {sheet_name!r} (from constraint {addr!r}) not found in workbook"
                    )
                ws = wb[sheet_name]
                value = ws[coord].value
                if value is None:
                    # Skip empty cells; caller can add explicit constraints if needed.
                    continue
                kind = _infer_kind_from_value(value)
                env[addr] = CellType(
                    kind=kind,
                    enum=EnumDomain(values=frozenset({value})),
                )
        finally:
            wb.close()

        return cls(cell_type_env=env, limits=limits or DynamicRefLimits())


@dataclass(frozen=True)
class FromWorkbook:
    """Metadata marker: domain is the current workbook value."""


def _has_from_workbook_marker(annotated_type: Any) -> bool:
    try:
        from typing import Annotated  # type: ignore
    except ImportError:  # pragma: no cover - Annotated always available in supported versions
        return False

    if get_origin(annotated_type) is not Annotated:
        return False
    args = get_args(annotated_type)
    if len(args) < 2:
        return False
    metadata = args[1:]
    return any(isinstance(m, FromWorkbook) for m in metadata)


def _split_addr_sheet_coord(addr: str) -> tuple[str, str]:
    """Split an address-style key into (sheet_name, coord) and normalize quoting."""
    if "!" not in addr:
        raise DynamicRefError(f"Constraint address must be sheet-qualified: {addr!r}")
    sheet_part, coord = addr.split("!", 1)
    sheet_part = sheet_part.strip()
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1].replace("''", "'")
    return sheet_part, coord


def _infer_kind_from_value(value: Any) -> CellKind:
    if isinstance(value, (int, float)):
        return CellKind.NUMBER
    if isinstance(value, bool):
        return CellKind.BOOL
    if isinstance(value, str):
        return CellKind.STRING
    return CellKind.ANY


@dataclass(frozen=True)
class GlobalWorkbookBounds(WorkbookBoundsProtocol):
    """Simple bounds implementation using Excel's global sheet limits."""

    sheet: str
    min_row: int = 1
    max_row: int = 1_048_576  # Excel row limit
    min_col: int = 1
    max_col: int = 16_384  # Excel column limit


def _sheet_from_addr(addr: str) -> str:
    """Return sheet part of address (e.g. 'Sheet1!A1' -> 'Sheet1')."""
    if "!" not in addr:
        return ""
    if addr.startswith("'"):
        end = addr.index("'", 1)
        return addr[1:end]
    return addr.split("!", 1)[0]


def expand_leaf_env_to_argument_env(
    argument_refs: set[str],
    get_cell_formula: Callable[[str], str | None],
    get_refs_from_formula: Callable[[str, str], set[str]],
    leaf_env: CellTypeEnv,
    limits: DynamicRefLimits,
    named_ranges: Mapping[str, tuple[str, str]] | None = None,
    named_range_ranges: Mapping[str, tuple[str, str, str]] | None = None,
) -> dict[str, CellType]:
    """Build a CellTypeEnv for all refs in the argument chain from leaf constraints only.

    The cell env targets leaves: only leaf (non-formula) addresses need to be in
    leaf_env. Intermediate (formula) cells are inferred by evaluating their formulas
    over their dependencies' domains; they do not need to be constrained. If an
    intermediate is in leaf_env, that type is used and we do not traverse that branch.
    When an intermediate cannot be inferred (e.g. its formula is OFFSET/INDIRECT and
    refs are empty after masking), it is assigned CellType(ANY); enumeration may then
    require a constraint for that cell.
    """
    cache: dict[str, CellType] = {}
    nr = named_ranges or {}
    nrr = named_range_ranges or {}

    def _formula_to_parse(raw: str) -> str:
        body = raw[1:] if raw.startswith("=") else raw
        qualified = _qualify_fragment(body, nr, nrr)
        return "=" + qualified

    def _values_to_cell_type(values: set[Any]) -> CellType:
        if not values:
            return CellType(kind=CellKind.ANY)
        kinds = {type(v) for v in values}
        if kinds <= {int, float}:
            return CellType(
                kind=CellKind.NUMBER,
                enum=EnumDomain(values=frozenset(values)),
            )
        if kinds <= {str}:
            return CellType(
                kind=CellKind.STRING,
                enum=EnumDomain(values=frozenset(values)),
            )
        if kinds <= {bool}:
            return CellType(
                kind=CellKind.BOOL,
                enum=EnumDomain(values=frozenset(values)),
            )
        return CellType(kind=CellKind.ANY, enum=EnumDomain(values=frozenset(values)))

    def cell_type_for(addr: str) -> CellType:
        if addr in cache:
            return cache[addr]
        if addr in leaf_env:
            cache[addr] = leaf_env[addr]
            return cache[addr]
        formula = get_cell_formula(addr)
        if formula is None:
            raise DynamicRefError(
                f"Missing constraint for leaf {addr!r} that feeds OFFSET/INDIRECT. "
                "Add constraints only for leaf cells (non-formula) in the argument subgraph."
            )
        refs = get_refs_from_formula(formula, _sheet_from_addr(addr))
        if not refs:
            try:
                formula_parse = _formula_to_parse(formula)
                ast = parse_ast(formula_parse)
            except FormulaParseError:
                cache[addr] = CellType(kind=CellKind.ANY)
                return cache[addr]
            try:
                val = evaluate_expr(ast, get_cell_value=lambda _: None, max_depth=limits.max_depth)
            except Exception:
                val = None
            if isinstance(val, Unsupported):
                cache[addr] = CellType(kind=CellKind.ANY)
                return cache[addr]
            if val is None or isinstance(val, XlError):
                cache[addr] = CellType(kind=CellKind.ANY)
                return cache[addr]
            cache[addr] = _values_to_cell_type({val})
            return cache[addr]
        ref_types = {r: cell_type_for(r) for r in refs}
        domains: dict[str, list[Any]] = {}
        for r, ct in ref_types.items():
            if ct.enum is not None:
                domains[r] = list(ct.enum.values)
            elif ct.interval is not None:
                domains[r] = _interval_to_values(ct.interval, limits)
            else:
                cache[addr] = CellType(kind=CellKind.ANY)
                return cache[addr]
        result_values: set[Any] = set()
        last_unsupported: Unsupported | None = None
        for assignment in product(*(domains[r] for r in refs)):
            addr_to_val = dict(zip(refs, assignment, strict=False))

            def get_cell_value(a: str, _av=addr_to_val) -> Any:
                return _av.get(a)

            try:
                formula_parse = _formula_to_parse(formula)
                ast = parse_ast(formula_parse)
            except FormulaParseError:
                cache[addr] = CellType(kind=CellKind.ANY)
                return cache[addr]
            val = evaluate_expr(
                ast,
                get_cell_value=get_cell_value,
                max_depth=limits.max_depth,
            )
            if isinstance(val, Unsupported):
                last_unsupported = val
                continue
            if isinstance(val, XlError):
                continue
            result_values.add(val)
        if not result_values:
            if last_unsupported is not None:
                cache[addr] = CellType(kind=CellKind.ANY)
                return cache[addr]
            cache[addr] = CellType(kind=CellKind.ANY)
            return cache[addr]
        cache[addr] = _values_to_cell_type(result_values)
        return cache[addr]

    for addr in argument_refs:
        cell_type_for(addr)
    return cache


def infer_dynamic_offset_targets(
    formula: str,
    *,
    current_sheet: str,
    cell_type_env: CellTypeEnv,
    limits: DynamicRefLimits | None = None,
    bounds: WorkbookBoundsProtocol | None = None,
    named_ranges: Mapping[str, tuple[str, str]] | None = None,
    named_range_ranges: Mapping[str, tuple[str, str, str]] | None = None,
    current_row: int | None = None,
    current_col: int | None = None,
) -> set[str]:
    """Infer the union of all possible OFFSET targets for a formula.

    This helper is intentionally focused and conservative:
    - Only OFFSET calls are analysed (INDIRECT is currently ignored).
    - Arguments may use a small Excel expression subset supported by
      ``core.expr_eval.evaluate_expr``.
    - Leaf cells referenced by OFFSET/INDEX arguments must have a numeric
      domain in ``cell_type_env`` unless they appear only in ref_only
      argument positions (see :mod:`excel_grapher.core.excel_function_meta`).
    - Integer interval domains must be finite and small enough to enumerate.
    """

    if not isinstance(formula, str) or not formula.startswith("="):
        return set()

    lim = limits or DynamicRefLimits()
    out: set[str] = set()

    calls = _find_function_calls_with_spans(formula, {"OFFSET"})
    for fn, inner, _span in calls:
        if fn != "OFFSET":
            continue
        targets = _infer_single_offset_call(
            inner,
            current_sheet=current_sheet,
            cell_type_env=cell_type_env,
            limits=lim,
            bounds=bounds,
            named_ranges=named_ranges,
            named_range_ranges=named_range_ranges,
            current_row=current_row,
            current_col=current_col,
        )
        out |= targets
        if len(out) > lim.max_cells:
            raise DynamicRefError(
                f"Dynamic ref cells exceed limit ({len(out)} > {lim.max_cells})"
            )

    return out


def _qualify_fragment(
    expr: str,
    named_ranges: Mapping[str, tuple[str, str]],
    named_range_ranges: Mapping[str, tuple[str, str, str]] | None,
) -> str:
    """Replace named range tokens in a formula fragment with sheet-qualified refs so the parser can parse it."""
    if not expr.strip():
        return expr
    # Replace longer names first so "Country_list" is not partially matched by "Count".
    all_names = sorted(
        set(named_ranges.keys()) | (set(named_range_ranges.keys()) if named_range_ranges else set()),
        key=lambda n: (-len(n), n),
    )
    result = expr
    for name in all_names:
        if name in named_ranges:
            sheet, addr = named_ranges[name]
            replacement = format_key(sheet, addr)
        elif named_range_ranges and name in named_range_ranges:
            sheet, start_a1, end_a1 = named_range_ranges[name]
            replacement = f"{format_key(sheet, start_a1)}:{format_key(sheet, end_a1)}"
        else:
            continue
        result = re.sub(rf"\b{re.escape(name)}\b(?!\s*!)", replacement, result)
    return result


def _infer_single_offset_call(
    inner_args: str,
    *,
    current_sheet: str,
    cell_type_env: CellTypeEnv,
    limits: DynamicRefLimits,
    bounds: WorkbookBoundsProtocol | None,
    named_ranges: Mapping[str, tuple[str, str]] | None = None,
    named_range_ranges: Mapping[str, tuple[str, str, str]] | None = None,
    current_row: int | None = None,
    current_col: int | None = None,
) -> set[str]:
    """Infer targets for a single OFFSET(...) call body."""

    args = _split_top_level_args(inner_args)
    if args is None or len(args) < 3 or len(args) > 5:
        raise DynamicRefError("OFFSET expects 3 to 5 arguments")

    named_ranges = named_ranges or {}
    named_range_ranges = named_range_ranges or {}

    base_expr = _qualify_fragment(args[0], named_ranges, named_range_ranges)
    rows_expr = _qualify_fragment(args[1], named_ranges, named_range_ranges)
    cols_expr = _qualify_fragment(args[2], named_ranges, named_range_ranges)
    height_expr = (
        _qualify_fragment(args[3], named_ranges, named_range_ranges)
        if len(args) >= 4 and args[3]
        else ""
    )
    width_expr = (
        _qualify_fragment(args[4], named_ranges, named_range_ranges)
        if len(args) >= 5 and args[4]
        else ""
    )

    try:
        base_ast = parse_ast("=" + base_expr)
        base_ranges = _resolve_offset_base(
            base_ast,
            current_sheet=current_sheet,
            cell_type_env=cell_type_env,
            limits=limits,
            current_row=current_row,
            current_col=current_col,
        )
    except DynamicRefError as exc:
        raise DynamicRefError(
            f"{exc} (OFFSET base expression {base_expr!r})"
        ) from exc

    rows_ast = parse_ast("=" + rows_expr)
    cols_ast = parse_ast("=" + cols_expr)
    height_ast = parse_ast("=" + height_expr) if height_expr else None
    width_ast = parse_ast("=" + width_expr) if width_expr else None

    leaf_addrs: set[str] = set()
    leaf_addrs |= _collect_addresses(rows_ast)
    leaf_addrs |= _collect_addresses(cols_ast)
    if height_ast is not None:
        leaf_addrs |= _collect_addresses(height_ast)
    if width_ast is not None:
        leaf_addrs |= _collect_addresses(width_ast)

    domains = _build_domains(leaf_addrs, cell_type_env, limits)

    eval_context = (
        {"row": current_row, "column": current_col}
        if current_row is not None and current_col is not None
        else None
    )

    targets: set[str] = set()
    for base_range in base_ranges:
        base_bounds = (
            GlobalWorkbookBounds(sheet=base_range.sheet) if bounds is None else bounds
        )

        for assignment in _enumerate_assignments(domains.values(), limits):
            addr_to_value = dict(zip(domains.keys(), assignment, strict=False))

            def get_cell_value(addr: str, addr_to_value_map=addr_to_value) -> float:
                try:
                    return addr_to_value_map[addr]
                except KeyError as exc:
                    raise DynamicRefError(
                        f"OFFSET argument formula references cell without domain: {addr!r}"
                    ) from exc

            rows_val = _eval_arg(rows_ast, get_cell_value, limits, context=eval_context)
            cols_val = _eval_arg(cols_ast, get_cell_value, limits, context=eval_context)
            height_val = (
                _eval_arg(height_ast, get_cell_value, limits, context=eval_context)
                if height_ast is not None
                else None
            )
            width_val = (
                _eval_arg(width_ast, get_cell_value, limits, context=eval_context)
                if width_ast is not None
                else None
            )

            if isinstance(rows_val, XlError) or isinstance(cols_val, XlError):
                continue
            if isinstance(height_val, XlError) or isinstance(width_val, XlError):
                continue

            result = offset_range(
                base_range,
                rows=rows_val,
                cols=cols_val,
                height=height_val,
                width=width_val,
                bounds=base_bounds,
            )
            if isinstance(result, ExcelRange):
                targets |= set(result.cell_addresses())

    return targets


def _eval_arg(
    node: AstNode | None,
    get_cell_value,
    limits: DynamicRefLimits,
    context: dict[str, int] | None = None,
) -> float | XlError:
    if node is None:
        return 0.0

    value = evaluate_expr(
        node,
        get_cell_value=get_cell_value,
        max_depth=limits.max_depth,
        context=context,
    )
    if isinstance(value, Unsupported):
        raise DynamicRefError(f"Unsupported argument expression: {value.reason or ''}")
    if isinstance(value, XlError):
        return value
    if isinstance(value, (int, float)):
        return float(value)
    raise DynamicRefError(f"Non-numeric OFFSET argument result: {value!r}")


def _resolve_offset_base(
    base_ast: AstNode,
    *,
    current_sheet: str,
    cell_type_env: CellTypeEnv,
    limits: DynamicRefLimits,
    current_row: int | None = None,
    current_col: int | None = None,
) -> list[ExcelRange]:
    """Resolve OFFSET base to a list of candidate ranges (one per cell when base is INDEX)."""
    if isinstance(base_ast, CellRefNode):
        r = _base_to_range(base_ast, current_sheet=current_sheet)
        return [r]
    if isinstance(base_ast, RangeNode):
        r = _base_to_range(base_ast, current_sheet=current_sheet)
        return [r]
    if isinstance(base_ast, FunctionCallNode) and base_ast.name == "INDEX":
        if len(base_ast.args) < 2 or len(base_ast.args) > 3:
            raise DynamicRefError(
                "OFFSET base INDEX must have 2 or 3 arguments (array, row_num, [column_num])"
            )
        array_ast, row_ast = base_ast.args[0], base_ast.args[1]
        col_ast = base_ast.args[2] if len(base_ast.args) >= 3 else None
        array_range = _base_to_range(array_ast, current_sheet=current_sheet)
        leaf_addrs_for_domains = _collect_addresses_needing_domain(row_ast) | (
            _collect_addresses_needing_domain(col_ast) if col_ast is not None else set()
        )
        domains = _build_domains(leaf_addrs_for_domains, cell_type_env, limits)
        bases: list[ExcelRange] = []
        for assignment in _enumerate_assignments(domains.values(), limits):
            addr_to_value = dict(zip(domains.keys(), assignment, strict=False))

            def get_cell_value(addr: str, m=addr_to_value) -> float:
                try:
                    return m[addr]
                except KeyError as exc:
                    raise DynamicRefError(
                        f"INDEX argument formula references cell without domain: {addr!r}"
                    ) from exc

            eval_context = (
                {"row": current_row, "column": current_col}
                if current_row is not None and current_col is not None
                else None
            )
            row_val = _eval_arg(row_ast, get_cell_value, limits, context=eval_context)
            col_val = (
                _eval_arg(col_ast, get_cell_value, limits, context=eval_context)
                if col_ast is not None
                else 1.0
            )
            if isinstance(row_val, XlError) or isinstance(col_val, XlError):
                continue
            r1, c1 = int(row_val), int(col_val)
            if r1 < 1 or c1 < 1:
                continue
            cell_row = array_range.start_row + r1 - 1
            cell_col = array_range.start_col + c1 - 1
            if cell_row > array_range.end_row or cell_col > array_range.end_col:
                continue
            bases.append(
                ExcelRange(
                    sheet=array_range.sheet,
                    start_row=cell_row,
                    start_col=cell_col,
                    end_row=cell_row,
                    end_col=cell_col,
                )
            )
        return bases
    raise DynamicRefError("OFFSET base must be a cell or range reference")


def _base_to_range(base_ast: AstNode, *, current_sheet: str) -> ExcelRange:
    if isinstance(base_ast, CellRefNode):
        sheet, coord = _split_address(base_ast.address, current_sheet=current_sheet)
        row, col = coordinate_to_tuple(coord)
        return ExcelRange(
            sheet=sheet,
            start_row=row,
            start_col=col,
            end_row=row,
            end_col=col,
        )
    if isinstance(base_ast, RangeNode):
        try:
            s1, coord1 = base_ast.start.split("!", 1)
            s2, coord2 = base_ast.end.split("!", 1)
        except ValueError as exc:
            raise DynamicRefError("OFFSET base range must be a single-sheet A1 range") from exc
        if s1 != s2:
            raise DynamicRefError("OFFSET base range must be on a single sheet")
        row1, col1 = coordinate_to_tuple(coord1)
        row2, col2 = coordinate_to_tuple(coord2)
        start_row, end_row = sorted((row1, row2))
        start_col, end_col = sorted((col1, col2))
        return ExcelRange(
            sheet=s1,
            start_row=start_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
        )
    raise DynamicRefError("OFFSET base must be a cell or range reference")


def _split_address(addr: str, *, current_sheet: str) -> tuple[str, str]:
    if "!" in addr:
        sheet, coord = addr.split("!", 1)
        return sheet, coord
    return current_sheet, addr


def _collect_addresses_needing_domain(node: AstNode) -> set[str]:
    """Return cell/range addresses that appear in a value context (need a numeric domain).

    Refs that appear only as ref_only arguments (e.g. ROW(ref), COLUMN(ref)) are
    excluded; their implementations use only the reference, not the cell value.
    """
    addrs: set[str] = set()

    def visit(
        n: AstNode,
        parent: AstNode | None = None,
        arg_index: int | None = None,
    ) -> None:
        if isinstance(n, CellRefNode):
            if not (
                parent is not None
                and isinstance(parent, FunctionCallNode)
                and arg_index is not None
                and is_ref_only_arg(parent.name, arg_index)
            ):
                addrs.add(n.address)
            return
        if isinstance(n, RangeNode):
            try:
                sheet, coord_start = n.start.split("!", 1)
                _sheet2, coord_end = n.end.split("!", 1)
            except ValueError:
                return
            row1, col1 = coordinate_to_tuple(coord_start)
            row2, col2 = coordinate_to_tuple(coord_end)
            rlo, rhi = sorted((row1, row2))
            clo, chi = sorted((col1, col2))
            from openpyxl.utils.cell import get_column_letter

            for r in range(rlo, rhi + 1):
                for c in range(clo, chi + 1):
                    col_letter = get_column_letter(c)
                    addrs.add(f"{sheet}!{col_letter}{r}")
            return
        if isinstance(n, FunctionCallNode):
            for i, arg in enumerate(n.args):
                visit(arg, n, i)
            return
        if hasattr(n, "left") and hasattr(n, "right"):
            visit(n.left, n, None)  # type: ignore[arg-type]
            visit(n.right, n, None)  # type: ignore[arg-type]
        if hasattr(n, "operand"):
            visit(n.operand, n, None)  # type: ignore[arg-type]

    visit(node)
    return addrs


def _collect_addresses(node: AstNode) -> set[str]:
    addrs: set[str] = set()

    def visit(n: AstNode) -> None:
        if isinstance(n, CellRefNode):
            addrs.add(n.address)
            return
        if isinstance(n, RangeNode):
            try:
                sheet, coord_start = n.start.split("!", 1)
                _sheet2, coord_end = n.end.split("!", 1)
            except ValueError:
                return
            row1, col1 = coordinate_to_tuple(coord_start)
            row2, col2 = coordinate_to_tuple(coord_end)
            rlo, rhi = sorted((row1, row2))
            clo, chi = sorted((col1, col2))
            for r in range(rlo, rhi + 1):
                for c in range(clo, chi + 1):
                    # coordinate_to_tuple gives (row, col); we need back to A1
                    from openpyxl.utils.cell import get_column_letter

                    col_letter = get_column_letter(c)
                    addrs.add(f"{sheet}!{col_letter}{r}")
            return
        if isinstance(n, FunctionCallNode):
            for arg in n.args:
                visit(arg)
            return
        # Binary/unary ops and other nodes: recurse into children where present.
        if hasattr(n, "left") and hasattr(n, "right"):
            visit(n.left)  # type: ignore[arg-type]
            visit(n.right)  # type: ignore[arg-type]
        if hasattr(n, "operand"):
            visit(n.operand)  # type: ignore[arg-type]

    visit(node)
    return addrs


def _build_domains(
    addrs: Iterable[str],
    env: CellTypeEnv,
    limits: DynamicRefLimits,
) -> dict[str, list[int]]:
    domains: dict[str, list[int]] = {}
    for addr in addrs:
        ct = env.get(addr)
        if ct is None:
            raise DynamicRefError(f"Missing CellType for {addr!r}")
        if ct.kind is not CellKind.NUMBER:
            raise DynamicRefError(f"CellType for {addr!r} must be numeric, got {ct.kind!r}")
        vals: list[int]
        if ct.enum is not None:
            vals = [int(v) for v in ct.enum.values]
        elif ct.interval is not None:
            vals = _interval_to_values(ct.interval, limits)
        else:
            raise DynamicRefError(
                f"CellType for {addr!r} has no enum or interval domain (e.g. formula uses "
                "OFFSET/INDIRECT and could not be inferred). Add a constraint for this cell."
            )
        if not vals:
            raise DynamicRefError(f"Empty domain for {addr!r}")
        domains[addr] = sorted(vals)

    total = 1
    for vs in domains.values():
        total *= len(vs)
        if total > limits.max_branches:
            raise DynamicRefError(
                f"Dynamic ref branches exceed limit ({total} > {limits.max_branches})"
            )
    return domains


def _interval_to_values(interval: IntIntervalDomain, limits: DynamicRefLimits) -> list[int]:
    if interval.min is None or interval.max is None:
        raise DynamicRefError("Unbounded intervals are not supported for dynamic refs")
    lo, hi = int(interval.min), int(interval.max)
    if hi < lo:
        raise DynamicRefError(f"Invalid interval domain [{lo}, {hi}]")
    count = hi - lo + 1
    if count > limits.max_branches:
        raise DynamicRefError(
            f"Interval size {count} exceeds branch limit {limits.max_branches}"
        )
    return list(range(lo, hi + 1))


def _enumerate_assignments(
    domains: Iterable[list[int]],
    limits: DynamicRefLimits,
) -> Iterable[tuple[int, ...]]:
    # Domains size has already been checked in _build_domains; this is a thin wrapper.
    return product(*domains)


def infer_dynamic_indirect_targets(
    formula: str,
    *,
    current_sheet: str,
    cell_type_env: CellTypeEnv,
    limits: DynamicRefLimits | None = None,
    bounds: WorkbookBoundsProtocol | None = None,
    named_ranges: Mapping[str, tuple[str, str]] | None = None,
    named_range_ranges: Mapping[str, tuple[str, str, str]] | None = None,
) -> set[str]:
    """Infer the union of all possible INDIRECT targets for a formula."""

    if not isinstance(formula, str) or not formula.startswith("="):
        return set()

    lim = limits or DynamicRefLimits()
    out: set[str] = set()

    calls = _find_function_calls_with_spans(formula, {"INDIRECT"})
    for fn, inner, _span in calls:
        if fn != "INDIRECT":
            continue
        targets = _infer_single_indirect_call(
            inner,
            current_sheet=current_sheet,
            cell_type_env=cell_type_env,
            limits=lim,
            bounds=bounds,
            named_ranges=named_ranges,
            named_range_ranges=named_range_ranges,
        )
        out |= targets
        if len(out) > lim.max_cells:
            raise DynamicRefError(
                f"Dynamic ref cells exceed limit ({len(out)} > {lim.max_cells})"
            )

    return out


def _infer_single_indirect_call(
    inner_args: str,
    *,
    current_sheet: str,
    cell_type_env: CellTypeEnv,
    limits: DynamicRefLimits,
    bounds: WorkbookBoundsProtocol | None,
    named_ranges: Mapping[str, tuple[str, str]] | None = None,
    named_range_ranges: Mapping[str, tuple[str, str, str]] | None = None,
) -> set[str]:
    args = _split_top_level_args(inner_args)
    if args is None or len(args) < 1 or len(args) > 2:
        raise DynamicRefError("INDIRECT expects 1 or 2 arguments")

    nr = named_ranges or {}
    nrr = named_range_ranges or {}
    text_expr = _qualify_fragment(args[0], nr, nrr)
    a1_expr = _qualify_fragment(args[1], nr, nrr) if len(args) == 2 else ""

    text_ast = parse_ast("=" + text_expr)
    a1_ast = parse_ast("=" + a1_expr) if a1_expr else None

    leaf_addrs: set[str] = set()
    leaf_addrs |= _collect_addresses(text_ast)
    if a1_ast is not None:
        leaf_addrs |= _collect_addresses(a1_ast)

    domains = _build_value_domains(leaf_addrs, cell_type_env, limits)

    targets: set[str] = set()

    for assignment in _enumerate_value_assignments(domains.values(), limits):
        addr_to_value = dict(zip(domains.keys(), assignment, strict=False))

        def get_cell_value(addr: str, addr_to_value_map=addr_to_value) -> Any:
            try:
                return addr_to_value_map[addr]
            except KeyError as exc:
                raise DynamicRefError(
                    f"INDIRECT argument formula references cell without domain: {addr!r}"
                ) from exc

        text_value = evaluate_expr(text_ast, get_cell_value=get_cell_value, max_depth=limits.max_depth)
        if isinstance(text_value, Unsupported):
            raise DynamicRefError(f"Unsupported INDIRECT text expression: {text_value.reason or ''}")
        if isinstance(text_value, XlError):
            continue
        if not isinstance(text_value, str):
            raise DynamicRefError(f"INDIRECT text argument must be a string, got {type(text_value).__name__}")

        if a1_ast is None:
            a1_flag = True
        else:
            a1_value = evaluate_expr(a1_ast, get_cell_value=get_cell_value, max_depth=limits.max_depth)
            if isinstance(a1_value, Unsupported):
                raise DynamicRefError(f"Unsupported INDIRECT A1/R1C1 flag expression: {a1_value.reason or ''}")
            if isinstance(a1_value, XlError):
                continue
            if isinstance(a1_value, bool):
                a1_flag = a1_value
            elif isinstance(a1_value, (int, float)):
                a1_flag = bool(a1_value)
            else:
                raise DynamicRefError(
                    f"INDIRECT A1/R1C1 flag must be boolean or numeric, got {type(a1_value).__name__}"
                )

        # Derive per-call bounds so sheet-qualified references are not rejected.
        sheet_for_bounds = _sheet_from_indirect_text(text_value, current_sheet=current_sheet)
        if bounds is None:
            local_bounds = GlobalWorkbookBounds(sheet=sheet_for_bounds)
        else:
            local_bounds = GlobalWorkbookBounds(
                sheet=sheet_for_bounds,
                min_row=bounds.min_row,
                max_row=bounds.max_row,
                min_col=bounds.min_col,
                max_col=bounds.max_col,
            )

        result = indirect_text_to_range(text_value, a1_flag, bounds=local_bounds)
        if isinstance(result, ExcelRange):
            targets |= set(result.cell_addresses())

    return targets


def _sheet_from_indirect_text(text: str, *, current_sheet: str) -> str:
    raw = text.strip()
    if not raw:
        return current_sheet
    if "!" in raw:
        sheet_text, _addr = raw.split("!", 1)
        return sheet_text or current_sheet
    return current_sheet


def _build_value_domains(
    addrs: Iterable[str],
    env: CellTypeEnv,
    limits: DynamicRefLimits,
) -> dict[str, list[Any]]:
    domains: dict[str, list[Any]] = {}
    for addr in addrs:
        ct = env.get(addr)
        if ct is None:
            raise DynamicRefError(f"Missing CellType for {addr!r}")
        values: list[Any]
        if ct.enum is not None:
            values = list(ct.enum.values)
        elif ct.interval is not None:
            values = _interval_to_values(ct.interval, limits)
        else:
            raise DynamicRefError(
                f"CellType for {addr!r} must have an interval or enum domain for INDIRECT analysis"
            )
        if not values:
            raise DynamicRefError(f"Empty domain for {addr!r}")
        domains[addr] = values

    total = 1
    for vs in domains.values():
        total *= len(vs)
        if total > limits.max_branches:
            raise DynamicRefError(
                f"Dynamic ref branches exceed limit ({total} > {limits.max_branches})"
            )
    return domains


def _enumerate_value_assignments(
    domains: Iterable[list[Any]],
    limits: DynamicRefLimits,
) -> Iterable[tuple[Any, ...]]:
    return product(*domains)


def _split_top_level_args(s: str) -> list[str] | None:
    """Minimal top-level argument splitter mirroring parser._split_top_level_args."""

    buf: list[str] = []
    args: list[str] = []
    depth = 0
    in_str = False
    i = 0
    while i < len(s):
        ch = s[i]
        if ch == '"':
            in_str = not in_str
            buf.append(ch)
            i += 1
            continue
        if in_str:
            buf.append(ch)
            i += 1
            continue
        if ch == "(":
            depth += 1
            buf.append(ch)
            i += 1
            continue
        if ch == ")":
            if depth == 0:
                return None
            depth -= 1
            buf.append(ch)
            i += 1
            continue
        if ch == "," and depth == 0:
            args.append("".join(buf).strip())
            buf = []
            i += 1
            continue
        buf.append(ch)
        i += 1
    if in_str or depth != 0:
        return None
    args.append("".join(buf).strip())
    return [a for a in args if a != ""]

