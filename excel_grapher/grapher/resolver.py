from __future__ import annotations

import re
from collections.abc import Callable
from dataclasses import dataclass
from typing import Protocol, cast

import openpyxl
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
)

from excel_grapher.core.addressing import offset_range
from excel_grapher.core.coercions import to_number
from excel_grapher.core.formula_ast import (
    AstNode,
    CellRefNode,
    FormulaParseError,
    FunctionCallNode,
    NumberNode,
    RangeNode,
)
from excel_grapher.core.formula_ast import (
    parse as parse_formula_ast,
)
from excel_grapher.core.types import CellValue, ExcelRange, XlError

from .parser import CellRef


class NameResolver(Protocol):
    def resolve(self, name: str) -> tuple[str, str] | None:
        """
        Return (sheet, A1) for a defined name, or None if unknown/unsupported.
        """


class DictNameResolver:
    def __init__(self, mapping: dict[str, tuple[str, str]]):
        self._mapping = mapping

    def resolve(self, name: str) -> tuple[str, str] | None:
        return self._mapping.get(name)


@dataclass(frozen=True)
class NamedRangeMaps:
    cell_map: dict[str, tuple[str, str]]
    range_map: dict[str, tuple[str, str, str]]


def _sheet_bounds(wb: openpyxl.Workbook) -> dict[str, tuple[int, int]]:
    """Return per-sheet (max_row, max_col) from workbook dimensions."""
    bounds: dict[str, tuple[int, int]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        max_row = getattr(ws, "max_row", None) or 1
        max_col = getattr(ws, "max_column", None) or 1
        if max_row < 1:
            max_row = 1
        if max_col < 1:
            max_col = 1
        bounds[name] = (max_row, max_col)
    return bounds


def _range_node_to_excel_range_bounded(
    node: RangeNode,
    bounds: dict[str, tuple[int, int]],
) -> ExcelRange | None:
    """Convert a RangeNode to ExcelRange, capping to sheet bounds."""
    try:
        sheet_start, coord_start = node.start.split("!", 1)
        sheet_end, coord_end = node.end.split("!", 1)
    except ValueError:
        return None
    if sheet_start != sheet_end:
        return None
    sheet = sheet_start.strip("'")
    try:
        col_letter1, row1 = coordinate_from_string(coord_start)
        col_letter2, row2 = coordinate_from_string(coord_end)
    except Exception:
        return None
    start_row = min(row1, row2)
    end_row = max(row1, row2)
    start_col = min(column_index_from_string(col_letter1), column_index_from_string(col_letter2))
    end_col = max(column_index_from_string(col_letter1), column_index_from_string(col_letter2))
    max_r, max_c = bounds.get(sheet, (1048576, 16384))
    end_row = min(end_row, max_r)
    end_col = min(end_col, max_c)
    start_row = max(1, start_row)
    start_col = max(1, start_col)
    return ExcelRange(
        sheet=sheet,
        start_row=start_row,
        start_col=start_col,
        end_row=end_row,
        end_col=end_col,
    )


def _base_node_to_excel_range(
    node: CellRefNode | RangeNode,
    bounds: dict[str, tuple[int, int]],
) -> ExcelRange | None:
    """Interpret base argument of OFFSET as an ExcelRange."""
    if isinstance(node, CellRefNode):
        try:
            sheet, coord = node.address.split("!", 1)
            sheet = sheet.strip("'")
            col_letter, row = coordinate_from_string(coord)
            col = column_index_from_string(col_letter)
            return ExcelRange(
                sheet=sheet,
                start_row=row,
                start_col=col,
                end_row=row,
                end_col=col,
            )
        except Exception:
            return None
    if isinstance(node, RangeNode):
        return _range_node_to_excel_range_bounded(node, bounds)
    return None


def _eval_number_for_defined_name(
    node: AstNode,
    get_cell_value: Callable[[str], CellValue],
    bounds: dict[str, tuple[int, int]],
) -> int | float | None:
    """Evaluate an AST node to a number for OFFSET args (rows, cols, height, width)."""
    if isinstance(node, NumberNode):
        return int(node.value) if node.value == int(node.value) else node.value
    if isinstance(node, CellRefNode):
        val = get_cell_value(node.address)
        n = to_number(val)
        if isinstance(n, XlError):
            return None
        return int(n) if n == int(n) else float(n)
    if isinstance(node, FunctionCallNode) and node.name.upper() == "COUNTA" and len(node.args) == 1:
        rng: ExcelRange | None = None
        if isinstance(node.args[0], RangeNode):
            rng = _range_node_to_excel_range_bounded(node.args[0], bounds)
        elif isinstance(node.args[0], CellRefNode):
            rng = _base_node_to_excel_range(node.args[0], bounds)
        if rng is None:
            return None
        count = 0
        for addr in rng.cell_addresses():
            v = get_cell_value(addr)
            if v is not None and v != "":
                count += 1
        return count
    return None


def _eval_offset_formula_to_range(
    node: FunctionCallNode,
    get_cell_value: Callable[[str], CellValue],
    bounds: dict[str, tuple[int, int]],
) -> tuple[str, str, str] | None:
    """Evaluate OFFSET(...) to (sheet, start_a1, end_a1) or None."""
    if node.name.upper() != "OFFSET" or len(node.args) < 3:
        return None
    base = _base_node_to_excel_range(node.args[0], bounds) if isinstance(node.args[0], (CellRefNode, RangeNode)) else None
    if base is None:
        return None
    rows = _eval_number_for_defined_name(node.args[1], get_cell_value, bounds)
    cols = _eval_number_for_defined_name(node.args[2], get_cell_value, bounds)
    if rows is None or cols is None:
        return None
    height = _eval_number_for_defined_name(node.args[3], get_cell_value, bounds) if len(node.args) >= 4 else None
    width = _eval_number_for_defined_name(node.args[4], get_cell_value, bounds) if len(node.args) >= 5 else None
    if height is not None and height <= 0:
        return None
    if width is not None and width <= 0:
        return None
    max_r, max_c = bounds.get(base.sheet, (1048576, 16384))

    class _Bounds:
        sheet = base.sheet
        min_row = 1
        min_col = 1
        max_row = max_r
        max_col = max_c

    result = offset_range(
        base,
        rows,
        cols,
        height,
        width,
        bounds=_Bounds(),
    )
    if isinstance(result, XlError):
        return None
    start_a1 = f"{get_column_letter(result.start_col)}{result.start_row}"
    end_a1 = f"{get_column_letter(result.end_col)}{result.end_row}"
    return (result.sheet, start_a1, end_a1)


def _eval_indirect_formula_to_range(
    node: FunctionCallNode,
    get_cell_value: Callable[[str], CellValue],
    bounds: dict[str, tuple[int, int]],
) -> tuple[str, str] | tuple[str, str, str] | None:
    """Evaluate INDIRECT(...) to (sheet, a1) or (sheet, start_a1, end_a1)."""
    from excel_grapher.core.formula_ast import StringNode

    if node.name.upper() != "INDIRECT" or len(node.args) < 1:
        return None
    if not isinstance(node.args[0], StringNode):
        return None
    text = node.args[0].value.strip()
    if "!" in text:
        sheet_part, addr_part = text.split("!", 1)
        sheet = sheet_part.strip("'")
    else:
        sheet = next(iter(bounds.keys()), "Sheet1")
        addr_part = text
    if ":" in addr_part:
        start_ref, end_ref = addr_part.split(":", 1)
        try:
            c1, r1 = coordinate_from_string(start_ref)
            c2, r2 = coordinate_from_string(end_ref)
            start_a1 = f"{c1}{r1}"
            end_a1 = f"{c2}{r2}"
            return (sheet, start_a1, end_a1)
        except Exception:
            return None
    try:
        c, r = coordinate_from_string(addr_part)
        a1 = f"{c}{r}"
        return (sheet, a1)
    except Exception:
        return None


def _normalize_formula_for_parse(formula: str, bounds: dict[str, tuple[int, int]]) -> str:
    """Strip $ and expand whole-column/whole-row refs so formula_ast can parse."""
    s = formula.replace("$", "")
    for sheet, (max_r, max_c) in bounds.items():
        col_letter = get_column_letter(max_c)
        s = re.sub(
            re.escape(sheet) + r"!\s*([A-Z]+)\s*:\s*\1\b",
            f"{sheet}!\\g<1>1:\\g<1>{max_r}",
            s,
            flags=re.IGNORECASE,
        )
        s = re.sub(
            re.escape(sheet) + r"!\s*(\d+)\s*:\s*\1\b",
            f"{sheet}!A\\g<1>:{col_letter}\\g<1>",
            s,
        )
    return s


def _try_resolve_formula_defined_name(
    attr_text: str,
    wb: openpyxl.Workbook,
) -> tuple[str, str, str] | tuple[str, str] | None:
    """If attr_text is an OFFSET/INDIRECT formula, evaluate to range or cell; else None."""
    formula = attr_text.strip()
    if not formula.upper().startswith("OFFSET(") and not formula.upper().startswith("INDIRECT("):
        return None
    if not formula.startswith("="):
        formula = "=" + formula
    bounds = _sheet_bounds(wb)
    formula = _normalize_formula_for_parse(formula, bounds)
    try:
        ast = parse_formula_ast(formula)
    except FormulaParseError:
        return None
    if not isinstance(ast, FunctionCallNode):
        return None

    def get_cell_value(addr: str) -> CellValue:
        try:
            sheet_part, a1 = addr.split("!", 1)
            sheet = sheet_part.strip("'")
            if sheet in wb.sheetnames:
                return wb[sheet][a1].value
        except Exception:
            pass
        return None

    if ast.name.upper() == "OFFSET":
        return _eval_offset_formula_to_range(ast, get_cell_value, bounds)
    if ast.name.upper() == "INDIRECT":
        return _eval_indirect_formula_to_range(ast, get_cell_value, bounds)
    return None


def build_named_range_map(wb: openpyxl.Workbook) -> NamedRangeMaps:
    """
    Map defined names to single-cell and range references.

    Only includes simple definitions like Sheet1!$A$1 or Sheet1!$A$1:$B$10
    (optionally quoted sheet name). Skips multi-area and complex formulas.
    Formula-based names (OFFSET, INDIRECT) are evaluated using workbook values.
    """
    cell_map: dict[str, tuple[str, str]] = {}
    range_map: dict[str, tuple[str, str, str]] = {}
    for name, defn in wb.defined_names.items():
        attr_text = getattr(defn, "attr_text", None)
        if not isinstance(attr_text, str) or not attr_text:
            continue
        if attr_text.startswith("{") or attr_text.startswith("#") or attr_text.startswith('"'):
            continue
        if attr_text.strip().upper().startswith(("OFFSET(", "INDIRECT(")):
            resolved = _try_resolve_formula_defined_name(attr_text, wb)
            if resolved is not None:
                if len(resolved) == 2:
                    cell_map[str(name)] = (resolved[0], resolved[1])
                elif len(resolved) == 3:
                    range_map[str(name)] = cast(tuple[str, str, str], resolved)
                continue
        if "," in attr_text:
            continue
        if ":" in attr_text:
            m = re.match(
                r"'?(?P<sheet>[^'!]+)'?!\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d+):\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d+)$",
                attr_text,
            )
            if not m:
                resolved = _try_resolve_formula_defined_name(attr_text, wb)
                if resolved is not None:
                    if len(resolved) == 2:
                        cell_map[str(name)] = (resolved[0], resolved[1])
                    elif len(resolved) == 3:
                        range_map[str(name)] = cast(tuple[str, str, str], resolved)
                continue
            sheet_name = m.group("sheet")
            start = f"{m.group('c1')}{m.group('r1')}"
            end = f"{m.group('c2')}{m.group('r2')}"
            range_map[str(name)] = (sheet_name, start, end)
            continue

        m = re.match(r"'?([^'!]+)'?!\$?([A-Z]{1,3})\$?(\d+)$", attr_text)
        if not m:
            continue
        sheet_name = m.group(1)
        col = m.group(2)
        row = m.group(3)
        cell_map[str(name)] = (sheet_name, f"{col}{row}")
    return NamedRangeMaps(cell_map=cell_map, range_map=range_map)


def qualify_cell_ref(ref: CellRef, current_sheet: str) -> tuple[str, str]:
    sheet = ref.sheet if ref.sheet is not None else current_sheet
    return sheet, f"{ref.column}{ref.row}"

