"""Code generator for converting Excel formulas to Python code."""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from typing import TYPE_CHECKING, AbstractSet, Any, Protocol, TypedDict

import openpyxl.utils.cell

from excel_grapher.grapher.graph import CycleError

from .export_runtime.embed import emit_runtime
from .name_utils import (
    address_to_python_name,
    excel_func_to_python,
    normalize_address,
    parse_address,
    quote_sheet_if_needed,
)
from .parser import (
    AstNode,
    BinaryOpNode,
    BoolNode,
    CellRefNode,
    EmptyArgNode,
    ErrorNode,
    FunctionCallNode,
    NumberNode,
    RangeNode,
    StringNode,
    UnaryOpNode,
    parse,
)
from .types import XlError

if TYPE_CHECKING:
    from excel_grapher.grapher import DependencyGraph  # noqa: F401


class GraphNode(Protocol):
    formula: str | None
    normalized_formula: str | None
    value: object | None


class GraphLike(Protocol):
    def get_node(self, address: str) -> GraphNode | None:
        ...

    def leaf_keys(self) -> list[str]:
        ...

    def formula_keys(self) -> list[str]:
        ...

    def dependencies(self, address: str) -> list[str]:
        ...

    leaf_classification: dict[str, str] | None


class GenerationParts(TypedDict):
    runtime_code: str
    inputs_block_lines: list[str]
    constants_block_lines: list[str]
    cell_code_lines: list[str]
    formula_cells: list[str]
    all_cells: list[str]
    needs_offset_table: bool
    targets: list[str]
    has_constants: bool
    used_xl_functions: frozenset[str]

# Operators that need wrapper functions for Excel semantics (error propagation)
_BINARY_OPS = {
    "+": "xl_add",
    "-": "xl_sub",
    "*": "xl_mul",
    "/": "xl_div",
    "^": "xl_pow",
    "=": "xl_eq",
    "<>": "xl_ne",
    "<": "xl_lt",
    ">": "xl_gt",
    "<=": "xl_le",
    ">=": "xl_ge",
}

# Unary operators that need wrapper functions
_UNARY_OPS = {
    "-": "xl_neg",
    "+": "xl_pos",
    "%": "xl_percent",
}


class CodeGenerator:
    """Generates Python code from Excel formulas."""

    def __init__(
        self,
        graph: DependencyGraph | GraphLike,
        *,
        iterate_enabled: bool | None = None,
        iterate_count: int = 100,
        iterate_delta: float = 0.001,
    ) -> None:
        """Initialize the code generator.

        Args:
            graph: Dependency graph from excel_grapher containing cell formulas.
            iterate_enabled: If True, :meth:`DependencyGraph.evaluation_order` rejects
                any must- or may-cycle (workbook iterative calc is unsupported in codegen).
                Typically set from :func:`excel_grapher.get_calc_settings`. ``None`` skips
                this check (default).
        """
        self.graph = graph
        self._iterate_enabled = iterate_enabled
        self._iterate_count = iterate_count
        self._iterate_delta = iterate_delta
        self._emitted: set[str] = set()
        self._needs_offset_runtime = False  # Set to True if dynamic OFFSET is used
        self._offset_runtime_sheets: set[str] = set()
        self._temp_var_counter = 0  # Counter for unique temp variable names
        self._ast_cache: dict[str, AstNode] = {}
        self._used_graph_closure: bool = False

    @staticmethod
    def _normalize_entrypoint_name(name: str) -> str:
        return CodeGenerator._normalize_package_name(name)

    def _normalize_entrypoints(
        self, entrypoints: Mapping[str, Sequence[str]] | None
    ) -> dict[str, list[str]]:
        if entrypoints is None:
            return {}
        if not isinstance(entrypoints, Mapping):
            raise TypeError("entrypoints must be a mapping of name to target list")
        if not entrypoints:
            raise ValueError("entrypoints must not be empty")
        normalized: dict[str, list[str]] = {}
        seen_names: dict[str, str] = {}
        for name, targets in entrypoints.items():
            if not isinstance(name, str):
                raise TypeError("entrypoints keys must be strings")
            if isinstance(targets, (str, bytes)) or not isinstance(targets, Sequence):
                raise TypeError("entrypoints values must be sequences of targets")
            if not targets:
                raise ValueError("entrypoints target lists must not be empty")
            normalized_name = self._normalize_entrypoint_name(name)
            if normalized_name in normalized:
                original = seen_names[normalized_name]
                raise ValueError(
                    "Entrypoint names normalize to the same identifier: "
                    f"{original!r} and {name!r}"
                )
            normalized_targets = [normalize_address(t) for t in targets]
            normalized[normalized_name] = normalized_targets
            seen_names[normalized_name] = name
        return normalized

    def _get_or_parse_ast(self, address: str) -> AstNode | None:
        """Parse and cache the AST for a formula cell.

        The cache key is the normalized, sheet-qualified cell address. The cache
        is cleared at the start of each generate() call.
        """
        normalized = normalize_address(address)
        if normalized in self._ast_cache:
            return self._ast_cache[normalized]

        node = self.graph.get_node(normalized)
        if node is None or node.formula is None:
            return None

        formula = node.normalized_formula or node.formula
        ast = parse(formula)
        self._ast_cache[normalized] = ast
        return ast

    def _emit_ast(self, node: AstNode) -> str:
        """Convert an AST node to a Python expression string.

        Args:
            node: AST node to convert.

        Returns:
            Python expression as a string.
        """
        if isinstance(node, EmptyArgNode):
            return "None"

        if isinstance(node, NumberNode):
            return repr(node.value)

        if isinstance(node, StringNode):
            return repr(node.value)

        if isinstance(node, BoolNode):
            return "True" if node.value else "False"

        if isinstance(node, ErrorNode):
            return f"XlError.{node.error.name}"

        if isinstance(node, CellRefNode):
            return self._emit_cell_eval(node.address)

        if isinstance(node, RangeNode):
            return self._emit_range(node)

        if isinstance(node, BinaryOpNode):
            return self._emit_binary_op(node)

        if isinstance(node, UnaryOpNode):
            return self._emit_unary_op(node)

        if isinstance(node, FunctionCallNode):
            return self._emit_function_call(node)

        raise ValueError(f"Unknown AST node type: {type(node)}")

    def _emit_range(self, node: RangeNode) -> str:
        """Emit a range as a 2D nested list of cell evaluations.

        The outer list contains rows, inner lists contain columns.
        For A1:B3, emits: [[xl_eval(ctx, "S!A1", cell_s_a1), xl_eval(ctx, "S!B1", ...)], ...]
        """
        rows = self._range_addresses_2d(node.start, node.end)
        row_strs = []
        for row_addrs in rows:
            cell_calls = [self._emit_cell_eval(addr) for addr in row_addrs]
            row_strs.append("[" + ", ".join(cell_calls) + "]")
        # Model ranges as object-dtype ndarrays so they fit `CellValue` and work
        # with runtime helpers like `flatten(*args)`.
        return f"np.array([{', '.join(row_strs)}], dtype=object)"

    def _emit_cell_eval(self, address: str) -> str:
        normalized = normalize_address(address)
        if self.graph is None:
            return f"xl_cell(ctx, {repr(normalized)})"
        node = self.graph.get_node(normalized)
        if node is not None and node.formula is not None:
            func_name = address_to_python_name(normalized)
            return f"xl_eval(ctx, {repr(normalized)}, {func_name})"
        return f"xl_cell(ctx, {repr(normalized)})"

    @staticmethod
    def _py_literal(value: Any) -> str:
        """Convert a Python value into a safe Python literal expression.

        The generated code must be syntactically valid Python. Values pulled from
        workbooks can include objects (e.g., openpyxl ArrayFormula) whose repr()
        is not a literal and would break the generated file if embedded.
        """
        if value is None:
            return "0"
        if isinstance(value, XlError):
            return f"XlError.{value.name}"
        if isinstance(value, (bool, int, float, str)):
            return repr(value)
        # Numpy scalars may appear; keep the runtime surface small by emitting their
        # native Python equivalent when available.
        if hasattr(value, "item"):
            try:
                return CodeGenerator._py_literal(value.item())
            except Exception:
                return "0"
        return "0"

    def _range_addresses_2d(self, start: str, end: str) -> list[list[str]]:
        """Generate all cell addresses in a range as a 2D list (rows x cols)."""
        start_sheet, start_cell = self._parse_address(start)
        end_sheet, end_cell = self._parse_address(end)

        # Use start sheet for all cells (Excel semantics)
        sheet = start_sheet

        start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
        end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)

        start_col_idx = openpyxl.utils.cell.column_index_from_string(start_col)
        end_col_idx = openpyxl.utils.cell.column_index_from_string(end_col)

        rows = []
        for row in range(start_row, end_row + 1):
            row_addrs = []
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
                row_addrs.append(f"{sheet}!{col_letter}{row}")
            rows.append(row_addrs)

        return rows

    def _range_addresses(self, start: str, end: str) -> list[str]:
        """Generate all cell addresses in a range (flat list)."""
        rows = self._range_addresses_2d(start, end)
        return [addr for row in rows for addr in row]

    @staticmethod
    def _format_cell_address(sheet: str, row: int, col: int) -> str:
        sheet_name = quote_sheet_if_needed(sheet)
        col_letter = openpyxl.utils.cell.get_column_letter(col)
        return f"{sheet_name}!{col_letter}{row}"

    def _targets_to_entries(self, targets: Sequence[str]) -> list[tuple[str, str]]:
        by_sheet: dict[str, list[tuple[int, int]]] = {}
        for address in targets:
            sheet, cell = parse_address(address)
            col_letters, row = openpyxl.utils.cell.coordinate_from_string(cell)
            col_idx = openpyxl.utils.cell.column_index_from_string(col_letters)
            by_sheet.setdefault(sheet, []).append((row, col_idx))

        entries: list[tuple[str, str]] = []

        for sheet, cells in by_sheet.items():
            cells_set = set(cells)
            if not cells_set:
                continue

            row_groups: dict[int, list[int]] = {}
            col_groups: dict[int, list[int]] = {}
            for row, col in cells_set:
                row_groups.setdefault(row, []).append(col)
                col_groups.setdefault(col, []).append(row)

            row_entries: list[tuple[str, str]] = []
            for row, cols in row_groups.items():
                cols = sorted(cols)
                start = prev = cols[0]
                for col in cols[1:]:
                    if col == prev + 1:
                        prev = col
                        continue
                    start_addr = self._format_cell_address(sheet, row, start)
                    end_addr = self._format_cell_address(sheet, row, prev)
                    if start == prev:
                        row_entries.append((start_addr, "xl_cell"))
                    else:
                        row_entries.append((f"{start_addr}:{end_addr}", "xl_range"))
                    start = prev = col
                start_addr = self._format_cell_address(sheet, row, start)
                end_addr = self._format_cell_address(sheet, row, prev)
                if start == prev:
                    row_entries.append((start_addr, "xl_cell"))
                else:
                    row_entries.append((f"{start_addr}:{end_addr}", "xl_range"))

            col_entries: list[tuple[str, str]] = []
            for col, rows in col_groups.items():
                rows = sorted(rows)
                start = prev = rows[0]
                for row in rows[1:]:
                    if row == prev + 1:
                        prev = row
                        continue
                    start_addr = self._format_cell_address(sheet, start, col)
                    end_addr = self._format_cell_address(sheet, prev, col)
                    if start == prev:
                        col_entries.append((start_addr, "xl_cell"))
                    else:
                        col_entries.append((f"{start_addr}:{end_addr}", "xl_range"))
                    start = prev = row
                start_addr = self._format_cell_address(sheet, start, col)
                end_addr = self._format_cell_address(sheet, prev, col)
                if start == prev:
                    col_entries.append((start_addr, "xl_cell"))
                else:
                    col_entries.append((f"{start_addr}:{end_addr}", "xl_range"))

            entries.extend(row_entries if len(row_entries) <= len(col_entries) else col_entries)

        entries.sort(key=lambda item: item[0])
        return entries

    @staticmethod
    def _emit_resolver_lines() -> list[str]:
        return [
            "# --- Formula resolver ---",
            "_RESOLVED_FORMULAS = {}",
            "def _address_to_func_name(address):",
            "    name = []",
            "    prev_underscore = False",
            "    for ch in address.lower():",
            "        if ch == \"'\":",
            "            continue",
            "        if \"a\" <= ch <= \"z\" or \"0\" <= ch <= \"9\":",
            "            name.append(ch)",
            "            prev_underscore = False",
            "        else:",
            "            if not prev_underscore:",
            "                name.append(\"_\")",
            "                prev_underscore = True",
            "    base = \"\".join(name).strip(\"_\")",
            "    return f\"cell_{base}\"",
            "",
            "def _resolve_formula(address):",
            "    fn = _RESOLVED_FORMULAS.get(address)",
            "    if fn is not None:",
            "        return fn",
            "    name = _address_to_func_name(address)",
            "    fn = globals().get(name)",
            "    if fn is not None:",
            "        _RESOLVED_FORMULAS[address] = fn",
            "    return fn",
            "",
        ]

    @staticmethod
    def _internals_runtime_import_names(
        used_xl_functions: AbstractSet[str], cell_code_lines: list[str]
    ) -> list[str]:
        """Names from the embedded runtime that formula cell bodies reference as globals."""
        blob = "\n".join(cell_code_lines)
        if "def " not in blob:
            return []
        names = set(used_xl_functions)
        names.discard("numpy")
        names.update({"xl_cell", "xl_eval"})
        if "numpy" in used_xl_functions or "np." in blob or "np.array" in blob:
            names.add("np")
        if "XlError" in blob:
            names.add("XlError")
        if "ExcelRange(" in blob:
            names.add("ExcelRange")
        return sorted(names)

    @staticmethod
    def _format_from_runtime_import(names: list[str]) -> str:
        if not names:
            return ""
        joined = ", ".join(names)
        prefix = "from .runtime import "
        if len(prefix) + len(joined) <= 88:
            return prefix + joined
        inner = ",\n    ".join(names)
        return f"{prefix}(\n    {inner},\n)"

    def _parse_address(self, address: str) -> tuple[str, str]:
        """Parse a sheet-qualified address into (quoted_sheet, cell) tuple.

        The sheet name is returned with quotes if needed for address construction.
        """
        sheet, cell = parse_address(address)
        return quote_sheet_if_needed(sheet), cell

    def _get_graph_leaf_classification(self) -> dict[str, str] | None:
        mapping = getattr(self.graph, "leaf_classification", None)
        if mapping is None:
            return None
        if not isinstance(mapping, Mapping):
            raise TypeError("leaf_classification must be a mapping of address to label")
        normalized: dict[str, str] = {}
        for key, value in mapping.items():
            if not isinstance(key, str):
                raise TypeError("leaf_classification keys must be strings")
            if value not in {"input", "constant"}:
                raise ValueError(
                    "leaf_classification values must be 'input' or 'constant'"
                )
            normalized[normalize_address(key)] = value
        return normalized

    def _collect_needed_leaves(self, all_cells: list[str]) -> set[str]:
        # Only emit leaf inputs that are actually needed for the target dependency closure.
        # This keeps generated output small and avoids embedding unrelated workbook artifacts.
        needed_leaves: set[str] = set()
        for addr in all_cells:
            node = self.graph.get_node(addr)
            if node is None or node.formula is not None:
                continue
            needed_leaves.add(normalize_address(addr))
        return needed_leaves

    @staticmethod
    def _normalize_constant_types(constant_types: set[str] | None) -> set[str]:
        if constant_types is None:
            return set()
        if isinstance(constant_types, (str, bytes)):
            raise TypeError("constant_types must be a set of strings")
        normalized = {str(item) for item in constant_types}
        allowed = {"number", "string"}
        invalid = normalized - allowed
        if invalid:
            raise ValueError(f"Unsupported constant_types: {sorted(invalid)!r}")
        return normalized

    @classmethod
    def _classification_from_graph(
        cls, graph_classification: dict[str, str] | None, needed_leaves: set[str]
    ) -> tuple[set[str], set[str]]:
        if graph_classification is None:
            return set(needed_leaves), set()
        constants = {
            addr
            for addr in needed_leaves
            if graph_classification.get(addr) == "constant"
        }
        inputs = set(needed_leaves) - constants
        return inputs, constants

    @staticmethod
    def _parse_constant_range(range_str: str) -> tuple[str, int, int, int, int]:
        if not isinstance(range_str, str):
            raise TypeError("constant_ranges entries must be strings")
        if "!" not in range_str:
            raise ValueError(f"Range must be sheet-qualified: {range_str}")
        sheet_part, cell_part = range_str.rsplit("!", 1)
        if ":" in cell_part:
            start_cell, end_cell = cell_part.split(":", 1)
        else:
            start_cell = end_cell = cell_part

        sheet, start = parse_address(f"{sheet_part}!{start_cell}")
        _, end = parse_address(f"{sheet_part}!{end_cell}")

        start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start)
        end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end)
        start_col_idx = openpyxl.utils.cell.column_index_from_string(start_col)
        end_col_idx = openpyxl.utils.cell.column_index_from_string(end_col)

        r1, r2 = (start_row, end_row) if start_row <= end_row else (end_row, start_row)
        c1, c2 = (
            (start_col_idx, end_col_idx)
            if start_col_idx <= end_col_idx
            else (end_col_idx, start_col_idx)
        )
        return (sheet, r1, c1, r2, c2)

    @classmethod
    def _normalize_constant_ranges(
        cls, constant_ranges: Sequence[str] | None
    ) -> list[tuple[str, int, int, int, int]]:
        if constant_ranges is None:
            return []
        if isinstance(constant_ranges, (str, bytes)):
            raise TypeError("constant_ranges must be a sequence of strings")
        return [cls._parse_constant_range(item) for item in constant_ranges]

    @classmethod
    def _normalize_input_ranges(
        cls, input_ranges: Sequence[str] | None
    ) -> list[tuple[str, int, int, int, int]]:
        if input_ranges is None:
            return []
        if isinstance(input_ranges, (str, bytes)):
            raise TypeError("input_ranges must be a sequence of strings")
        return [cls._parse_constant_range(item) for item in input_ranges]

    @staticmethod
    def _apply_input_ranges_override(
        needed_leaves: set[str],
        constants: set[str],
        input_ranges: list[tuple[str, int, int, int, int]],
    ) -> tuple[set[str], set[str]]:
        """Drop constants that fall in input_ranges; input ranges win over constant rules."""
        if not input_ranges:
            return set(needed_leaves) - constants, constants
        constants = set(constants)
        for key in needed_leaves:
            if CodeGenerator._leaf_in_constant_ranges(key, input_ranges):
                constants.discard(key)
        inputs = set(needed_leaves) - constants
        return inputs, constants

    @staticmethod
    def _leaf_value_matches_constant_type(
        value: object | None, constant_types: set[str]
    ) -> bool:
        if not constant_types:
            return False
        if value is None:
            value = 0
        if isinstance(value, bool):
            return False
        return ("number" in constant_types and isinstance(value, (int, float))) or (
            "string" in constant_types and isinstance(value, str)
        )

    @staticmethod
    def _leaf_in_constant_ranges(
        address: str, constant_ranges: list[tuple[str, int, int, int, int]]
    ) -> bool:
        if not constant_ranges:
            return False
        sheet, cell = parse_address(normalize_address(address))
        col_str, row = openpyxl.utils.cell.coordinate_from_string(cell)
        col = openpyxl.utils.cell.column_index_from_string(col_str)
        for range_sheet, r1, c1, r2, c2 in constant_ranges:
            if sheet != range_sheet:
                continue
            if r1 <= row <= r2 and c1 <= col <= c2:
                return True
        return False

    def classify_leaf_nodes(
        self,
        targets: list[str],
        *,
        constant_types: set[str] | None = None,
        constant_ranges: Sequence[str] | None = None,
        constant_blanks: bool = False,
        input_ranges: Sequence[str] | None = None,
        attach_to_graph: bool = False,
    ) -> tuple[set[str], set[str]]:
        normalized_targets = [normalize_address(t) for t in targets]
        all_cells = self._collect_all_cells(normalized_targets)
        needed_leaves = self._collect_needed_leaves(all_cells)

        normalized_constant_types = self._normalize_constant_types(constant_types)
        normalized_constant_ranges = self._normalize_constant_ranges(constant_ranges)
        normalized_input_ranges = self._normalize_input_ranges(input_ranges)
        explicit_constant_rules = bool(
            constant_types or constant_ranges or constant_blanks
        )
        use_graph_classification = not explicit_constant_rules and not input_ranges

        if use_graph_classification:
            graph_classification = self._get_graph_leaf_classification()
            inputs, constants = self._classification_from_graph(
                graph_classification, needed_leaves
            )
        elif explicit_constant_rules:
            inputs, constants = self._classify_leaf_nodes(
                needed_leaves,
                constant_types=normalized_constant_types,
                constant_ranges=normalized_constant_ranges,
                constant_blanks=constant_blanks,
                input_ranges=normalized_input_ranges,
            )
        else:
            graph_classification = self._get_graph_leaf_classification()
            inputs, constants = self._classification_from_graph(
                graph_classification, needed_leaves
            )
            inputs, constants = self._apply_input_ranges_override(
                needed_leaves, constants, normalized_input_ranges
            )

        if attach_to_graph:
            classification = {addr: "input" for addr in inputs}
            classification.update({addr: "constant" for addr in constants})
            self.graph.leaf_classification = classification  # type: ignore[assignment]

        return inputs, constants

    def _classify_leaf_nodes(
        self,
        needed_leaves: set[str],
        *,
        constant_types: set[str],
        constant_ranges: list[tuple[str, int, int, int, int]],
        constant_blanks: bool,
        input_ranges: list[tuple[str, int, int, int, int]] | None = None,
    ) -> tuple[set[str], set[str]]:
        input_ranges = input_ranges or []
        constants: set[str] = set()
        for key in needed_leaves:
            if self._leaf_in_constant_ranges(key, constant_ranges):
                constants.add(key)
                continue
            node = self.graph.get_node(key)
            value = None if node is None else node.value
            if constant_blanks and value is None:
                constants.add(key)
                continue
            if self._leaf_value_matches_constant_type(value, constant_types):
                constants.add(key)
        return self._apply_input_ranges_override(
            needed_leaves, constants, input_ranges
        )

    def _emit_binary_op(self, node: BinaryOpNode) -> str:
        """Emit a binary operation."""
        left = self._emit_ast(node.left)
        right = self._emit_ast(node.right)
        op = node.op

        # Concatenation: & -> xl_concat
        if op == "&":
            return f"xl_concat({left}, {right})"

        # All other operators use wrapper functions for error propagation
        if op in _BINARY_OPS:
            func = _BINARY_OPS[op]
            return f"{func}({left}, {right})"

        raise ValueError(f"Unknown operator: {op}")

    def _emit_unary_op(self, node: UnaryOpNode) -> str:
        """Emit a unary operation."""
        operand = self._emit_ast(node.operand)
        op = node.op
        if op in _UNARY_OPS:
            func = _UNARY_OPS[op]
            return f"{func}({operand})"
        raise ValueError(f"Unknown unary operator: {op}")

    def _emit_function_call(self, node: FunctionCallNode) -> str:
        """Emit a function call.

        For functions that need numpy arrays (LOOKUP, VLOOKUP, HLOOKUP, INDEX,
        MATCH, SUMPRODUCT), range arguments are wrapped with np.array().
        IF, OFFSET are handled specially.
        """
        func_name = excel_func_to_python(node.name)
        upper_name = node.name.upper()

        # IF needs special handling - emit as Python conditional for lazy evaluation
        if upper_name == "IF":
            return self._emit_if(node)

        # IFERROR needs special handling - only evaluate fallback if value is error
        if upper_name == "IFERROR":
            return self._emit_iferror(node)

        # IFNA needs special handling - only evaluate fallback if value is #N/A
        if upper_name in {"IFNA", "_XLFN.IFNA"}:
            return self._emit_ifna(node)

        # CHOOSE needs special handling - only evaluate the selected argument
        if upper_name == "CHOOSE":
            return self._emit_choose(node)

        # OFFSET needs special handling - try static resolution first
        if upper_name == "OFFSET":
            return self._emit_offset(node)

        # ROW needs special handling - references should not be evaluated
        if upper_name == "ROW":
            return self._emit_row(node)

        # COLUMN needs special handling - references should not be evaluated
        if upper_name == "COLUMN":
            return self._emit_column(node)

        # COLUMNS needs special handling - references should not be evaluated
        if upper_name == "COLUMNS":
            return self._emit_columns(node)

        # Functions that need numpy arrays for their array/table arguments
        # Maps function name -> set of argument indices that need np.array wrapping
        numpy_array_args: dict[str, set[int]] = {
            "LOOKUP": {1, 2},  # lookup_vector/array + optional result_vector
            "VLOOKUP": {1},  # table_array is 2nd arg (index 1)
            "HLOOKUP": {1},  # table_array is 2nd arg (index 1)
            "INDEX": {0},  # array is 1st arg (index 0)
            "MATCH": {1},  # lookup_array is 2nd arg (index 1)
            "SUMPRODUCT": set(range(10)),  # all args can be arrays
        }

        needs_numpy_wrap = numpy_array_args.get(upper_name, set())

        emitted_args = []
        for i, arg in enumerate(node.args):
            emitted = self._emit_ast(arg)
            # Wrap range arguments with np.array() for functions that need it
            # Use dtype=object to preserve original Python types (mixed str/int/float)
            if i in needs_numpy_wrap and isinstance(arg, RangeNode):
                emitted = f"np.array({emitted}, dtype=object)"
            emitted_args.append(emitted)

        args = ", ".join(emitted_args)
        return f"{func_name}({args})"

    def _next_temp_var(self) -> str:
        """Generate a unique temporary variable name."""
        self._temp_var_counter += 1
        return f"_t{self._temp_var_counter}"

    def _emit_iferror(self, node: FunctionCallNode) -> str:
        """Emit IFERROR as a Python conditional for lazy evaluation.

        IFERROR(value, value_if_error)

        Emits as: (value_if_error if isinstance(value, XlError) else value)
        The value is evaluated once and stored, then checked for errors.
        The value_if_error is only evaluated if value is an error.
        """
        if len(node.args) < 2:
            return "XlError.VALUE"

        value_expr = self._emit_ast(node.args[0])
        error_expr = self._emit_ast(node.args[1])

        # Store value in uniquely-named temp var to avoid evaluating twice
        var = self._next_temp_var()
        return f"(({error_expr}) if isinstance(({var} := {value_expr}), XlError) else {var})"

    def _emit_ifna(self, node: FunctionCallNode) -> str:
        """Emit IFNA as a Python conditional for lazy evaluation.

        IFNA(value, value_if_na)

        Emits as: (value_if_na if value is #N/A else value)
        The value is evaluated once and stored, then checked for #N/A.
        The value_if_na is only evaluated if value is #N/A.
        """
        if len(node.args) < 2:
            return "XlError.VALUE"

        value_expr = self._emit_ast(node.args[0])
        na_expr = self._emit_ast(node.args[1])

        var = self._next_temp_var()
        return f"(({na_expr}) if (({var} := {value_expr}) == XlError.NA) else {var})"

    def _emit_if(self, node: FunctionCallNode) -> str:
        """Emit IF as a Python conditional expression for lazy evaluation.

        IF(condition, true_val, [false_val])

        Emits as a nested conditional that:
        1. Returns error if condition is an error
        2. Otherwise lazily evaluates only the relevant branch

        This ensures only the relevant branch is evaluated, which is critical
        for breaking circular references that Excel handles via lazy evaluation.
        """
        if len(node.args) < 2:
            return "XlError.VALUE"

        cond_expr = self._emit_ast(node.args[0])
        true_expr = self._emit_ast(node.args[1])
        false_expr = self._emit_ast(node.args[2]) if len(node.args) > 2 else "False"

        # Excel-style boolean coercion is not Python truthiness:
        # - "FALSE" should behave like False
        # - "0" should produce #VALUE! (per to_bool)
        # We must coerce via to_bool(), and keep lazy branch evaluation.
        cond_var = self._next_temp_var()
        bool_var = self._next_temp_var()
        return (
            f"({bool_var} if isinstance(({bool_var} := to_bool(({cond_var} := {cond_expr}))), XlError) "
            f"else (({true_expr}) if {bool_var} else ({false_expr})))"
        )

    def _emit_row(self, node: FunctionCallNode) -> str:
        if len(node.args) < 1:
            return "XlError.VALUE"

        arg = node.args[0]
        if isinstance(arg, CellRefNode):
            sheet, cell = parse_address(arg.address)
            col_str, row = openpyxl.utils.cell.coordinate_from_string(cell)
            col = openpyxl.utils.cell.column_index_from_string(col_str)
            return f"xl_row(ExcelRange({repr(sheet)}, {row}, {col}, {row}, {col}))"
        if isinstance(arg, RangeNode):
            sheet, r1, c1, r2, c2 = self._range_coords(arg.start, arg.end)
            return f"xl_row(ExcelRange({repr(sheet)}, {r1}, {c1}, {r2}, {c2}))"
        if isinstance(arg, FunctionCallNode) and arg.name.upper() == "OFFSET":
            return f"xl_row({self._emit_offset_ref(arg)})"

        return f"xl_row({self._emit_ast(arg)})"

    def _emit_column(self, node: FunctionCallNode) -> str:
        if len(node.args) < 1:
            return "XlError.VALUE"

        arg = node.args[0]
        if isinstance(arg, CellRefNode):
            sheet, cell = parse_address(arg.address)
            col_str, row = openpyxl.utils.cell.coordinate_from_string(cell)
            col = openpyxl.utils.cell.column_index_from_string(col_str)
            return f"xl_column(ExcelRange({repr(sheet)}, {row}, {col}, {row}, {col}))"
        if isinstance(arg, RangeNode):
            sheet, r1, c1, r2, c2 = self._range_coords(arg.start, arg.end)
            return f"xl_column(ExcelRange({repr(sheet)}, {r1}, {c1}, {r2}, {c2}))"
        if isinstance(arg, FunctionCallNode) and arg.name.upper() == "OFFSET":
            return f"xl_column({self._emit_offset_ref(arg)})"

        return f"xl_column({self._emit_ast(arg)})"

    def _emit_columns(self, node: FunctionCallNode) -> str:
        if len(node.args) < 1:
            return "XlError.VALUE"

        arg = node.args[0]
        if isinstance(arg, CellRefNode):
            sheet, cell = parse_address(arg.address)
            col_str, row = openpyxl.utils.cell.coordinate_from_string(cell)
            col = openpyxl.utils.cell.column_index_from_string(col_str)
            return f"xl_columns(ExcelRange({repr(sheet)}, {row}, {col}, {row}, {col}))"
        if isinstance(arg, RangeNode):
            sheet, r1, c1, r2, c2 = self._range_coords(arg.start, arg.end)
            return f"xl_columns(ExcelRange({repr(sheet)}, {r1}, {c1}, {r2}, {c2}))"
        if isinstance(arg, FunctionCallNode) and arg.name.upper() == "OFFSET":
            return f"xl_columns({self._emit_offset_ref(arg)})"

        return f"xl_columns({self._emit_ast(arg)})"

    def _emit_choose(self, node: FunctionCallNode) -> str:
        """Emit CHOOSE as chained conditionals for lazy evaluation.

        CHOOSE(index_num, value1, [value2], ...)

        Emits as chained conditionals that only evaluate the selected value.
        This is critical for breaking circular references that Excel handles
        via lazy evaluation.
        """
        if len(node.args) < 2:
            return "XlError.VALUE"

        index_expr = self._emit_ast(node.args[0])
        value_exprs = [self._emit_ast(arg) for arg in node.args[1:]]

        # Store index in temp vars to avoid evaluating twice and to keep typing clean.
        # We coerce via to_int() (Excel-style numeric coercion + error propagation)
        # to avoid `int(CellValue)` in generated code (which type-checkers reject).
        var = self._next_temp_var()
        idx_var = self._next_temp_var()

        # Build chained conditionals: if idx==1 then val1 else if idx==2 then val2 ...
        # Start from the innermost (last value or VALUE error for out of bounds)
        result = "XlError.VALUE"
        for i, val_expr in reversed(list(enumerate(value_exprs, start=1))):
            result = f"(({val_expr}) if {idx_var} == {i} else ({result}))"

        # Wrap with error/bounds checking
        return (
            f"({var} if isinstance(({var} := {index_expr}), XlError) "
            f"else ({idx_var} if isinstance(({idx_var} := to_int({var})), XlError) "
            f"else XlError.VALUE if {idx_var} < 1 or {idx_var} > {len(value_exprs)} else {result}))"
        )

    def _is_constant_number(self, node: AstNode) -> bool:
        """Check if a node is a constant numeric value.

        Handles both NumberNode and unary negation of NumberNode (e.g., -2).
        """
        if isinstance(node, NumberNode):
            return True
        if isinstance(node, UnaryOpNode) and node.op == "-":
            return isinstance(node.operand, NumberNode)
        return False

    def _get_constant_number(self, node: AstNode) -> float:
        """Extract the numeric value from a constant number node.

        Assumes _is_constant_number() has already returned True.
        """
        if isinstance(node, NumberNode):
            return node.value
        if (
            isinstance(node, UnaryOpNode)
            and node.op == "-"
            and isinstance(node.operand, NumberNode)
        ):
            return -node.operand.value
        raise ValueError(f"Not a constant number: {node}")

    def _can_offset_be_static(self, node: FunctionCallNode) -> bool:
        """Check if an OFFSET call can be statically resolved.

        Returns True if reference is a cell or range and all offsets/sizes are constants.
        """
        if len(node.args) < 3:
            return False

        ref_node = node.args[0]
        rows_node = node.args[1]
        cols_node = node.args[2]
        height_node = node.args[3] if len(node.args) > 3 else None
        width_node = node.args[4] if len(node.args) > 4 else None

        return (
            isinstance(ref_node, (CellRefNode, RangeNode))
            and self._is_constant_number(rows_node)
            and self._is_constant_number(cols_node)
            and (height_node is None or self._is_constant_number(height_node))
            and (width_node is None or self._is_constant_number(width_node))
        )

    def _emit_offset(self, node: FunctionCallNode) -> str:
        """Emit OFFSET function, trying static resolution first.

        OFFSET(reference, rows, cols, [height], [width])

        If all offset arguments are constants, resolves to direct cell/range reference.
        Otherwise, falls back to runtime xl_offset() function.
        """
        if len(node.args) < 3:
            # Invalid OFFSET - need at least reference, rows, cols
            return "XlError.VALUE"

        ref_node = node.args[0]
        rows_node = node.args[1]
        cols_node = node.args[2]
        height_node = node.args[3] if len(node.args) > 3 else None
        width_node = node.args[4] if len(node.args) > 4 else None

        # Try static resolution if reference is a cell and offsets are constants
        if self._can_offset_be_static(node):
            assert isinstance(ref_node, (CellRefNode, RangeNode))
            base_address = (
                ref_node.address if isinstance(ref_node, CellRefNode) else ref_node.start
            )
            base_h, base_w = self._offset_base_shape(ref_node)
            height = (
                int(self._get_constant_number(height_node))
                if height_node is not None
                else base_h
            )
            width = (
                int(self._get_constant_number(width_node))
                if width_node is not None
                else base_w
            )
            return self._emit_offset_static(
                base_address,
                int(self._get_constant_number(rows_node)),
                int(self._get_constant_number(cols_node)),
                height,
                width,
            )

        # Fall back to runtime resolution
        self._needs_offset_runtime = True
        return self._emit_offset_dynamic(
            ref_node, rows_node, cols_node, height_node, width_node
        )

    def _offset_base_shape(self, ref_node: AstNode) -> tuple[int, int]:
        """Return (height, width) for an OFFSET base reference."""
        if isinstance(ref_node, CellRefNode):
            return (1, 1)
        if isinstance(ref_node, RangeNode):
            _, r1, c1, r2, c2 = self._range_coords(ref_node.start, ref_node.end)
            return (r2 - r1 + 1, c2 - c1 + 1)
        return (1, 1)

    def _range_coords(self, start: str, end: str) -> tuple[str, int, int, int, int]:
        """Parse a range into (sheet, start_row, start_col, end_row, end_col).

        Uses Excel semantics: start sheet applies to the whole range.
        """
        start_sheet, start_cell = parse_address(start)
        _, end_cell = parse_address(end)

        start_col_str, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
        end_col_str, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)

        start_col = openpyxl.utils.cell.column_index_from_string(start_col_str)
        end_col = openpyxl.utils.cell.column_index_from_string(end_col_str)

        r1, r2 = (start_row, end_row) if start_row <= end_row else (end_row, start_row)
        c1, c2 = (start_col, end_col) if start_col <= end_col else (end_col, start_col)
        return (start_sheet, r1, c1, r2, c2)

    def _emit_offset_static(
        self, base_address: str, rows: int, cols: int, height: int, width: int
    ) -> str:
        """Emit statically resolved OFFSET as direct cell/range reference."""
        base_sheet, base_cell = parse_address(base_address)
        base_col_str, base_row = openpyxl.utils.cell.coordinate_from_string(base_cell)
        base_col = openpyxl.utils.cell.column_index_from_string(base_col_str)

        # Compute target position
        target_row = base_row + rows
        target_col = base_col + cols

        if target_row < 1 or target_col < 1:
            # Invalid reference
            return "XlError.REF"

        target_col_str = openpyxl.utils.cell.get_column_letter(target_col)

        if height == 1 and width == 1:
            # Single cell reference
            target_addr = f"{quote_sheet_if_needed(base_sheet)}!{target_col_str}{target_row}"
            return self._emit_cell_eval(target_addr)
        else:
            # Range reference - emit as 2D array
            end_row = target_row + height - 1
            end_col = target_col + width - 1
            end_col_str = openpyxl.utils.cell.get_column_letter(end_col)

            start_addr = f"{quote_sheet_if_needed(base_sheet)}!{target_col_str}{target_row}"
            end_addr = f"{quote_sheet_if_needed(base_sheet)}!{end_col_str}{end_row}"

            # Generate 2D array like _emit_range does
            rows_list = self._range_addresses_2d(start_addr, end_addr)
            row_strs = []
            for row_addrs in rows_list:
                cell_calls = [self._emit_cell_eval(addr) for addr in row_addrs]
                row_strs.append("[" + ", ".join(cell_calls) + "]")
            return f"np.array([{', '.join(row_strs)}], dtype=object)"

    def _emit_offset_dynamic(
        self,
        ref_node: AstNode,
        rows_node: AstNode,
        cols_node: AstNode,
        height_node: AstNode | None,
        width_node: AstNode | None,
    ) -> str:
        """Emit dynamic OFFSET that resolves at runtime."""
        # For dynamic OFFSET, we need to pass the base reference info
        if isinstance(ref_node, CellRefNode):
            base_sheet, base_cell = parse_address(ref_node.address)
            self._offset_runtime_sheets.add(base_sheet)
            base_col_str, base_row = openpyxl.utils.cell.coordinate_from_string(base_cell)
            base_col = openpyxl.utils.cell.column_index_from_string(base_col_str)
            ref_info = f"({repr(base_sheet)}, {base_row}, {base_col})"
        elif isinstance(ref_node, RangeNode):
            base_sheet, r1, c1, r2, c2 = self._range_coords(ref_node.start, ref_node.end)
            self._offset_runtime_sheets.add(base_sheet)
            ref_info = f"({repr(base_sheet)}, {r1}, {c1}, {r2}, {c2})"
        else:
            # If reference is not a simple cell, we can't handle it
            return "XlError.REF"

        rows_expr = self._emit_ast(rows_node)
        cols_expr = self._emit_ast(cols_node)
        height_expr = "None" if height_node is None else self._emit_ast(height_node)
        width_expr = "None" if width_node is None else self._emit_ast(width_node)

        return f"xl_offset(ctx, {ref_info}, {rows_expr}, {cols_expr}, {height_expr}, {width_expr})"

    def _emit_offset_ref(self, node: FunctionCallNode) -> str:
        if len(node.args) < 3:
            return "XlError.VALUE"

        ref_node = node.args[0]
        rows_node = node.args[1]
        cols_node = node.args[2]
        height_node = node.args[3] if len(node.args) > 3 else None
        width_node = node.args[4] if len(node.args) > 4 else None

        if isinstance(ref_node, CellRefNode):
            base_sheet, base_cell = parse_address(ref_node.address)
            base_col_str, base_row = openpyxl.utils.cell.coordinate_from_string(base_cell)
            base_col = openpyxl.utils.cell.column_index_from_string(base_col_str)
            ref_info = f"({repr(base_sheet)}, {base_row}, {base_col})"
        elif isinstance(ref_node, RangeNode):
            base_sheet, r1, c1, r2, c2 = self._range_coords(ref_node.start, ref_node.end)
            ref_info = f"({repr(base_sheet)}, {r1}, {c1}, {r2}, {c2})"
        else:
            return "XlError.REF"

        rows_expr = self._emit_ast(rows_node)
        cols_expr = self._emit_ast(cols_node)
        height_expr = "None" if height_node is None else self._emit_ast(height_node)
        width_expr = "None" if width_node is None else self._emit_ast(width_node)

        return f"xl_offset_ref({ref_info}, {rows_expr}, {cols_expr}, {height_expr}, {width_expr})"

    def _emit_cell(self, address: str) -> str:
        """Emit a Python function for a single formula cell.

        Args:
            address: Sheet-qualified cell address (e.g., 'Sheet1!A1')

        Returns:
            Python function definition as a string.
        """
        normalized = normalize_address(address)
        func_name = address_to_python_name(normalized)
        node = self.graph.get_node(normalized)

        if node is None or node.formula is None:
            raise ValueError(f"Not a formula cell: {normalized}")

        lines: list[str] = []
        lines.append(f"def {func_name}(ctx):")
        doc = f"Formula: {node.formula}".replace("'''", "\\'''")
        lines.append(f"    '''{doc}'''")
        # Reset temp var counter for each cell to keep variable names short
        self._temp_var_counter = 0
        ast = self._get_or_parse_ast(normalized)
        assert ast is not None
        expr = self._emit_ast(ast)
        lines.append(f"    return {expr}")

        return "\n".join(lines)

    def _collect_dependencies(self, address: str) -> list[str]:
        """Collect all cell addresses that a cell depends on (recursively).

        Cycles are allowed in the dependency graph. Excel permits circular
        references when broken by conditional evaluation (IF, IFERROR, etc.).
        The generated code handles cycles at runtime via EvalContext tracking
        and lazy evaluation of function arguments.

        Missing cells (referenced but not in graph) are included in the output
        so stub functions can be generated for them.

        Args:
            address: Starting cell address.

        Returns:
            List of all dependent cell addresses in dependency order.
            Includes missing cells (not in graph) that are referenced by formulas.
        """
        visited: set[str] = set()
        in_progress: set[str] = set()  # Currently being visited (for cycle detection)
        order: list[str] = []

        def visit(addr: str) -> None:
            # Normalize address to match Node.key format
            addr = normalize_address(addr)

            # Skip if already fully visited or currently being visited (cycle)
            if addr in visited or addr in in_progress:
                return

            in_progress.add(addr)

            node = self.graph.get_node(addr)
            if node is None:
                # Cell not in graph - still add to order so we generate a stub
                order.append(addr)
                in_progress.discard(addr)
                visited.add(addr)
                return

            # If it's a formula, parse and find cell references
            if node.formula is not None:
                ast = self._get_or_parse_ast(addr)
                assert ast is not None
                deps = self._extract_cell_refs(ast)
                for dep in deps:
                    visit(dep)

            order.append(addr)
            in_progress.discard(addr)
            visited.add(addr)

        visit(address)
        return order

    def _extract_cell_refs(self, node: AstNode) -> list[str]:
        """Extract all cell references from an AST node."""
        refs: list[str] = []

        if isinstance(node, CellRefNode):
            refs.append(node.address)
        elif isinstance(node, RangeNode):
            refs.extend(self._range_addresses(node.start, node.end))
        elif isinstance(node, BinaryOpNode):
            refs.extend(self._extract_cell_refs(node.left))
            refs.extend(self._extract_cell_refs(node.right))
        elif isinstance(node, UnaryOpNode):
            refs.extend(self._extract_cell_refs(node.operand))
        elif isinstance(node, FunctionCallNode):
            # For OFFSET that can be statically resolved, extract target cells
            if node.name.upper() == "OFFSET" and self._can_offset_be_static(node):
                refs.extend(self._extract_offset_target_refs(node))
            for arg in node.args:
                refs.extend(self._extract_cell_refs(arg))

        return refs

    def _extract_offset_target_refs(self, node: FunctionCallNode) -> list[str]:
        """Extract target cell references from a statically resolvable OFFSET."""
        ref_node = node.args[0]
        rows_node = node.args[1]
        cols_node = node.args[2]
        height_node = node.args[3] if len(node.args) > 3 else None
        width_node = node.args[4] if len(node.args) > 4 else None

        if isinstance(ref_node, CellRefNode):
            base_address = ref_node.address
            base_h, base_w = (1, 1)
        elif isinstance(ref_node, RangeNode):
            base_address = ref_node.start
            base_h, base_w = self._offset_base_shape(ref_node)
        else:
            return []

        base_sheet, base_cell = parse_address(base_address)
        base_col_str, base_row = openpyxl.utils.cell.coordinate_from_string(base_cell)
        base_col = openpyxl.utils.cell.column_index_from_string(base_col_str)

        rows = int(self._get_constant_number(rows_node))
        cols = int(self._get_constant_number(cols_node))
        height = int(self._get_constant_number(height_node)) if height_node is not None else base_h
        width = int(self._get_constant_number(width_node)) if width_node is not None else base_w

        target_row = base_row + rows
        target_col = base_col + cols

        if target_row < 1 or target_col < 1:
            return []

        refs = []
        for r in range(target_row, target_row + height):
            for c in range(target_col, target_col + width):
                col_str = openpyxl.utils.cell.get_column_letter(c)
                refs.append(f"{quote_sheet_if_needed(base_sheet)}!{col_str}{r}")

        return refs

    def _extract_xl_functions(self, node: AstNode) -> set[str]:
        """Extract all xl_* function names and markers used in an AST node.

        Special markers:
        - "XlError": XlError enum is needed (e.g., error literals like #N/A)
        - "numpy": numpy is needed for np.array() wrapping of ranges
        """
        funcs: set[str] = set()

        if isinstance(node, ErrorNode):
            # Error literal requires XlError enum
            funcs.add("XlError")
        elif isinstance(node, FunctionCallNode):
            upper_name = node.name.upper()

            # IF, IFERROR, CHOOSE are special - emitted as native Python conditionals
            if upper_name == "IF":
                funcs.add("XlError")
                funcs.add("to_bool")
            elif upper_name == "IFERROR" or upper_name in {"IFNA", "_XLFN.IFNA"}:
                funcs.add("XlError")
            elif upper_name == "CHOOSE":
                funcs.add("XlError")
                funcs.add("to_int")
            elif upper_name == "ROW":
                funcs.add("xl_row")
                if node.args:
                    ref = node.args[0]
                    if (
                        isinstance(ref, FunctionCallNode)
                        and ref.name.upper() == "OFFSET"
                    ):
                        funcs.add("xl_offset_ref")
                        for off_arg in ref.args:
                            funcs.update(self._extract_xl_functions(off_arg))
                    else:
                        funcs.update(self._extract_xl_functions(ref))
            elif upper_name in {"COLUMN", "COLUMNS"}:
                funcs.add("xl_column" if upper_name == "COLUMN" else "xl_columns")
                if node.args:
                    ref = node.args[0]
                    if (
                        isinstance(ref, FunctionCallNode)
                        and ref.name.upper() == "OFFSET"
                    ):
                        funcs.add("xl_offset_ref")
                        for off_arg in ref.args:
                            funcs.update(self._extract_xl_functions(off_arg))
                    else:
                        funcs.update(self._extract_xl_functions(ref))
            # OFFSET is special - only add xl_offset if it can't be statically resolved
            elif upper_name == "OFFSET":
                if not self._can_offset_be_static(node):
                    funcs.add("xl_offset")
            else:
                funcs.add(excel_func_to_python(node.name))

            # Check if this function needs numpy array wrapping for range args
            numpy_array_args = {
                "LOOKUP": {1, 2},
                "VLOOKUP": {1},
                "HLOOKUP": {1},
                "INDEX": {0},
                "MATCH": {1},
                "SUMPRODUCT": set(range(10)),
            }
            if upper_name in numpy_array_args:
                for i, arg in enumerate(node.args):
                    if i in numpy_array_args[upper_name] and isinstance(arg, RangeNode):
                        funcs.add("numpy")
                        break
            for arg in node.args:
                funcs.update(self._extract_xl_functions(arg))
        elif isinstance(node, BinaryOpNode):
            # Binary operators use xl_* functions for error propagation
            if node.op == "&":
                funcs.add("xl_concat")
            elif node.op in _BINARY_OPS:
                funcs.add(_BINARY_OPS[node.op])
            funcs.update(self._extract_xl_functions(node.left))
            funcs.update(self._extract_xl_functions(node.right))
        elif isinstance(node, UnaryOpNode):
            # Unary operators use xl_* functions for error propagation
            if node.op in _UNARY_OPS:
                funcs.add(_UNARY_OPS[node.op])
            funcs.update(self._extract_xl_functions(node.operand))

        return funcs

    def generate(
        self,
        targets: list[str],
        *,
        entrypoints: Mapping[str, Sequence[str]] | None = None,
        constant_types: set[str] | None = None,
        constant_ranges: Sequence[str] | None = None,
        constant_blanks: bool = False,
        input_ranges: Sequence[str] | None = None,
    ) -> str:
        """Generate standalone Python code for target cells.

        Args:
            targets: List of target cell addresses to compute.
            entrypoints: Optional mapping of entrypoint names to target lists.
            input_ranges: Sheet-qualified ranges whose leaf cells are treated as inputs.
                When a cell would otherwise be a constant, ``input_ranges`` take precedence.

        Returns:
            Standalone Python source code as a string.
        """
        normalized_entrypoints = self._normalize_entrypoints(entrypoints)
        normalized_targets = [normalize_address(t) for t in targets]
        entrypoint_targets: list[str] = []
        seen_targets: set[str] = set()
        for entrypoint_list in normalized_entrypoints.values():
            for target in entrypoint_list:
                if target not in seen_targets:
                    seen_targets.add(target)
                    entrypoint_targets.append(target)

        compute_all_targets = (
            normalized_targets
            or (entrypoint_targets if normalized_entrypoints else normalized_targets)
        )
        dependency_targets: list[str] = []
        seen_dependency: set[str] = set()
        for target in compute_all_targets:
            if target not in seen_dependency:
                seen_dependency.add(target)
                dependency_targets.append(target)
        for target in entrypoint_targets:
            if target not in seen_dependency:
                seen_dependency.add(target)
                dependency_targets.append(target)

        parts = self._generate_parts(
            compute_all_targets,
            dependency_targets=dependency_targets,
            constant_types=constant_types,
            constant_ranges=constant_ranges,
            constant_blanks=constant_blanks,
            input_ranges=input_ranges,
        )
        runtime_code = parts["runtime_code"]
        cell_code_lines = parts["cell_code_lines"]
        _formula_cells = parts["formula_cells"]
        _all_cells = parts["all_cells"]
        normalized_targets = parts["targets"]

        entrypoint_entries = {
            name: self._targets_to_entries(entrypoint_list)
            for name, entrypoint_list in normalized_entrypoints.items()
        }
        targets_entries = self._targets_to_entries(normalized_targets)
        _needs_range_helper = any(
            handler == "xl_range"
            for entries in entrypoint_entries.values()
            for _, handler in entries
        ) or any(handler == "xl_range" for _, handler in targets_entries)

        # Combine: runtime + inputs + formulas + entry point
        lines: list[str] = [runtime_code, ""]

        lines.extend(parts["inputs_block_lines"])
        if parts["has_constants"]:
            lines.append("")
            lines.extend(parts["constants_block_lines"])
        lines.append("")
        lines.append("")

        # Export formula cell implementations and a resolver.
        lines.append("# --- Formula cell functions ---\n")
        lines.extend(cell_code_lines)
        lines.extend(self._emit_resolver_lines())

        # Generate entry point helpers
        lines.append("def make_context(inputs=None):")
        lines.append('    """Create an EvalContext with merged inputs."""')
        lines.append("    merged = dict(DEFAULT_INPUTS)")
        if parts["has_constants"]:
            lines.append("    merged.update(CONSTANTS)")
        lines.append("    if inputs is not None:")
        lines.append("        merged.update(inputs)")
        lines.append(
            "    return EvalContext("
            "inputs=merged, resolver=_resolve_formula, "
            f"iterative_enabled={bool(self._iterate_enabled)}, "
            f"iterate_count={int(self._iterate_count)}, "
            f"iterate_delta={float(self._iterate_delta)!r})"
        )
        lines.append("")
        lines.append("")
        for name, entrypoint_list in normalized_entrypoints.items():
            targets_name = f"TARGETS_{name.upper()}"
            lines.append(f"{targets_name} = {{")
            for target, handler in self._targets_to_entries(entrypoint_list):
                lines.append(f"    {repr(target)}: {handler},")
            lines.append("}")
            lines.append("")
            lines.append("")
            lines.append(f"def compute_{name}(inputs=None, *, ctx=None):")
            lines.append(
                f'    """Compute {name} target cells and return results."""'
            )
            lines.append("    if ctx is None:")
            lines.append("        ctx = make_context(inputs)")
            lines.append("    elif inputs is not None:")
            lines.append(
                "        warnings.warn("
                '"inputs will be ignored because ctx was provided", '
                "UserWarning, stacklevel=2)"
            )
            if self._iterate_enabled:
                lines.append(f"    return xl_iterative_compute(ctx, {targets_name})")
            else:
                lines.append(
                    f"    return {{target: handler(ctx, target) for target, handler in {targets_name}.items()}}"
                )
            lines.append("")
            lines.append("")
        lines.append("TARGETS = {")
        for target, handler in self._targets_to_entries(normalized_targets):
            lines.append(f"    {repr(target)}: {handler},")
        lines.append("}")
        lines.append("")
        lines.append("")
        lines.append("def compute_all(inputs=None, *, ctx=None):")
        lines.append('    """Compute all target cells and return results."""')
        lines.append("    if ctx is None:")
        lines.append("        ctx = make_context(inputs)")
        lines.append("    elif inputs is not None:")
        lines.append(
            "        warnings.warn("
            '"inputs will be ignored because ctx was provided", '
            "UserWarning, stacklevel=2)"
        )
        if self._iterate_enabled:
            lines.append("    return xl_iterative_compute(ctx, TARGETS)")
        else:
            lines.append("    return {target: handler(ctx, target) for target, handler in TARGETS.items()}")
        lines.append("")

        return "\n".join(lines)

    def generate_modules(
        self,
        targets: list[str],
        *,
        entrypoints: Mapping[str, Sequence[str]] | None = None,
        package_name: str = "exported",
        constant_types: set[str] | None = None,
        constant_ranges: Sequence[str] | None = None,
        constant_blanks: bool = False,
        input_ranges: Sequence[str] | None = None,
    ) -> dict[str, str]:
        """Generate a multi-module Python package for target cells.

        Returns a mapping of file paths (including a normalized package directory) to file
        contents. The package name is normalized to an importable directory name.

        The generated package has six flat files:
        - __init__.py: exports compute_all and DEFAULT_INPUTS
        - entrypoint.py: compute_all implementation
        - inputs.py: DEFAULT_INPUTS
        - constants.py: CONSTANTS (may be empty)
        - runtime.py: embedded Excel runtime (emit_runtime)
        - internals.py: formula cell functions + resolver dispatch
        """
        pkg = self._normalize_package_name(package_name)
        normalized_entrypoints = self._normalize_entrypoints(entrypoints)
        normalized_targets = [normalize_address(t) for t in targets]
        entrypoint_targets: list[str] = []
        seen_targets: set[str] = set()
        for entrypoint_list in normalized_entrypoints.values():
            for target in entrypoint_list:
                if target not in seen_targets:
                    seen_targets.add(target)
                    entrypoint_targets.append(target)
        compute_all_targets = (
            normalized_targets
            or (entrypoint_targets if normalized_entrypoints else normalized_targets)
        )
        dependency_targets: list[str] = []
        seen_dependency: set[str] = set()
        for target in compute_all_targets:
            if target not in seen_dependency:
                seen_dependency.add(target)
                dependency_targets.append(target)
        for target in entrypoint_targets:
            if target not in seen_dependency:
                seen_dependency.add(target)
                dependency_targets.append(target)

        parts = self._generate_parts(
            compute_all_targets,
            dependency_targets=dependency_targets,
            constant_types=constant_types,
            constant_ranges=constant_ranges,
            constant_blanks=constant_blanks,
            input_ranges=input_ranges,
        )
        runtime_code = parts["runtime_code"]
        cell_code_lines = parts["cell_code_lines"]
        _formula_cells = parts["formula_cells"]
        _all_cells = parts["all_cells"]
        normalized_targets = parts["targets"]

        entrypoint_entries = {
            name: self._targets_to_entries(entrypoint_list)
            for name, entrypoint_list in normalized_entrypoints.items()
        }
        targets_entries = self._targets_to_entries(normalized_targets)
        needs_range_helper = any(
            handler == "xl_range"
            for entries in entrypoint_entries.values()
            for _, handler in entries
        ) or any(handler == "xl_range" for _, handler in targets_entries)

        inputs_py = "\n".join(
            [
                "from __future__ import annotations",
                "",
                "# --- Default inputs (leaf cells) ---",
                *parts["inputs_block_lines"][1:],  # drop the comment already included above
                "",
            ]
        )

        runtime_py = runtime_code.rstrip() + "\n"

        constants_lines_out: list[str] = [
            "from __future__ import annotations",
            "",
        ]
        if parts["constants_block_lines"]:
            constants_lines_out.extend(parts["constants_block_lines"])
        else:
            constants_lines_out.append("CONSTANTS = {}")
        constants_lines_out.append("")
        constants_py = "\n".join(constants_lines_out).rstrip() + "\n"

        internals_import_names = self._internals_runtime_import_names(
            parts["used_xl_functions"], cell_code_lines
        )
        runtime_import_block = self._format_from_runtime_import(internals_import_names)
        internals_lines: list[str] = ["from __future__ import annotations", ""]
        if runtime_import_block:
            internals_lines.append(runtime_import_block)
            internals_lines.append("")
        internals_lines.append("# --- Formula cell functions ---\n")
        internals_lines.extend(cell_code_lines)
        internals_lines.extend(self._emit_resolver_lines())
        internals_py = "\n".join(internals_lines).rstrip() + "\n"

        runtime_entry_names = ["EvalContext", "xl_cell"]
        if needs_range_helper:
            runtime_entry_names.append("xl_range")
        if self._iterate_enabled:
            runtime_entry_names.append("xl_iterative_compute")
        runtime_entry_names.sort()
        runtime_imports = self._format_from_runtime_import(runtime_entry_names)

        entrypoint_lines: list[str] = [
            "from __future__ import annotations",
            "",
            "from .constants import CONSTANTS",
            "from .inputs import DEFAULT_INPUTS",
            "from .internals import _resolve_formula",
            runtime_imports,
            "import warnings",
            "",
            "",
            "def make_context(inputs=None):",
            '    """Create an EvalContext with merged inputs."""',
            "    merged = dict(DEFAULT_INPUTS)",
            "    merged.update(CONSTANTS)",
            "    if inputs is not None:",
            "        merged.update(inputs)",
            (
                "    return EvalContext("
                "inputs=merged, resolver=_resolve_formula, "
                f"iterative_enabled={bool(self._iterate_enabled)}, "
                f"iterate_count={int(self._iterate_count)}, "
                f"iterate_delta={float(self._iterate_delta)!r})"
            ),
            "",
            "",
        ]
        for name in normalized_entrypoints:
            targets_name = f"TARGETS_{name.upper()}"
            entrypoint_lines.append(f"{targets_name} = {{")
            for target, handler in entrypoint_entries[name]:
                entrypoint_lines.append(f"    {repr(target)}: {handler},")
            entrypoint_lines.append("}")
            entrypoint_lines.extend(
                [
                    "",
                    "",
                    f"def compute_{name}(inputs=None, *, ctx=None):",
                    f'    """Compute {name} target cells and return results."""',
                    "    if ctx is None:",
                    "        ctx = make_context(inputs)",
                    "    elif inputs is not None:",
                    "        warnings.warn(",
                    '            "inputs will be ignored because ctx was provided",',
                    "            UserWarning,",
                    "            stacklevel=2,",
                    "        )",
                    (
                        f"    return xl_iterative_compute(ctx, {targets_name})"
                        if self._iterate_enabled
                        else f"    return {{target: handler(ctx, target) for target, handler in {targets_name}.items()}}"
                    ),
                    "",
                    "",
                ]
            )
        entrypoint_lines.append("TARGETS = {")
        for target, handler in targets_entries:
            entrypoint_lines.append(f"    {repr(target)}: {handler},")
        entrypoint_lines.extend(
            [
                "}",
                "",
                "",
                "def compute_all(inputs=None, *, ctx=None):",
                '    """Compute all target cells and return results."""',
                "    if ctx is None:",
                "        ctx = make_context(inputs)",
                "    elif inputs is not None:",
                "        warnings.warn(",
                '            "inputs will be ignored because ctx was provided",',
                "            UserWarning,",
                "            stacklevel=2,",
                "        )",
                (
                    "    return xl_iterative_compute(ctx, TARGETS)"
                    if self._iterate_enabled
                    else "    return {target: handler(ctx, target) for target, handler in TARGETS.items()}"
                ),
                "",
            ]
        )
        entrypoint_py = "\n".join(entrypoint_lines)

        entrypoint_exports = ["compute_all", "make_context"]
        entrypoint_exports.extend(f"compute_{name}" for name in normalized_entrypoints)
        entrypoint_imports = ", ".join(entrypoint_exports)
        all_exports = entrypoint_exports + ["DEFAULT_INPUTS"]
        init_py = "\n".join(
            [
                "from __future__ import annotations",
                "",
                f"from .entrypoint import {entrypoint_imports}  # noqa: F401",
                "from .inputs import DEFAULT_INPUTS  # noqa: F401",
                "",
                f"__all__ = {all_exports!r}",
                "",
            ]
        )

        return {
            f"{pkg}/__init__.py": init_py,
            f"{pkg}/entrypoint.py": entrypoint_py,
            f"{pkg}/inputs.py": inputs_py,
            f"{pkg}/constants.py": constants_py,
            f"{pkg}/runtime.py": runtime_py,
            f"{pkg}/internals.py": internals_py,
        }

    @staticmethod
    def _normalize_package_name(name: str) -> str:
        # Keep the name importable as a Python package (identifier-ish).
        out = re.sub(r"[^A-Za-z0-9_]+", "_", name.strip())
        out = re.sub(r"_+", "_", out).strip("_")
        if not out:
            out = "exported"
        if out[0].isdigit():
            out = "_" + out
        return out.lower()

    def _generate_parts(
        self,
        targets: list[str],
        *,
        dependency_targets: list[str] | None = None,
        constant_types: set[str] | None = None,
        constant_ranges: Sequence[str] | None = None,
        constant_blanks: bool = False,
        input_ranges: Sequence[str] | None = None,
    ) -> GenerationParts:
        """Generate shared intermediate artifacts for single-file and modular exports."""
        # Reset state for this generation
        self._needs_offset_runtime = False
        self._offset_runtime_sheets.clear()
        self._ast_cache.clear()
        self._used_graph_closure = False

        normalized_targets = [normalize_address(t) for t in targets]
        normalized_dependency_targets = (
            [normalize_address(t) for t in dependency_targets]
            if dependency_targets is not None
            else normalized_targets
        )

        # Collect all dependencies for all targets.
        #
        # When we are given a real excel_grapher.DependencyGraph, prefer its dependency edges
        # as the single source of truth for the export surface area. This ensures the exported
        # package can evaluate the full excel_grapher dependency closure (for the workbook's
        # cached state) without missing-cell KeyErrors.
        all_cells = self._collect_all_cells(normalized_dependency_targets)

        # Generate formula cell functions and collect used xl_* functions
        self._emitted.clear()
        cell_code_lines: list[str] = []
        used_xl_functions: set[str] = set()
        formula_cells: set[str] = set()

        for address in all_cells:
            if address in self._emitted:
                continue
            self._emitted.add(address)
            node = self.graph.get_node(address)
            if node is not None and node.formula is not None:
                formula_cells.add(normalize_address(address))
                cell_code_lines.append(self._emit_cell(address))
                cell_code_lines.append("")
                cell_code_lines.append("")

                ast = self._get_or_parse_ast(address)
                assert ast is not None
                used_xl_functions.update(self._extract_xl_functions(ast))

        # If dynamic OFFSET was used, ensure the runtime implementation is embedded.
        #
        # For graphs without excel_grapher dependency edges (common in unit tests that only add
        # nodes), we expand the export surface area to include the full graph so dynamic OFFSET
        # reads do not hit missing-cell KeyErrors. For real workbook graphs (excel_grapher
        # closure), we intentionally *do not* widen the export surface area here.
        if self._needs_offset_runtime:
            used_xl_functions.add("xl_offset")
            if not self._used_graph_closure:
                all_graph_cells = (
                    list(self.graph.leaf_keys()) + list(self.graph.formula_keys())
                )
                if self._offset_runtime_sheets:
                    all_graph_cells = [
                        addr
                        for addr in all_graph_cells
                        if parse_address(normalize_address(addr))[0]
                        in self._offset_runtime_sheets
                    ]
                for address in all_graph_cells:
                    if address in self._emitted:
                        continue
                    self._emitted.add(address)
                    all_cells.append(address)
                    node = self.graph.get_node(address)
                    if node is not None and node.formula is not None:
                        formula_cells.add(normalize_address(address))
                        cell_code_lines.append(self._emit_cell(address))
                        cell_code_lines.append("")
                        cell_code_lines.append("")

                        ast = self._get_or_parse_ast(address)
                        assert ast is not None
                        used_xl_functions.update(self._extract_xl_functions(ast))

        # Always include per-call evaluation scaffolding.
        # XlError is commonly referenced by generated code (error literals, IF/IFERROR, and
        # potentially leaf inputs), so keep it available.
        runtime_symbols = set(used_xl_functions) | {
            "EvalContext",
            "xl_cell",
            "xl_eval",
            "xl_range",
            "XlError",
        }
        if self._iterate_enabled:
            runtime_symbols.add("xl_iterative_compute")
        runtime_code = emit_runtime(runtime_symbols, include_offset_table=False)
        runtime_code = runtime_code.rstrip()

        normalized_constant_types = self._normalize_constant_types(constant_types)
        normalized_constant_ranges = self._normalize_constant_ranges(constant_ranges)
        normalized_input_ranges = self._normalize_input_ranges(input_ranges)
        include_constants = bool(constant_types or constant_ranges or constant_blanks)
        graph_classification = None
        if not include_constants:
            graph_classification = self._get_graph_leaf_classification()
        inputs_block_lines, constants_block_lines = self._emit_default_inputs_lines(
            all_cells,
            constant_types=normalized_constant_types,
            constant_ranges=normalized_constant_ranges,
            constant_blanks=constant_blanks,
            graph_classification=graph_classification,
            include_constants=include_constants,
            input_ranges=normalized_input_ranges,
        )
        if graph_classification is not None and constants_block_lines:
            include_constants = True

        return {
            "runtime_code": runtime_code,
            "inputs_block_lines": inputs_block_lines,
            "constants_block_lines": constants_block_lines,
            "cell_code_lines": cell_code_lines,
            "formula_cells": sorted(formula_cells),
            "all_cells": all_cells,
            "needs_offset_table": self._needs_offset_runtime,
            "targets": normalized_targets,
            "has_constants": include_constants,
            "used_xl_functions": frozenset(used_xl_functions),
        }

    def _collect_all_cells(self, targets: list[str]) -> list[str]:
        """Collect an ordered list of addresses to emit for the given targets.

        For excel_grapher.DependencyGraph instances, this uses the graph dependency edges and
        evaluation order as the export closure. For other GraphLike implementations, it falls
        back to the CodeGenerator AST-based dependency walk.
        """
        # Prefer graph-driven closure when excel_grapher provides an evaluation order AND
        # has dependency edges populated. Many unit tests build a DependencyGraph with nodes
        # only (no edges); for those we must fall back to AST-based dependency discovery.
        eval_order = getattr(self.graph, "evaluation_order", None)
        if callable(eval_order):
            # Heuristic: only use graph edges if any target has at least one dependency edge.
            # (Graphs constructed via create_dependency_graph(...) will satisfy this for
            # non-leaf targets; test graphs that only add nodes will not.)
            has_edges = any(bool(self.graph.dependencies(t)) for t in targets)
            if not has_edges:
                return self._collect_all_cells_via_ast(targets)

            self._used_graph_closure = True
            closure: set[str] = set()
            stack = list(targets)
            while stack:
                addr = normalize_address(stack.pop())
                if addr in closure:
                    continue
                node = self.graph.get_node(addr)
                if node is None:
                    continue
                closure.add(addr)
                for dep in self.graph.dependencies(addr):
                    dep_n = normalize_address(dep)
                    if dep_n not in closure and self.graph.get_node(dep_n) is not None:
                        stack.append(dep_n)

            if self._iterate_enabled:
                ordered = []
            else:
                try:
                    ordered = list(eval_order(strict=False))
                except CycleError:
                    raise
                except Exception:
                    ordered = []

            out: list[str] = []
            seen: set[str] = set()
            for addr in ordered:
                a = normalize_address(addr)
                if a in seen or a not in closure:
                    continue
                seen.add(a)
                out.append(a)
            for addr in sorted(closure):
                if addr not in seen:
                    out.append(addr)
            return out

        return self._collect_all_cells_via_ast(targets)

    def _collect_all_cells_via_ast(self, targets: list[str]) -> list[str]:
        """AST-based dependency walk (works for GraphLike test doubles)."""
        out: list[str] = []
        seen: set[str] = set()
        for target in targets:
            deps = self._collect_dependencies(target)
            for dep in deps:
                dep_n = normalize_address(dep)
                if dep_n not in seen:
                    seen.add(dep_n)
                    out.append(dep_n)
        return out

    def _emit_default_inputs_lines(
        self,
        all_cells: list[str],
        *,
        constant_types: set[str],
        constant_ranges: list[tuple[str, int, int, int, int]],
        constant_blanks: bool,
        graph_classification: dict[str, str] | None,
        include_constants: bool,
        input_ranges: list[tuple[str, int, int, int, int]],
    ) -> tuple[list[str], list[str]]:
        default_lines: list[str] = []
        default_lines.append("# --- Default inputs (leaf cells) ---")
        default_lines.append("DEFAULT_INPUTS = {")
        needed_leaves = self._collect_needed_leaves(all_cells)

        if include_constants:
            inputs, constants = self._classify_leaf_nodes(
                needed_leaves,
                constant_types=constant_types,
                constant_ranges=constant_ranges,
                constant_blanks=constant_blanks,
                input_ranges=input_ranges,
            )
        else:
            inputs, constants = self._classification_from_graph(
                graph_classification, needed_leaves
            )
            inputs, constants = self._apply_input_ranges_override(
                needed_leaves, constants, input_ranges
            )

        for key in sorted(inputs):
            node = self.graph.get_node(key)
            value = 0 if node is None else node.value
            default_lines.append(f"    {repr(key)}: {self._py_literal(value)},")
        default_lines.append("}")

        constants_lines: list[str] = []
        if constants:
            constants_lines.append("# --- Constant leaf values ---")
            constants_lines.append("CONSTANTS = {")
            for key in sorted(constants):
                node = self.graph.get_node(key)
                value = 0 if node is None else node.value
                constants_lines.append(f"    {repr(key)}: {self._py_literal(value)},")
            constants_lines.append("}")

        return default_lines, constants_lines

    def _emit_offset_cell_table_lines(self, all_cells: list[str]) -> list[str]:
        lines: list[str] = []
        lines.append("# --- Cell lookup table for dynamic OFFSET ---")
        lines.append("_CELL_TABLE = {")
        for address in all_cells:
            normalized = normalize_address(address)
            sheet, cell = parse_address(normalized)
            col_str, row = openpyxl.utils.cell.coordinate_from_string(cell)
            col = openpyxl.utils.cell.column_index_from_string(col_str)
            lines.append(f"    ({repr(sheet)}, {row}, {col}): {repr(normalized)},")
        lines.append("}")
        return lines
