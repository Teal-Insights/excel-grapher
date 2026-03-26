#!/usr/bin/env python3
"""
Map dependencies for LIC-DSF indicator rows using excel-grapher.

This script traces the dependency closure for key indicators across sheets
and validates against calcChain.xml.

Dynamic refs (OFFSET/INDIRECT) are resolved via a constraint-based config.
Iterative workflow: run the script; if DynamicRefError is raised, the message
includes the formula cell that needs a constraint. Inspect that cell and the
row/column headers in the workbook to decide plausible input domains, add the
address to LicDsfConstraints (with Annotated[int, Between(lo, hi)] or
Literal[...]), then re-run until the graph
builds.
"""

import time
from pathlib import Path
from typing import (  # noqa: F401 - Annotated/Literal used when adding constraints
    Annotated,
    Any,
    Literal,
    TypedDict,
    cast,
)

import openpyxl
import openpyxl.utils.cell

from excel_grapher import (
    CycleError,
    DynamicRefConfig,
    DynamicRefError,
    constrain,
    create_dependency_graph,
    format_cell_key,
    get_calc_settings,
    to_graphviz,
    validate_graph,
)
from excel_grapher.core.cell_types import Between  # noqa: F401 - used when adding constraints
from excel_grapher.grapher.dynamic_refs import (
    FromWorkbook,  # noqa: F401 - used when adding constraints
)

# Row labels for the multi-row stress-test blocks (same row layout in each block).
# Blank string means that row is skipped when splitting by row.
STRESS_TEST_ROW_LABELS: list[str] = [
    "Baseline",
    "A1. Key variables at their historical averages in 2024-2034 2/",
    "B1. Real GDP growth",
    "B2. Primary balance",
    "B3. Exports",
    "B4. Other flows 3/",
    "B5. Depreciation",
    "B6. Combination of B1-B5",
    "",  # blank row
    "C1. Combined contingent liabilities",
    "C2. Natural disaster",
    "C3. Commodity price",
    "C4. Market Financing",
    "A2. Alternative Scenario :[Customize, enter title]",
]

# Explicit cell ranges to extract (sheet-qualified A1 range, e.g. "'Chart Data'!D10:D17").
# All cells in each range are included as graph targets.
# Multi-row stress-test blocks are split by row using STRESS_TEST_ROW_LABELS (blank row skipped).
_CHART_DATA_FIXED: list[tuple[str, str]] = [
    ("External DSA risk rating signals", "'Chart Data'!D10:D17"),
    ("Fiscal (Total Public Debt) risk rating signals", "'Chart Data'!I10:I14"),
    ("Applicable tailored stress test signals", "'Chart Data'!I17:I19"),
    ("Fiscal space for moderate risk category", "'Chart Data'!E25:E27"),
    ("Overall rating", "'Chart Data'!L10:L11"),
]

_STRESS_TEST_BLOCKS: list[tuple[str, int]] = [
    ("PV of Debt-to-GDP Ratio", 239),
    ("PV of Debt-to-Revenue Ratio", 281),
    ("Debt Service-to-Revenue Ratio", 318),
    ("Debt Service-to-GDP Ratio", 351),
]


def _chart_data_ranges() -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = list(_CHART_DATA_FIXED)
    for metric_label, start_row in _STRESS_TEST_BLOCKS:
        for i, row_label in enumerate(STRESS_TEST_ROW_LABELS):
            if not row_label:
                continue
            row = start_row + i
            out.append((f"{metric_label} - {row_label}", f"'Chart Data'!D{row}:X{row}"))
    return out


CHART_DATA_RANGES: list[tuple[str, str]] = _chart_data_ranges()

LiteralType = cast(Any, Literal)


def constrain_constant_range(
    constraints: type[Any],
    data: dict[str, Any],
    workbook_path: Path,
    *,
    sheet_name: str,
    range_a1: str,
) -> None:
    """Fill ``constraints`` and ``data`` with Literal types from a rectangular cell range."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        ws = wb[sheet_name]
        for row in ws[range_a1]:
            for cell in row:
                if cell.value is None:
                    continue
                key = f"{sheet_name}!{cell.coordinate}"
                val = cell.value
                constraints.__annotations__[key] = LiteralType[val]
                data[key] = val
    finally:
        wb.close()


# Dated template; adjust filename if using a different snapshot.
WORKBOOK_PATH = Path("example/data/lic-dsf-template-2026-01-31.xlsm")

# Set True to resolve OFFSET/INDIRECT from cached workbook values (no constraints).
# Set False to use constraint-based resolution; add address-style keys below as you hit DynamicRefError.
USE_CACHED_DYNAMIC_REFS = False

# Constraint types for cells that feed OFFSET/INDIRECT. Keys must be address-style (e.g. "Sheet1!B1").
# Add entries when the script raises DynamicRefError: the message lists leaf cells that need
# constraints. Add each to __annotations__ (with Annotated[int, Between(lo, hi)],
# Annotated[..., FromWorkbook()], or Literal[...]) then re-run. Repeat until the graph builds.
class LicDsfConstraints(TypedDict, total=False):
    pass

# Lookup switches; treat as constants
constrain(LicDsfConstraints, "lookup!AF4", Literal["New"])
constrain(LicDsfConstraints, "lookup!AF5", Literal["Old"])

# Marker to use for applicable tailored stress test; we can treat as a constant
constrain(LicDsfConstraints, "'Chart Data'!I21", Literal[1])

# PV_Base!B9xx = CONCAT("$", A9xx, "$", $A$<row>) → INDIRECT($B9xx). Row-index cells A917, A941, A965 (fixed).
# Treat these as constants derived from the current workbook values.
constrain(LicDsfConstraints, "PV_Base!A917", Literal[64])
constrain(LicDsfConstraints, "PV_Base!A941", Literal[90])
constrain(LicDsfConstraints, "PV_Base!A965", Literal[115])

constrain(LicDsfConstraints, "PV_Base!A965", Annotated[float, Between(min=0)])

# A918:A938, A942:A962, A966:A986 each has a single cached letter D, E, …, X.
for _start, _end in [(918, 939), (942, 963), (966, 987)]:
    for _row in range(_start, _end):
        _letter = chr(ord("D") + _row - _start)
        LicDsfConstraints.__annotations__[f"PV_Base!A{_row}"] = LiteralType[_letter]

# Language selector and lookup table (feed INDIRECT/VLOOKUP for language-dependent refs).
# START!L10 = VLOOKUP(K10, lookup!BB4:BC7, 2); evaluator does not support VLOOKUP, so L10 is constrained too.
_LANG = Literal["English", "French", "Portuguese", "Spanish"]
_LANG_LOOKUP = Literal[
    "English", "French", "Portuguese", "Spanish", "Français", "Portugues", "Español"
]
constrain(LicDsfConstraints, "START!L10", _LANG)
constrain(LicDsfConstraints, "START!K10", _LANG)
constrain(LicDsfConstraints, "lookup!BB4:BC7", _LANG_LOOKUP)


def _constrain_lookup_countries(constraints: type[Any]) -> None:
    """Constrain lookup!C4:C73 — LIC-DSF eligible country names.

    Generated by example/codegen_lookup_countries.py.
    """
    _countries: list[tuple[int, str]] = [
        (4, 'Afghanistan'),
        (5, 'Bangladesh'),
        (6, 'Benin'),
        (7, 'Bhutan'),
        (8, 'Burkina Faso'),
        (9, 'Burundi'),
        (10, 'Cambodia'),
        (11, 'Cameroon'),
        (12, 'Cabo Verde'),
        (13, 'Central African Republic'),
        (14, 'Chad'),
        (15, 'Comoros'),
        (16, 'Congo, DR'),
        (17, 'Congo, Republic of'),
        (18, "Cote d'Ivoire"),
        (19, 'Djibouti'),
        (20, 'Dominica'),
        (21, 'Eritrea'),
        (22, 'Ethiopia'),
        (23, 'Gambia, The'),
        (24, 'Ghana'),
        (25, 'Grenada'),
        (26, 'Guinea'),
        (27, 'Guinea-Bissau'),
        (28, 'Guyana'),
        (29, 'Haiti'),
        (30, 'Honduras'),
        (31, 'Kenya'),
        (32, 'Kiribati'),
        (33, 'Kyrgyz Republic'),
        (34, 'Lao PDR'),
        (35, 'Lesotho'),
        (36, 'Liberia'),
        (37, 'Madagascar'),
        (38, 'Malawi'),
        (39, 'Maldives'),
        (40, 'Mali'),
        (41, 'Marshall Islands'),
        (42, 'Mauritania'),
        (43, 'Micronesia'),
        (44, 'Moldova'),
        (45, 'Mozambique'),
        (46, 'Myanmar'),
        (47, 'Nepal'),
        (48, 'Nicaragua'),
        (49, 'Niger'),
        (50, 'Papua New Guinea'),
        (51, 'Rwanda'),
        (52, 'Samoa'),
        (53, 'Sao Tome & Principe'),
        (54, 'Senegal'),
        (55, 'Sierra Leone'),
        (56, 'Solomon Islands'),
        (57, 'Somalia'),
        (58, 'South Sudan'),
        (59, 'St. Lucia'),
        (60, 'St. Vincent & the Grenadines'),
        (61, 'Sudan'),
        (62, 'Tajikistan'),
        (63, 'Tanzania'),
        (64, 'Timor-Leste'),
        (65, 'Togo'),
        (66, 'Tonga'),
        (67, 'Tuvalu'),
        (68, 'Uganda'),
        (69, 'Uzbekistan'),
        (70, 'Vanuatu'),
        (71, 'Yemen, Republic of'),
        (72, 'Zambia'),
        (73, 'Zimbabwe'),
    ]
    for _row, _name in _countries:
        constraints.__annotations__[f"lookup!C{_row}"] = LiteralType[_name]


_constrain_lookup_countries(LicDsfConstraints)


def check_constraints(lic_dsf_constraints: type[TypedDict]) -> None:
    cells_to_constrain = ['C4_Market_financing!C13', 'C4_Market_financing!C19:C4_Market_financing!C32', 'C4_Market_financing!C34', 'C4_Market_financing!C40:C4_Market_financing!C53', 'C4_Market_financing!C4:C4_Market_financing!C7', 'C4_Market_financing!C9:C4_Market_financing!C11', 'C4_Market_financing!D19:C4_Market_financing!D22', 'C4_Market_financing!D29:C4_Market_financing!D32', 'C4_Market_financing!D34', 'C4_Market_financing!D40:C4_Market_financing!D53', 'C4_Market_financing!D4:C4_Market_financing!D6', 'C4_Market_financing!D77', 'C4_Market_financing!D9:C4_Market_financing!D11', 'C4_Market_financing!E19:C4_Market_financing!E34', 'C4_Market_financing!E40:C4_Market_financing!E47', 'C4_Market_financing!E4:C4_Market_financing!E7', 'C4_Market_financing!E9:C4_Market_financing!E11', 'C4_Market_financing!F19:C4_Market_financing!F22', 'C4_Market_financing!F29:C4_Market_financing!F34', 'C4_Market_financing!F40', 'C4_Market_financing!F42:C4_Market_financing!F47', 'C4_Market_financing!F4:C4_Market_financing!F7', 'C4_Market_financing!F9:C4_Market_financing!F11', 'C4_Market_financing!G19:C4_Market_financing!G34', 'C4_Market_financing!G40', 'C4_Market_financing!G42:C4_Market_financing!G47', 'C4_Market_financing!G4:C4_Market_financing!G7', 'C4_Market_financing!G9:C4_Market_financing!G11', 'Chart Data!I21', 'Ext_Debt_Data!AA403:Ext_Debt_Data!AG403', 'Ext_Debt_Data!F383:Ext_Debt_Data!F384', 'Input 1 - Basics!C18', 'Input 1 - Basics!C25', 'Input 1 - Basics!C33', 'Input 3 - Macro-Debt data(DMX)!AB100:Input 3 - Macro-Debt data(DMX)!AQ100', 'Input 3 - Macro-Debt data(DMX)!AB109:Input 3 - Macro-Debt data(DMX)!AQ109', 'Input 3 - Macro-Debt data(DMX)!AB111:Input 3 - Macro-Debt data(DMX)!AQ111', 'Input 3 - Macro-Debt data(DMX)!AB113:Input 3 - Macro-Debt data(DMX)!AQ113', 'Input 3 - Macro-Debt data(DMX)!AB116:Input 3 - Macro-Debt data(DMX)!AQ116', 'Input 3 - Macro-Debt data(DMX)!AB120:Input 3 - Macro-Debt data(DMX)!AQ120', 'Input 3 - Macro-Debt data(DMX)!AB122:Input 3 - Macro-Debt data(DMX)!AQ122', 'Input 3 - Macro-Debt data(DMX)!AB126:Input 3 - Macro-Debt data(DMX)!AQ126', 'Input 3 - Macro-Debt data(DMX)!AB128:Input 3 - Macro-Debt data(DMX)!AQ128', 'Input 3 - Macro-Debt data(DMX)!AB12:Input 3 - Macro-Debt data(DMX)!AQ13', 'Input 3 - Macro-Debt data(DMX)!AB132:Input 3 - Macro-Debt data(DMX)!AQ132', 'Input 3 - Macro-Debt data(DMX)!AB141:Input 3 - Macro-Debt data(DMX)!AQ144', 'Input 3 - Macro-Debt data(DMX)!AB155:Input 3 - Macro-Debt data(DMX)!AQ155', 'Input 3 - Macro-Debt data(DMX)!AB157:Input 3 - Macro-Debt data(DMX)!AQ157', 'Input 3 - Macro-Debt data(DMX)!AB166:Input 3 - Macro-Debt data(DMX)!AQ169', 'Input 3 - Macro-Debt data(DMX)!AB175:Input 3 - Macro-Debt data(DMX)!AQ175', 'Input 3 - Macro-Debt data(DMX)!AB177:Input 3 - Macro-Debt data(DMX)!AQ178', 'Input 3 - Macro-Debt data(DMX)!AB180:Input 3 - Macro-Debt data(DMX)!AQ180', 'Input 3 - Macro-Debt data(DMX)!AB19:Input 3 - Macro-Debt data(DMX)!AQ20', 'Input 3 - Macro-Debt data(DMX)!AB22:Input 3 - Macro-Debt data(DMX)!AQ22', 'Input 3 - Macro-Debt data(DMX)!AB24:Input 3 - Macro-Debt data(DMX)!AQ24', 'Input 3 - Macro-Debt data(DMX)!AB26:Input 3 - Macro-Debt data(DMX)!AQ27', 'Input 3 - Macro-Debt data(DMX)!AB30:Input 3 - Macro-Debt data(DMX)!AQ30', 'Input 3 - Macro-Debt data(DMX)!AB34:Input 3 - Macro-Debt data(DMX)!AQ35', 'Input 3 - Macro-Debt data(DMX)!AB38:Input 3 - Macro-Debt data(DMX)!AQ38', 'Input 3 - Macro-Debt data(DMX)!AB41:Input 3 - Macro-Debt data(DMX)!AQ41', 'Input 3 - Macro-Debt data(DMX)!AB43:Input 3 - Macro-Debt data(DMX)!AQ43', 'Input 3 - Macro-Debt data(DMX)!AB52:Input 3 - Macro-Debt data(DMX)!AQ52', 'Input 3 - Macro-Debt data(DMX)!AB55:Input 3 - Macro-Debt data(DMX)!AQ55', 'Input 3 - Macro-Debt data(DMX)!AB57:Input 3 - Macro-Debt data(DMX)!AQ59', 'Input 3 - Macro-Debt data(DMX)!AB65:Input 3 - Macro-Debt data(DMX)!AQ65', 'Input 3 - Macro-Debt data(DMX)!AB70:Input 3 - Macro-Debt data(DMX)!AQ70', 'Input 3 - Macro-Debt data(DMX)!AB72:Input 3 - Macro-Debt data(DMX)!AQ72', 'Input 3 - Macro-Debt data(DMX)!AB74:Input 3 - Macro-Debt data(DMX)!AQ74', 'Input 3 - Macro-Debt data(DMX)!AB77:Input 3 - Macro-Debt data(DMX)!AQ77', 'Input 3 - Macro-Debt data(DMX)!AB81:Input 3 - Macro-Debt data(DMX)!AQ81', 'Input 3 - Macro-Debt data(DMX)!AB83:Input 3 - Macro-Debt data(DMX)!AQ83', 'Input 3 - Macro-Debt data(DMX)!AB87:Input 3 - Macro-Debt data(DMX)!AQ87', 'Input 3 - Macro-Debt data(DMX)!AB89:Input 3 - Macro-Debt data(DMX)!AQ89', 'Input 3 - Macro-Debt data(DMX)!AB93:Input 3 - Macro-Debt data(DMX)!AQ93', 'Input 3 - Macro-Debt data(DMX)!AB95:Input 3 - Macro-Debt data(DMX)!AQ95', 'Input 3 - Macro-Debt data(DMX)!AR100', 'Input 3 - Macro-Debt data(DMX)!AR109', 'Input 3 - Macro-Debt data(DMX)!AR111', 'Input 3 - Macro-Debt data(DMX)!AR113', 'Input 3 - Macro-Debt data(DMX)!AR116', 'Input 3 - Macro-Debt data(DMX)!AR120', 'Input 3 - Macro-Debt data(DMX)!AR122', 'Input 3 - Macro-Debt data(DMX)!AR126', 'Input 3 - Macro-Debt data(DMX)!AR128', 'Input 3 - Macro-Debt data(DMX)!AR12:Input 3 - Macro-Debt data(DMX)!AR13', 'Input 3 - Macro-Debt data(DMX)!AR132', 'Input 3 - Macro-Debt data(DMX)!AR141:Input 3 - Macro-Debt data(DMX)!AR144', 'Input 3 - Macro-Debt data(DMX)!AR147', 'Input 3 - Macro-Debt data(DMX)!AR155', 'Input 3 - Macro-Debt data(DMX)!AR157', 'Input 3 - Macro-Debt data(DMX)!AR166:Input 3 - Macro-Debt data(DMX)!AR169', 'Input 3 - Macro-Debt data(DMX)!AR175', 'Input 3 - Macro-Debt data(DMX)!AR177:Input 3 - Macro-Debt data(DMX)!AR178', 'Input 3 - Macro-Debt data(DMX)!AR180', 'Input 3 - Macro-Debt data(DMX)!AR19:Input 3 - Macro-Debt data(DMX)!AR20', 'Input 3 - Macro-Debt data(DMX)!AR22', 'Input 3 - Macro-Debt data(DMX)!AR24', 'Input 3 - Macro-Debt data(DMX)!AR26:Input 3 - Macro-Debt data(DMX)!AR27', 'Input 3 - Macro-Debt data(DMX)!AR30', 'Input 3 - Macro-Debt data(DMX)!AR34:Input 3 - Macro-Debt data(DMX)!AR35', 'Input 3 - Macro-Debt data(DMX)!AR38', 'Input 3 - Macro-Debt data(DMX)!AR41', 'Input 3 - Macro-Debt data(DMX)!AR43', 'Input 3 - Macro-Debt data(DMX)!AR52', 'Input 3 - Macro-Debt data(DMX)!AR55', 'Input 3 - Macro-Debt data(DMX)!AR57:Input 3 - Macro-Debt data(DMX)!AR59', 'Input 3 - Macro-Debt data(DMX)!AR65', 'Input 3 - Macro-Debt data(DMX)!AR70', 'Input 3 - Macro-Debt data(DMX)!AR72', 'Input 3 - Macro-Debt data(DMX)!AR74', 'Input 3 - Macro-Debt data(DMX)!AR77', 'Input 3 - Macro-Debt data(DMX)!AR81', 'Input 3 - Macro-Debt data(DMX)!AR83', 'Input 3 - Macro-Debt data(DMX)!AR87', 'Input 3 - Macro-Debt data(DMX)!AR89', 'Input 3 - Macro-Debt data(DMX)!AR93', 'Input 3 - Macro-Debt data(DMX)!AR95', 'Input 3 - Macro-Debt data(DMX)!BP65', 'Input 3 - Macro-Debt data(DMX)!BP70', 'Input 3 - Macro-Debt data(DMX)!BP72', 'Input 3 - Macro-Debt data(DMX)!BP74', 'Input 3 - Macro-Debt data(DMX)!BP77', 'Input 3 - Macro-Debt data(DMX)!BP81', 'Input 3 - Macro-Debt data(DMX)!BP83', 'Input 3 - Macro-Debt data(DMX)!BP87', 'Input 3 - Macro-Debt data(DMX)!BP89', 'Input 3 - Macro-Debt data(DMX)!BP93', 'Input 3 - Macro-Debt data(DMX)!M12:Input 3 - Macro-Debt data(DMX)!M13', 'Input 3 - Macro-Debt data(DMX)!M35', 'Input 3 - Macro-Debt data(DMX)!N12:Input 3 - Macro-Debt data(DMX)!N13', 'Input 3 - Macro-Debt data(DMX)!N142', 'Input 3 - Macro-Debt data(DMX)!N166:Input 3 - Macro-Debt data(DMX)!N167', 'Input 3 - Macro-Debt data(DMX)!N20', 'Input 3 - Macro-Debt data(DMX)!N34:Input 3 - Macro-Debt data(DMX)!N35', 'Input 3 - Macro-Debt data(DMX)!N41', 'Input 3 - Macro-Debt data(DMX)!N43', 'Input 3 - Macro-Debt data(DMX)!N53', 'Input 3 - Macro-Debt data(DMX)!N59', 'Input 3 - Macro-Debt data(DMX)!V12:Input 3 - Macro-Debt data(DMX)!V13', 'Input 3 - Macro-Debt data(DMX)!V20', 'Input 3 - Macro-Debt data(DMX)!V35', 'Input 3 - Macro-Debt data(DMX)!W12:Input 3 - Macro-Debt data(DMX)!W13', 'Input 3 - Macro-Debt data(DMX)!W138:Input 3 - Macro-Debt data(DMX)!W139', 'Input 3 - Macro-Debt data(DMX)!W142', 'Input 3 - Macro-Debt data(DMX)!W161:Input 3 - Macro-Debt data(DMX)!W164', 'Input 3 - Macro-Debt data(DMX)!W166:Input 3 - Macro-Debt data(DMX)!W167', 'Input 3 - Macro-Debt data(DMX)!W19:Input 3 - Macro-Debt data(DMX)!W20', 'Input 3 - Macro-Debt data(DMX)!W34:Input 3 - Macro-Debt data(DMX)!W35', 'Input 3 - Macro-Debt data(DMX)!W41', 'Input 3 - Macro-Debt data(DMX)!W43', 'Input 3 - Macro-Debt data(DMX)!W51:Input 3 - Macro-Debt data(DMX)!W53', 'Input 3 - Macro-Debt data(DMX)!W55', 'Input 3 - Macro-Debt data(DMX)!W57:Input 3 - Macro-Debt data(DMX)!W59', 'Input 3 - Macro-Debt data(DMX)!X100', 'Input 3 - Macro-Debt data(DMX)!X109', 'Input 3 - Macro-Debt data(DMX)!X111', 'Input 3 - Macro-Debt data(DMX)!X113', 'Input 3 - Macro-Debt data(DMX)!X116', 'Input 3 - Macro-Debt data(DMX)!X120', 'Input 3 - Macro-Debt data(DMX)!X122', 'Input 3 - Macro-Debt data(DMX)!X126', 'Input 3 - Macro-Debt data(DMX)!X128', 'Input 3 - Macro-Debt data(DMX)!X12:Input 3 - Macro-Debt data(DMX)!X13', 'Input 3 - Macro-Debt data(DMX)!X132', 'Input 3 - Macro-Debt data(DMX)!X141:Input 3 - Macro-Debt data(DMX)!X144', 'Input 3 - Macro-Debt data(DMX)!X147', 'Input 3 - Macro-Debt data(DMX)!X149:Input 3 - Macro-Debt data(DMX)!X150', 'Input 3 - Macro-Debt data(DMX)!X152', 'Input 3 - Macro-Debt data(DMX)!X154:Input 3 - Macro-Debt data(DMX)!X155', 'Input 3 - Macro-Debt data(DMX)!X157', 'Input 3 - Macro-Debt data(DMX)!X166:Input 3 - Macro-Debt data(DMX)!X169', 'Input 3 - Macro-Debt data(DMX)!X172:Input 3 - Macro-Debt data(DMX)!X173', 'Input 3 - Macro-Debt data(DMX)!X175', 'Input 3 - Macro-Debt data(DMX)!X177:Input 3 - Macro-Debt data(DMX)!X178', 'Input 3 - Macro-Debt data(DMX)!X180', 'Input 3 - Macro-Debt data(DMX)!X19:Input 3 - Macro-Debt data(DMX)!X20', 'Input 3 - Macro-Debt data(DMX)!X22', 'Input 3 - Macro-Debt data(DMX)!X24', 'Input 3 - Macro-Debt data(DMX)!X26:Input 3 - Macro-Debt data(DMX)!X27', 'Input 3 - Macro-Debt data(DMX)!X30', 'Input 3 - Macro-Debt data(DMX)!X35', 'Input 3 - Macro-Debt data(DMX)!X41', 'Input 3 - Macro-Debt data(DMX)!X52', 'Input 3 - Macro-Debt data(DMX)!X55', 'Input 3 - Macro-Debt data(DMX)!X57:Input 3 - Macro-Debt data(DMX)!X58', 'Input 3 - Macro-Debt data(DMX)!X65', 'Input 3 - Macro-Debt data(DMX)!X70', 'Input 3 - Macro-Debt data(DMX)!X72', 'Input 3 - Macro-Debt data(DMX)!X74', 'Input 3 - Macro-Debt data(DMX)!X77', 'Input 3 - Macro-Debt data(DMX)!X81', 'Input 3 - Macro-Debt data(DMX)!X83', 'Input 3 - Macro-Debt data(DMX)!X87', 'Input 3 - Macro-Debt data(DMX)!X89', 'Input 3 - Macro-Debt data(DMX)!X93', 'Input 3 - Macro-Debt data(DMX)!X95', 'Input 3 - Macro-Debt data(DMX)!Y100:Input 3 - Macro-Debt data(DMX)!AA100', 'Input 3 - Macro-Debt data(DMX)!Y109:Input 3 - Macro-Debt data(DMX)!AA109', 'Input 3 - Macro-Debt data(DMX)!Y111:Input 3 - Macro-Debt data(DMX)!AA111', 'Input 3 - Macro-Debt data(DMX)!Y113:Input 3 - Macro-Debt data(DMX)!AA113', 'Input 3 - Macro-Debt data(DMX)!Y116:Input 3 - Macro-Debt data(DMX)!AA116', 'Input 3 - Macro-Debt data(DMX)!Y120:Input 3 - Macro-Debt data(DMX)!AA120', 'Input 3 - Macro-Debt data(DMX)!Y122:Input 3 - Macro-Debt data(DMX)!AA122', 'Input 3 - Macro-Debt data(DMX)!Y126:Input 3 - Macro-Debt data(DMX)!AA126', 'Input 3 - Macro-Debt data(DMX)!Y128:Input 3 - Macro-Debt data(DMX)!AA132', 'Input 3 - Macro-Debt data(DMX)!Y12:Input 3 - Macro-Debt data(DMX)!AA13', 'Input 3 - Macro-Debt data(DMX)!Y141:Input 3 - Macro-Debt data(DMX)!AA144', 'Input 3 - Macro-Debt data(DMX)!Y155:Input 3 - Macro-Debt data(DMX)!AA155', 'Input 3 - Macro-Debt data(DMX)!Y157:Input 3 - Macro-Debt data(DMX)!AA157', 'Input 3 - Macro-Debt data(DMX)!Y166:Input 3 - Macro-Debt data(DMX)!AA169', 'Input 3 - Macro-Debt data(DMX)!Y175:Input 3 - Macro-Debt data(DMX)!AA175', 'Input 3 - Macro-Debt data(DMX)!Y177:Input 3 - Macro-Debt data(DMX)!AA178', 'Input 3 - Macro-Debt data(DMX)!Y180:Input 3 - Macro-Debt data(DMX)!AA180', 'Input 3 - Macro-Debt data(DMX)!Y19:Input 3 - Macro-Debt data(DMX)!AA20', 'Input 3 - Macro-Debt data(DMX)!Y22:Input 3 - Macro-Debt data(DMX)!AA22', 'Input 3 - Macro-Debt data(DMX)!Y24:Input 3 - Macro-Debt data(DMX)!AA24', 'Input 3 - Macro-Debt data(DMX)!Y26:Input 3 - Macro-Debt data(DMX)!AA27', 'Input 3 - Macro-Debt data(DMX)!Y30:Input 3 - Macro-Debt data(DMX)!AA30', 'Input 3 - Macro-Debt data(DMX)!Y34:Input 3 - Macro-Debt data(DMX)!AA35', 'Input 3 - Macro-Debt data(DMX)!Y38:Input 3 - Macro-Debt data(DMX)!AA38', 'Input 3 - Macro-Debt data(DMX)!Y41:Input 3 - Macro-Debt data(DMX)!AA41', 'Input 3 - Macro-Debt data(DMX)!Y43:Input 3 - Macro-Debt data(DMX)!AA43', 'Input 3 - Macro-Debt data(DMX)!Y52:Input 3 - Macro-Debt data(DMX)!AA52', 'Input 3 - Macro-Debt data(DMX)!Y55:Input 3 - Macro-Debt data(DMX)!AA55', 'Input 3 - Macro-Debt data(DMX)!Y57:Input 3 - Macro-Debt data(DMX)!AA59', 'Input 3 - Macro-Debt data(DMX)!Y65:Input 3 - Macro-Debt data(DMX)!AA65', 'Input 3 - Macro-Debt data(DMX)!Y70:Input 3 - Macro-Debt data(DMX)!AA70', 'Input 3 - Macro-Debt data(DMX)!Y72:Input 3 - Macro-Debt data(DMX)!AA72', 'Input 3 - Macro-Debt data(DMX)!Y74:Input 3 - Macro-Debt data(DMX)!AA74', 'Input 3 - Macro-Debt data(DMX)!Y77:Input 3 - Macro-Debt data(DMX)!AA77', 'Input 3 - Macro-Debt data(DMX)!Y81:Input 3 - Macro-Debt data(DMX)!AA81', 'Input 3 - Macro-Debt data(DMX)!Y83:Input 3 - Macro-Debt data(DMX)!AA83', 'Input 3 - Macro-Debt data(DMX)!Y87:Input 3 - Macro-Debt data(DMX)!AA87', 'Input 3 - Macro-Debt data(DMX)!Y89:Input 3 - Macro-Debt data(DMX)!AA89', 'Input 3 - Macro-Debt data(DMX)!Y93:Input 3 - Macro-Debt data(DMX)!AA93', 'Input 3 - Macro-Debt data(DMX)!Y95:Input 3 - Macro-Debt data(DMX)!AA95', 'Input 4 - External Financing!AG10:Input 4 - External Financing!AM10', 'Input 4 - External Financing!AG19:Input 4 - External Financing!AM19', 'Input 4 - External Financing!AG21:Input 4 - External Financing!AM21', 'Input 4 - External Financing!AG23:Input 4 - External Financing!AM23', 'Input 4 - External Financing!AG26:Input 4 - External Financing!AM26', 'Input 4 - External Financing!AG30:Input 4 - External Financing!AM30', 'Input 4 - External Financing!AG32:Input 4 - External Financing!AM32', 'Input 4 - External Financing!AG36:Input 4 - External Financing!AM36', 'Input 4 - External Financing!AG38:Input 4 - External Financing!AM38', 'Input 4 - External Financing!AG42:Input 4 - External Financing!AM42', 'Input 4 - External Financing!F10', 'Input 4 - External Financing!F19', 'Input 4 - External Financing!F21', 'Input 4 - External Financing!F23', 'Input 4 - External Financing!F26', 'Input 4 - External Financing!F30', 'Input 4 - External Financing!F32', 'Input 4 - External Financing!F36', 'Input 4 - External Financing!F38:Input 4 - External Financing!F42', 'Input 4 - External Financing!F45', 'Input 4 - External Financing!G10:Input 4 - External Financing!H10', 'Input 4 - External Financing!G19:Input 4 - External Financing!H19', 'Input 4 - External Financing!G21:Input 4 - External Financing!H21', 'Input 4 - External Financing!G23:Input 4 - External Financing!H23', 'Input 4 - External Financing!G26:Input 4 - External Financing!H26', 'Input 4 - External Financing!G30:Input 4 - External Financing!H30', 'Input 4 - External Financing!G32:Input 4 - External Financing!H32', 'Input 4 - External Financing!G36:Input 4 - External Financing!H36', 'Input 4 - External Financing!G38:Input 4 - External Financing!H42', 'Input 5 - Local-debt Financing!AB250', 'Input 5 - Local-debt Financing!AB274', 'Input 5 - Local-debt Financing!AB298', 'Input 5 - Local-debt Financing!AB322', 'Input 5 - Local-debt Financing!AB488', 'Input 5 - Local-debt Financing!AB512', 'Input 5 - Local-debt Financing!AB581', 'Input 5 - Local-debt Financing!AB63', 'Input 5 - Local-debt Financing!AC63', 'Input 5 - Local-debt Financing!AD108', 'Input 5 - Local-debt Financing!AD110', 'Input 5 - Local-debt Financing!AD188', 'Input 5 - Local-debt Financing!AD191', 'Input 5 - Local-debt Financing!AD193', 'Input 5 - Local-debt Financing!AD93', 'Input 5 - Local-debt Financing!AD95', 'Input 5 - Local-debt Financing!AE108', 'Input 5 - Local-debt Financing!AE110', 'Input 5 - Local-debt Financing!AE250', 'Input 5 - Local-debt Financing!AE254', 'Input 5 - Local-debt Financing!AE274', 'Input 5 - Local-debt Financing!AE278', 'Input 5 - Local-debt Financing!AE298', 'Input 5 - Local-debt Financing!AE302', 'Input 5 - Local-debt Financing!AE322', 'Input 5 - Local-debt Financing!AE392', 'Input 5 - Local-debt Financing!AE461', 'Input 5 - Local-debt Financing!AE93', 'Input 5 - Local-debt Financing!AE95', 'Input 5 - Local-debt Financing!AF108', 'Input 5 - Local-debt Financing!AF110', 'Input 5 - Local-debt Financing!AF250', 'Input 5 - Local-debt Financing!AF274', 'Input 5 - Local-debt Financing!AF298', 'Input 5 - Local-debt Financing!AF322', 'Input 5 - Local-debt Financing!AF392', 'Input 5 - Local-debt Financing!AF461', 'Input 5 - Local-debt Financing!AF488', 'Input 5 - Local-debt Financing!AF512', 'Input 5 - Local-debt Financing!AF93', 'Input 5 - Local-debt Financing!AF95', 'Input 5 - Local-debt Financing!AG108:Input 5 - Local-debt Financing!AJ108', 'Input 5 - Local-debt Financing!AG110:Input 5 - Local-debt Financing!AJ110', 'Input 5 - Local-debt Financing!AG250:Input 5 - Local-debt Financing!AJ250', 'Input 5 - Local-debt Financing!AG254:Input 5 - Local-debt Financing!AJ254', 'Input 5 - Local-debt Financing!AG274:Input 5 - Local-debt Financing!AJ274', 'Input 5 - Local-debt Financing!AG278:Input 5 - Local-debt Financing!AJ278', 'Input 5 - Local-debt Financing!AG298:Input 5 - Local-debt Financing!AJ298', 'Input 5 - Local-debt Financing!AG302:Input 5 - Local-debt Financing!AJ302', 'Input 5 - Local-debt Financing!AG322:Input 5 - Local-debt Financing!AJ322', 'Input 5 - Local-debt Financing!AG392:Input 5 - Local-debt Financing!AJ392', 'Input 5 - Local-debt Financing!AG461:Input 5 - Local-debt Financing!AJ461', 'Input 5 - Local-debt Financing!AG468:Input 5 - Local-debt Financing!AJ468', 'Input 5 - Local-debt Financing!AG488:Input 5 - Local-debt Financing!AJ488', 'Input 5 - Local-debt Financing!AG492:Input 5 - Local-debt Financing!AJ492', 'Input 5 - Local-debt Financing!AG512:Input 5 - Local-debt Financing!AJ512', 'Input 5 - Local-debt Financing!AG93:Input 5 - Local-debt Financing!AJ93', 'Input 5 - Local-debt Financing!AG95:Input 5 - Local-debt Financing!AJ95', 'Input 5 - Local-debt Financing!AK250:Input 5 - Local-debt Financing!AX250', 'Input 5 - Local-debt Financing!AK254:Input 5 - Local-debt Financing!AX254', 'Input 5 - Local-debt Financing!AK274:Input 5 - Local-debt Financing!AX274', 'Input 5 - Local-debt Financing!AK278:Input 5 - Local-debt Financing!AX278', 'Input 5 - Local-debt Financing!AK298:Input 5 - Local-debt Financing!AX298', 'Input 5 - Local-debt Financing!AK302:Input 5 - Local-debt Financing!AX302', 'Input 5 - Local-debt Financing!AK322:Input 5 - Local-debt Financing!AX322', 'Input 5 - Local-debt Financing!AK392:Input 5 - Local-debt Financing!AX392', 'Input 5 - Local-debt Financing!AK461:Input 5 - Local-debt Financing!AX461', 'Input 5 - Local-debt Financing!AK468:Input 5 - Local-debt Financing!AX468', 'Input 5 - Local-debt Financing!AK488:Input 5 - Local-debt Financing!AX488', 'Input 5 - Local-debt Financing!AK492:Input 5 - Local-debt Financing!AX492', 'Input 5 - Local-debt Financing!AK512:Input 5 - Local-debt Financing!AX512', 'Input 5 - Local-debt Financing!AY254', 'Input 5 - Local-debt Financing!AY278', 'Input 5 - Local-debt Financing!AY302', 'Input 5 - Local-debt Financing!AY392', 'Input 5 - Local-debt Financing!AY468', 'Input 5 - Local-debt Financing!AY492', 'Input 5 - Local-debt Financing!BA250', 'Input 5 - Local-debt Financing!BA274', 'Input 5 - Local-debt Financing!BA298', 'Input 5 - Local-debt Financing!BA322', 'Input 5 - Local-debt Financing!BA392', 'Input 5 - Local-debt Financing!BA463', 'Input 5 - Local-debt Financing!BB250:Input 5 - Local-debt Financing!BT250', 'Input 5 - Local-debt Financing!BB274:Input 5 - Local-debt Financing!BT274', 'Input 5 - Local-debt Financing!BB298:Input 5 - Local-debt Financing!BT298', 'Input 5 - Local-debt Financing!BB322:Input 5 - Local-debt Financing!BT322', 'Input 5 - Local-debt Financing!BB392:Input 5 - Local-debt Financing!BT392', 'Input 5 - Local-debt Financing!BB463:Input 5 - Local-debt Financing!BT463', 'Input 5 - Local-debt Financing!BB488:Input 5 - Local-debt Financing!BT488', 'Input 5 - Local-debt Financing!BB512:Input 5 - Local-debt Financing!BT512', 'Input 5 - Local-debt Financing!BU392', 'Input 5 - Local-debt Financing!BU463', 'Input 5 - Local-debt Financing!C10', 'Input 5 - Local-debt Financing!C100:Input 5 - Local-debt Financing!C101', 'Input 5 - Local-debt Financing!C104:Input 5 - Local-debt Financing!C106', 'Input 5 - Local-debt Financing!C108:Input 5 - Local-debt Financing!C110', 'Input 5 - Local-debt Financing!C16', 'Input 5 - Local-debt Financing!C18', 'Input 5 - Local-debt Financing!C20', 'Input 5 - Local-debt Financing!C22', 'Input 5 - Local-debt Financing!C78', 'Input 5 - Local-debt Financing!C83', 'Input 5 - Local-debt Financing!C86', 'Input 5 - Local-debt Financing!C89:Input 5 - Local-debt Financing!C91', 'Input 5 - Local-debt Financing!C93:Input 5 - Local-debt Financing!C95', 'Input 5 - Local-debt Financing!D10', 'Input 5 - Local-debt Financing!D104', 'Input 5 - Local-debt Financing!D106', 'Input 5 - Local-debt Financing!D108', 'Input 5 - Local-debt Financing!D110', 'Input 5 - Local-debt Financing!D16', 'Input 5 - Local-debt Financing!D18', 'Input 5 - Local-debt Financing!D20', 'Input 5 - Local-debt Financing!D22', 'Input 5 - Local-debt Financing!D93', 'Input 5 - Local-debt Financing!D95', 'Input 5 - Local-debt Financing!E104', 'Input 5 - Local-debt Financing!E106', 'Input 5 - Local-debt Financing!E108', 'Input 5 - Local-debt Financing!E110', 'Input 5 - Local-debt Financing!E93', 'Input 5 - Local-debt Financing!E95', 'Input 5 - Local-debt Financing!F104', 'Input 5 - Local-debt Financing!F106', 'Input 5 - Local-debt Financing!F108', 'Input 5 - Local-debt Financing!F110', 'Input 5 - Local-debt Financing!F83', 'Input 5 - Local-debt Financing!F93', 'Input 5 - Local-debt Financing!F95', 'Input 5 - Local-debt Financing!H230', 'Input 5 - Local-debt Financing!H254', 'Input 5 - Local-debt Financing!H278', 'Input 5 - Local-debt Financing!H302', 'Input 5 - Local-debt Financing!H327', 'Input 5 - Local-debt Financing!H397', 'Input 5 - Local-debt Financing!I16', 'Input 5 - Local-debt Financing!I18', 'Input 5 - Local-debt Financing!I20', 'Input 5 - Local-debt Financing!I22', 'Input 5 - Local-debt Financing!I461', 'Input 5 - Local-debt Financing!I488', 'Input 5 - Local-debt Financing!I581', 'Input 5 - Local-debt Financing!I63', 'Input 5 - Local-debt Financing!J488:Input 5 - Local-debt Financing!M488', 'Input 5 - Local-debt Financing!J581:Input 5 - Local-debt Financing!M581', 'Input 5 - Local-debt Financing!J63:Input 5 - Local-debt Financing!M63', 'Input 5 - Local-debt Financing!N16', 'Input 5 - Local-debt Financing!N18', 'Input 5 - Local-debt Financing!N20', 'Input 5 - Local-debt Financing!N22', 'Input 5 - Local-debt Financing!N488', 'Input 5 - Local-debt Financing!N581', 'Input 5 - Local-debt Financing!N63', 'Input 5 - Local-debt Financing!O488:Input 5 - Local-debt Financing!AA488', 'Input 5 - Local-debt Financing!O581:Input 5 - Local-debt Financing!AA581', 'Input 5 - Local-debt Financing!O63:Input 5 - Local-debt Financing!AA63', 'Input 6 - Tailored Tests!C6', 'Input 6(optional)-Standard Test!C17', 'Input 6(optional)-Standard Test!C4:Input 6(optional)-Standard Test!C5', 'Input 6(optional)-Standard Test!C7:Input 6(optional)-Standard Test!C8', 'Input 6(optional)-Standard Test!D18', 'Input 6(optional)-Standard Test!D8:Input 6(optional)-Standard Test!D9', 'Input 8 - SDR!AG37', 'Input 8 - SDR!B6:Input 8 - SDR!B7', 'Input 8 - SDR!C11:Input 8 - SDR!C12', 'Input 8 - SDR!D11:Input 8 - SDR!V12', 'Input 8 - SDR!D14:Input 8 - SDR!V14', 'Input 8 - SDR!W14', 'Input 8 - SDR!X27', 'Input 8 - SDR!Y28', 'PV Stress!D147', 'PV Stress!D161', 'PV Stress!D4', 'PV Stress!E161:PV Stress!G161', 'PV Stress!H147:PV Stress!X147', 'PV Stress!Y148:PV Stress!AF148', 'PV Stress!Y162:PV Stress!AF162', 'PV Stress!Y30:PV Stress!AF30', 'PV_Base!AF23', 'PV_Base!AF272', 'PV_Base!AF298', 'PV_Base!AF350', 'PV_Base!AF376', 'PV_Base!AF480', 'PV_Base!AF506', 'PV_Base!AF610', 'PV_Base!AF636', 'PV_Base!AF740', 'PV_Base!AF766', 'PV_Base!AF818', 'PV_Base!AF844', 'PV_Base!AF896', 'PV_Base!BD366', 'PV_Base!BD470', 'PV_Base!BD496', 'PV_Base!BD600', 'PV_Base!BD626', 'PV_Base!BD730', 'PV_Base!BD756', 'PV_Base!BD808', 'PV_Base!BD834', 'PV_Base!BD886', 'PV_Base!D258', 'PV_Base!D27', 'PV_Base!D276', 'PV_Base!D284', 'PV_Base!D302', 'PV_Base!D336', 'PV_Base!D354', 'PV_Base!D362', 'PV_Base!D380', 'PV_Base!D466', 'PV_Base!D484', 'PV_Base!D49', 'PV_Base!D492', 'PV_Base!D510', 'PV_Base!D596', 'PV_Base!D614', 'PV_Base!D622', 'PV_Base!D640', 'PV_Base!D726', 'PV_Base!D744', 'PV_Base!D752', 'PV_Base!D770', 'PV_Base!D804', 'PV_Base!D822', 'PV_Base!D830', 'PV_Base!D848', 'PV_Base!D882', 'PV_Base!D9', 'PV_Base!D900', 'PV_LC_NR1!AF102', 'PV_LC_NR1!AF121', 'PV_LC_NR1!AF140', 'PV_LC_NR1!AF159', 'PV_LC_NR1!AF178', 'PV_LC_NR1!AF197', 'PV_LC_NR1!AF216', 'PV_LC_NR1!AF235', 'PV_LC_NR1!AF254', 'PV_LC_NR1!AF26', 'PV_LC_NR1!AF273', 'PV_LC_NR1!AF292', 'PV_LC_NR1!AF311', 'PV_LC_NR1!AF330', 'PV_LC_NR1!AF349', 'PV_LC_NR1!AF368', 'PV_LC_NR1!AF387', 'PV_LC_NR1!AF406', 'PV_LC_NR1!AF45', 'PV_LC_NR1!AF64', 'PV_LC_NR1!AF83', 'PV_LC_NR1!BB106', 'PV_LC_NR1!BB125', 'PV_LC_NR1!BB144', 'PV_LC_NR1!BB163', 'PV_LC_NR1!BB182', 'PV_LC_NR1!BB201', 'PV_LC_NR1!BB220', 'PV_LC_NR1!BB239', 'PV_LC_NR1!BB258', 'PV_LC_NR1!BB277', 'PV_LC_NR1!BB296', 'PV_LC_NR1!BB30', 'PV_LC_NR1!BB315', 'PV_LC_NR1!BB334', 'PV_LC_NR1!BB353', 'PV_LC_NR1!BB372', 'PV_LC_NR1!BB391', 'PV_LC_NR1!BB410', 'PV_LC_NR1!BB49', 'PV_LC_NR1!BB68', 'PV_LC_NR1!BB87', 'PV_LC_NR1!BD7', 'PV_LC_NR1!C28', 'PV_LC_NR1!D104', 'PV_LC_NR1!D107', 'PV_LC_NR1!D118', 'PV_LC_NR1!D123', 'PV_LC_NR1!D126', 'PV_LC_NR1!D137', 'PV_LC_NR1!D142', 'PV_LC_NR1!D145', 'PV_LC_NR1!D156', 'PV_LC_NR1!D161', 'PV_LC_NR1!D164', 'PV_LC_NR1!D175', 'PV_LC_NR1!D180', 'PV_LC_NR1!D183', 'PV_LC_NR1!D194', 'PV_LC_NR1!D199', 'PV_LC_NR1!D202', 'PV_LC_NR1!D213', 'PV_LC_NR1!D218', 'PV_LC_NR1!D221', 'PV_LC_NR1!D23', 'PV_LC_NR1!D232', 'PV_LC_NR1!D237', 'PV_LC_NR1!D240', 'PV_LC_NR1!D251', 'PV_LC_NR1!D256', 'PV_LC_NR1!D259', 'PV_LC_NR1!D270', 'PV_LC_NR1!D275', 'PV_LC_NR1!D278', 'PV_LC_NR1!D289', 'PV_LC_NR1!D294', 'PV_LC_NR1!D297', 'PV_LC_NR1!D308', 'PV_LC_NR1!D31', 'PV_LC_NR1!D313', 'PV_LC_NR1!D316', 'PV_LC_NR1!D327', 'PV_LC_NR1!D332', 'PV_LC_NR1!D335', 'PV_LC_NR1!D346', 'PV_LC_NR1!D351', 'PV_LC_NR1!D354', 'PV_LC_NR1!D365', 'PV_LC_NR1!D370', 'PV_LC_NR1!D373', 'PV_LC_NR1!D384', 'PV_LC_NR1!D389', 'PV_LC_NR1!D392', 'PV_LC_NR1!D403', 'PV_LC_NR1!D408', 'PV_LC_NR1!D411', 'PV_LC_NR1!D42', 'PV_LC_NR1!D47', 'PV_LC_NR1!D50', 'PV_LC_NR1!D61', 'PV_LC_NR1!D66', 'PV_LC_NR1!D69', 'PV_LC_NR1!D80', 'PV_LC_NR1!D85', 'PV_LC_NR1!D88', 'PV_LC_NR1!D99', 'PV_LC_NR1!Y6:PV_LC_NR1!AE6', 'PV_LC_NR3!AF102', 'PV_LC_NR3!AF121', 'PV_LC_NR3!AF140', 'PV_LC_NR3!AF159', 'PV_LC_NR3!AF178', 'PV_LC_NR3!AF197', 'PV_LC_NR3!AF216', 'PV_LC_NR3!AF235', 'PV_LC_NR3!AF254', 'PV_LC_NR3!AF26', 'PV_LC_NR3!AF273', 'PV_LC_NR3!AF292', 'PV_LC_NR3!AF311', 'PV_LC_NR3!AF330', 'PV_LC_NR3!AF349', 'PV_LC_NR3!AF368', 'PV_LC_NR3!AF387', 'PV_LC_NR3!AF406', 'PV_LC_NR3!AF45', 'PV_LC_NR3!AF64', 'PV_LC_NR3!AF83', 'PV_LC_NR3!BB106', 'PV_LC_NR3!BB125', 'PV_LC_NR3!BB144', 'PV_LC_NR3!BB163', 'PV_LC_NR3!BB182', 'PV_LC_NR3!BB201', 'PV_LC_NR3!BB220', 'PV_LC_NR3!BB239', 'PV_LC_NR3!BB258', 'PV_LC_NR3!BB277', 'PV_LC_NR3!BB296', 'PV_LC_NR3!BB30', 'PV_LC_NR3!BB315', 'PV_LC_NR3!BB334', 'PV_LC_NR3!BB353', 'PV_LC_NR3!BB372', 'PV_LC_NR3!BB391', 'PV_LC_NR3!BB410', 'PV_LC_NR3!BB49', 'PV_LC_NR3!BB68', 'PV_LC_NR3!BB87', 'PV_LC_NR3!BD7', 'PV_LC_NR3!C28', 'PV_LC_NR3!D104', 'PV_LC_NR3!D107', 'PV_LC_NR3!D118', 'PV_LC_NR3!D123', 'PV_LC_NR3!D126', 'PV_LC_NR3!D137', 'PV_LC_NR3!D142', 'PV_LC_NR3!D145', 'PV_LC_NR3!D156', 'PV_LC_NR3!D161', 'PV_LC_NR3!D164', 'PV_LC_NR3!D175', 'PV_LC_NR3!D180', 'PV_LC_NR3!D183', 'PV_LC_NR3!D194', 'PV_LC_NR3!D199', 'PV_LC_NR3!D202', 'PV_LC_NR3!D213', 'PV_LC_NR3!D218', 'PV_LC_NR3!D221', 'PV_LC_NR3!D23', 'PV_LC_NR3!D232', 'PV_LC_NR3!D237', 'PV_LC_NR3!D240', 'PV_LC_NR3!D251', 'PV_LC_NR3!D256', 'PV_LC_NR3!D259', 'PV_LC_NR3!D270', 'PV_LC_NR3!D275', 'PV_LC_NR3!D278', 'PV_LC_NR3!D289', 'PV_LC_NR3!D294', 'PV_LC_NR3!D297', 'PV_LC_NR3!D308', 'PV_LC_NR3!D31', 'PV_LC_NR3!D313', 'PV_LC_NR3!D316', 'PV_LC_NR3!D327', 'PV_LC_NR3!D332', 'PV_LC_NR3!D335', 'PV_LC_NR3!D346', 'PV_LC_NR3!D351', 'PV_LC_NR3!D354', 'PV_LC_NR3!D365', 'PV_LC_NR3!D370', 'PV_LC_NR3!D373', 'PV_LC_NR3!D384', 'PV_LC_NR3!D389', 'PV_LC_NR3!D392', 'PV_LC_NR3!D403', 'PV_LC_NR3!D408', 'PV_LC_NR3!D411', 'PV_LC_NR3!D42', 'PV_LC_NR3!D47', 'PV_LC_NR3!D50', 'PV_LC_NR3!D61', 'PV_LC_NR3!D66', 'PV_LC_NR3!D69', 'PV_LC_NR3!D80', 'PV_LC_NR3!D85', 'PV_LC_NR3!D88', 'PV_LC_NR3!D99', 'PV_LC_NR3!Y6:PV_LC_NR3!AE6', 'PV_baseline_com!AF111', 'PV_baseline_com!AF137', 'PV_baseline_com!AF33', 'PV_baseline_com!AF59', 'PV_baseline_com!AF85', 'PV_baseline_com!BD101', 'PV_baseline_com!BD127', 'PV_baseline_com!BD23', 'PV_baseline_com!BD49', 'PV_baseline_com!BD75', 'PV_baseline_com!D110', 'PV_baseline_com!D123', 'PV_baseline_com!D136', 'PV_baseline_com!D19', 'PV_baseline_com!D32', 'PV_baseline_com!D45', 'PV_baseline_com!D58', 'PV_baseline_com!D7', 'PV_baseline_com!D71', 'PV_baseline_com!D84', 'PV_baseline_com!D97', 'PV_baseline_com!H110:PV_baseline_com!AE110', 'PV_baseline_com!H136:PV_baseline_com!AE136', 'PV_baseline_com!H32:PV_baseline_com!AE32', 'PV_baseline_com!H58:PV_baseline_com!AE58', 'PV_baseline_com!H84:PV_baseline_com!AE84', 'PV_stress_com!AF115', 'PV_stress_com!AF141', 'PV_stress_com!AF37', 'PV_stress_com!AF63', 'PV_stress_com!AF89', 'PV_stress_com!BD105', 'PV_stress_com!BD131', 'PV_stress_com!BD27', 'PV_stress_com!BD53', 'PV_stress_com!BD79', 'PV_stress_com!D101', 'PV_stress_com!D114', 'PV_stress_com!D127', 'PV_stress_com!D140', 'PV_stress_com!D23', 'PV_stress_com!D36', 'PV_stress_com!D49', 'PV_stress_com!D62', 'PV_stress_com!D75', 'PV_stress_com!D88', 'PV_stress_com!D9', 'PV_stress_com!H114:PV_stress_com!AE114', 'PV_stress_com!H140:PV_stress_com!AE140', 'PV_stress_com!H36:PV_stress_com!AE36', 'PV_stress_com!H62:PV_stress_com!AE62', 'PV_stress_com!H88:PV_stress_com!AE88', 'lookup!AF4', 'translation!C451:translation!C452', 'translation!C898', 'translation!C90', 'translation!D451:translation!F452', 'translation!D898:translation!F898']
    missing = [c for c in cells_to_constrain if c not in lic_dsf_constraints.__annotations__]
    if missing:
        raise ValueError(f"Missing constraints for: {missing}")


def parse_range_spec(spec: str) -> tuple[str, str]:
    """
    Parse a sheet-qualified range spec into (sheet_name, range_a1).

    Accepts specs like "'Chart Data'!D10:D17" or "Sheet1!A1:B2".
    """
    if "!" not in spec:
        raise ValueError(f"Range spec must contain '!': {spec!r}")
    sheet_part, range_part = spec.split("!", 1)
    sheet_part = sheet_part.strip()
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1].replace("''", "'")
    return sheet_part, range_part.strip()


def cells_in_range(sheet: str, range_a1: str) -> list[str]:
    """
    Expand an A1 range to a list of sheet-qualified cell keys.

    range_a1 may be a single cell ("D10") or a range ("D10:D17", "D239:X252").
    """
    if ":" in range_a1:
        start_a1, end_a1 = range_a1.split(":", 1)
        start_a1 = start_a1.strip()
        end_a1 = end_a1.strip()
    else:
        start_a1 = end_a1 = range_a1.strip()

    c1, r1 = openpyxl.utils.cell.coordinate_from_string(start_a1)
    c2, r2 = openpyxl.utils.cell.coordinate_from_string(end_a1)
    start_col_idx = openpyxl.utils.cell.column_index_from_string(c1)
    end_col_idx = openpyxl.utils.cell.column_index_from_string(c2)
    rlo, rhi = (r1, r2) if r1 <= r2 else (r2, r1)
    clo, chi = (start_col_idx, end_col_idx) if start_col_idx <= end_col_idx else (end_col_idx, start_col_idx)

    out: list[str] = []
    for row in range(rlo, rhi + 1):
        for col_idx in range(clo, chi + 1):
            col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
            out.append(format_cell_key(sheet, col_letter, row))
    return out


def main() -> None:
    print("=" * 70)
    print("LIC-DSF Indicator Dependency Mapping")
    print("=" * 70)
    
    if not WORKBOOK_PATH.exists():
        print(f"Error: Workbook not found at {WORKBOOK_PATH}")
        return
    
    # Check constraints
    check_constraints(LicDsfConstraints)

    # Discover targets: explicit ranges (all cells) and indicator rows (formula cells only)
    print("\n1. Collecting target cells...")
    all_targets: list[str] = []

    for label, spec in CHART_DATA_RANGES:
        sheet_name, range_a1 = parse_range_spec(spec)
        targets = cells_in_range(sheet_name, range_a1)
        print(f"   {label}: {spec} -> {len(targets)} cells")
        all_targets.extend(targets)

    print(f"\n   Total targets: {len(all_targets)}")
    
    if not all_targets:
        print("No formula cells found. Exiting.")
        return
    
    # Build dependency graph (constraint-based or cached for OFFSET/INDIRECT)
    print("\n2. Building dependency graph...")
    t_build = time.perf_counter()
    dynamic_refs: DynamicRefConfig | None = None
    if not USE_CACHED_DYNAMIC_REFS:
        dynamic_refs = DynamicRefConfig.from_constraints_and_workbook(
            LicDsfConstraints,
            WORKBOOK_PATH,
        )
    try:
        graph = create_dependency_graph(
            WORKBOOK_PATH,
            all_targets,
            load_values=False,
            max_depth=50,
            dynamic_refs=dynamic_refs,
            use_cached_dynamic_refs=USE_CACHED_DYNAMIC_REFS,
        )
    except DynamicRefError as e:
        print(f"\n   DynamicRefError: {e}")
        print(
            "   Add the reported cell's argument cells to LicDsfConstraints (address-style keys)"
            " using Annotated[..., Between(...)] / Annotated[..., FromWorkbook()] as needed,"
            " then re-run. Or set USE_CACHED_DYNAMIC_REFS=True to resolve from cached values."
        )
        raise

    build_s = time.perf_counter() - t_build
    print(f"   Graph build time: {build_s:.2f}s")

    print(f"   Nodes in graph: {len(graph)}")
    print(f"   Leaf nodes: {sum(1 for _ in graph.leaves())}")
    print(f"   Formula nodes: {len(graph) - sum(1 for _ in graph.leaves())}")
    
    # Group nodes by sheet
    sheets: dict[str, int] = {}
    for key in graph:
        node = graph.get_node(key)
        if node:
            sheets[node.sheet] = sheets.get(node.sheet, 0) + 1
    
    print("\n   Nodes by sheet:")
    for sheet_name in sorted(sheets.keys()):
        print(f"      {sheet_name}: {sheets[sheet_name]}")

    # Workbook calc settings (useful context for interpreting cycles)
    print("\n3. Workbook calculation settings...")
    settings = get_calc_settings(WORKBOOK_PATH)
    print(f"   Iterate enabled: {settings.iterate_enabled}")
    print(f"   Iterate count:   {settings.iterate_count}")
    print(f"   Iterate delta:   {settings.iterate_delta}")

    # Cycle analysis (must-cycle vs may-cycle)
    print("\n4. Cycle analysis...")
    report = graph.cycle_report()
    print(f"   Must-cycles: {len(report.must_cycles)}")
    print(f"   May-cycles:  {len(report.may_cycles)}")
    if report.example_must_cycle_path:
        print(
            f"   Example must-cycle path: {' -> '.join(report.example_must_cycle_path)}"
        )
    if report.example_may_cycle_path:
        print(
            f"   Example may-cycle path:  {' -> '.join(report.example_may_cycle_path)}"
        )
    
    # Validate against calcChain.xml
    print("\n5. Validating against calcChain.xml...")
    scope = {parse_range_spec(spec)[0] for _label, spec in CHART_DATA_RANGES}
    result = validate_graph(graph, WORKBOOK_PATH, scope=scope)
    
    print(f"   Valid: {result.is_valid}")
    for msg in result.messages:
        print(f"   {msg}")
    
    if result.in_graph_not_in_chain:
        print(
            f"\n   Cells in graph but not in calcChain ({len(result.in_graph_not_in_chain)}):"
        )
        for cell in sorted(result.in_graph_not_in_chain)[:10]:
            print(f"      {cell}")
        if len(result.in_graph_not_in_chain) > 10:
            print(f"      ... and {len(result.in_graph_not_in_chain) - 10} more")
    
    # Evaluation order stats
    print("\n6. Computing evaluation order...")
    try:
        # Non-strict mode will warn and exclude nodes involved in may-cycles, but
        # still fails on must-cycles.
        order = graph.evaluation_order(strict=False)
        print(f"   Evaluation order computed: {len(order)} nodes")
        print(f"   First 5 (leaves): {order[:5]}")
        print(f"   Last 5 (targets): {order[-5:]}")
    except CycleError as e:
        kind = "must-cycle" if e.is_must_cycle else "may-cycle"
        print(f"   Error ({kind}): {e}")
        if e.cycle_path:
            print(f"   Cycle path: {' -> '.join(e.cycle_path)}")
    
    # Optional: save a small subgraph visualization
    print("\n7. Sample visualization (first target's immediate deps)...")
    if all_targets:
        sample_target = all_targets[0]
        sample_deps = graph.dependencies(sample_target)
        print(f"   {sample_target} depends on {len(sample_deps)} cells:")
        for dep in sorted(sample_deps)[:5]:
            guard = graph.edge_attrs(sample_target, dep).get("guard")
            if guard is None:
                print(f"      {dep}")
            else:
                print(f"      {dep}  [guarded: {guard}]")
        if len(sample_deps) > 5:
            print(f"      ... and {len(sample_deps) - 5} more")

        # Emit a DOT snippet for quick inspection (guarded edges render dashed + labeled).
        try:
            dot = to_graphviz(graph, highlight={sample_target}, rankdir="LR")
            print("\n   GraphViz DOT (truncated to first ~40 lines):")
            for line in dot.splitlines()[:40]:
                print(f"      {line}")
            if len(dot.splitlines()) > 40:
                print("      ...")
        except Exception as e:
            print(f"   Could not render GraphViz DOT: {e}")
    
    print("\n" + "=" * 70)
    print("Done.")


if __name__ == "__main__":
    main()
